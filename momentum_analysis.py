"""Sales-momentum + ABC focus report (`--momentum`).

Two lenses to decide what to push vs prune:

  Momentum — recent window (`MOMENTUM_WINDOW_DAYS`) qty vs the prior window:
    🚀 Akselerasi / 📉 Menurun (±`MOMENTUM_GROWTH_THRESHOLD`) / ➡️ Stabil /
    🆕 Baru naik (prior 0) / 💤 Berhenti (recent 0). Needs ≥ `MOMENTUM_MIN_QTY` total.
  ABC — Pareto by trailing profit (`MOMENTUM_TRAILING_DAYS`): cumulative share ≤
    `ABC_A_SHARE` = A, ≤ `ABC_B_SHARE` = B, else C. profit = omzet + admin − HPP_WA×qty.

The recommendation combines them (e.g. A-class accelerating → push; A-class declining
→ protect; C-class declining → prune). Output: output/Analisa_Momentum_ABC.xlsx.
Zero-config (built from BisaJual + HPP).
"""
from __future__ import annotations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    ABC_A_SHARE, ABC_B_SHARE, BLUE_FILL_COLOR, FMT_NUM, FMT_RP, FONT_NAME,
    GREEN_FILL_COLOR, HEADER_BG_COLOR, HEADER_TEXT_COLOR, LIGHT_GRAY_COLOR,
    MOMENTUM_GROWTH_THRESHOLD, MOMENTUM_MIN_QTY, MOMENTUM_TRAILING_DAYS,
    MOMENTUM_WINDOW_DAYS, RED_FILL_COLOR, TITLE_COLOR, YELLOW_FILL_COLOR,
)

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
GREEN_FILL = PatternFill("solid", start_color=GREEN_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
RED_FILL = PatternFill("solid", start_color=RED_FILL_COLOR)
BLUE_FILL = PatternFill("solid", start_color=BLUE_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)
FMT_PCT1 = '0.0%'
FMT_GROWTH = '+0%;-0%'


def _recommend(abc: str, mom: str) -> tuple[str, str]:
    if abc == "A" and "🚀" in mom:
        return "⭐ Dorong", "Juara yang masih naik — amankan stok & iklan."
    if abc == "A" and "📉" in mom:
        return "⚠ Lindungi", "Juara turun — cek harga/stok/kompetitor segera."
    if abc == "A" and "💤" in mom:
        return "🚨 Investigasi", "Juara berhenti jual — stockout/delisting? cek sekarang."
    if abc == "B" and "🚀" in mom:
        return "📈 Naikkan", "Kandidat juara berikutnya — beri dorongan."
    if abc == "C" and ("📉" in mom or "💤" in mom):
        return "✂ Pangkas", "Kontribusi kecil & menurun — stop reorder / clearance."
    if "🆕" in mom:
        return "👀 Pantau", "Pendatang baru — pantau apakah lanjut naik."
    return "—", ""


def analyze_momentum(jual: pd.DataFrame, hpp_agg: pd.DataFrame, today) -> pd.DataFrame:
    """One row per SKU with recent activity: momentum tag, ABC class, trailing profit
    + share, and a push/prune recommendation."""
    needed = {"SKU", "tanggal_pesan", "qty_jual", "omzet"}
    if jual is None or len(jual) == 0 or not needed.issubset(jual.columns):
        return pd.DataFrame()

    hpp_wa = hpp_agg.set_index("SKU")["hpp_wa"].to_dict() if len(hpp_agg) else {}
    j = jual[(jual["qty_jual"] > 0) & jual["omzet"].notna()].copy()
    j["_admin"] = j.get("tambahan", 0).fillna(0) + j.get("kode_unik", 0).fillna(0)

    cut_last = today - pd.Timedelta(days=MOMENTUM_WINDOW_DAYS)
    cut_prior = today - pd.Timedelta(days=2 * MOMENTUM_WINDOW_DAYS)
    cut_trail = today - pd.Timedelta(days=MOMENTUM_TRAILING_DAYS)

    def agg(d):
        return d.groupby("SKU").agg(qty=("qty_jual", "sum"), omzet=("omzet", "sum"),
                                    admin=("_admin", "sum"))

    last = agg(j[j["tanggal_pesan"] >= cut_last])
    prior = agg(j[(j["tanggal_pesan"] < cut_last) & (j["tanggal_pesan"] >= cut_prior)])
    trail = agg(j[j["tanggal_pesan"] >= cut_trail])
    skus = trail.index.union(last.index).union(prior.index)

    rows = []
    for sku in skus:
        lq = float(last["qty"].get(sku, 0.0))
        pq = float(prior["qty"].get(sku, 0.0))
        tq = float(trail["qty"].get(sku, 0.0))
        tom = float(trail["omzet"].get(sku, 0.0))
        tad = float(trail["admin"].get(sku, 0.0))
        hpp = hpp_wa.get(sku, np.nan)
        profit = (tom + tad - hpp * tq) if pd.notna(hpp) else np.nan

        vol = lq + pq
        if vol < MOMENTUM_MIN_QTY:
            mom, growth = "• Data tipis", np.nan
        elif pq == 0 and lq > 0:
            mom, growth = "🆕 Baru naik", np.nan
        elif lq == 0 and pq > 0:
            mom, growth = "💤 Berhenti", -1.0
        else:
            growth = lq / pq - 1.0
            mom = ("🚀 Akselerasi" if growth >= MOMENTUM_GROWTH_THRESHOLD
                   else "📉 Menurun" if growth <= -MOMENTUM_GROWTH_THRESHOLD else "➡️ Stabil")
        rows.append({"SKU": sku, "last_qty": lq, "prior_qty": pq, "growth": growth,
                     "momentum": mom, "trail_qty": tq, "trail_profit": profit})

    df = pd.DataFrame(rows)
    if len(df) == 0:
        print("✓ Momentum/ABC: tidak ada penjualan dalam window")
        return df

    # ABC by trailing profit (cumulative over positive-profit total)
    df["_p"] = df["trail_profit"].fillna(0.0)
    df = df.sort_values("_p", ascending=False).reset_index(drop=True)
    pos_total = float(df.loc[df["_p"] > 0, "_p"].sum())
    cum, classes, cum_shares = 0.0, [], []
    for v in df["_p"]:
        if v > 0 and pos_total > 0:
            cum += v
            share = cum / pos_total
            classes.append("A" if share <= ABC_A_SHARE else "B" if share <= ABC_B_SHARE else "C")
            cum_shares.append(share)
        else:
            classes.append("C")
            cum_shares.append(np.nan)
    df["abc"] = classes
    df["cum_share"] = cum_shares
    # share of the positive-profit total; NaN for loss-making / HPP-less SKUs (matches cum_share)
    if pos_total > 0:
        df["profit_share"] = np.where(df["_p"] > 0, df["_p"] / pos_total, np.nan)
    else:
        df["profit_share"] = np.nan
    df[["rec", "saran"]] = df.apply(lambda r: pd.Series(_recommend(r["abc"], r["momentum"])), axis=1)
    df = df.drop(columns="_p")

    n_a = int((df["abc"] == "A").sum())
    n_acc = int(df["momentum"].str.contains("Akselerasi").sum())
    n_dec = int(df["momentum"].str.contains("Menurun").sum())
    print(f"✓ Momentum/ABC: {len(df):,} SKU — kelas A={n_a}, akselerasi={n_acc}, menurun={n_dec}")
    return df


# ---------------------------------------------------------------------------
def _style_header(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 28


def _growth_cell(ws, row, col, growth):
    if pd.isna(growth):
        ws.cell(row=row, column=col, value="—").font = NORMAL_FONT
    else:
        c = ws.cell(row=row, column=col, value=float(growth))
        c.font = NORMAL_FONT
        c.number_format = FMT_GROWTH


def write_momentum_report(filepath: Path, df: pd.DataFrame, today) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Ringkasan"
    ws["A1"] = "MOMENTUM & ABC — APA YANG DIDORONG vs DIPANGKAS"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:F1")
    win = MOMENTUM_WINDOW_DAYS
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  Momentum = qty {win} hari terakhir vs {win} hari "
                f"sebelumnya.  ABC = Pareto profit {MOMENTUM_TRAILING_DAYS//30} bln terakhir "
                f"(A ≤ {ABC_A_SHARE*100:.0f}% kumulatif, B ≤ {ABC_B_SHARE*100:.0f}%, sisanya C).")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:F2")

    r = 4
    if len(df):
        n_by_class = df["abc"].value_counts().to_dict()
        a_profit = float(df.loc[df["abc"] == "A", "trail_profit"].sum(skipna=True))
        tot_profit = float(df.loc[df["trail_profit"] > 0, "trail_profit"].sum(skipna=True))
        a_share = (a_profit / tot_profit * 100) if tot_profit > 0 else 0
        lines = [
            ("SKU dengan penjualan (window)", len(df), FMT_NUM),
            (f"Kelas A ({a_share:.0f}% dari profit)", n_by_class.get("A", 0), FMT_NUM),
            ("Kelas B", n_by_class.get("B", 0), FMT_NUM),
            ("Kelas C", n_by_class.get("C", 0), FMT_NUM),
            ("🚀 Akselerasi", int(df["momentum"].str.contains("Akselerasi").sum()), FMT_NUM),
            ("📉 Menurun", int(df["momentum"].str.contains("Menurun").sum()), FMT_NUM),
        ]
        for label, val, fmt in lines:
            ws.cell(row=r, column=1, value=label).font = BOLD_FONT
            c = ws.cell(row=r, column=2, value=val)
            c.font = BOLD_FONT
            c.number_format = fmt
            r += 1

        # Alerts: A-class declining/stopped (protect)
        r += 1
        ws.cell(row=r, column=1, value="⚠ Juara (A) yang turun — prioritas lindungi").font = TITLE_FONT
        r += 1
        _style_header(ws, r, ["SKU", "Momentum", "Growth", "Profit Share", "Aksi"],
                      [34, 14, 10, 13, 16])
        r += 1
        alerts = df[(df["abc"] == "A") & (df["momentum"].str.contains("Menurun|Berhenti"))].head(12)
        for _, p in alerts.iterrows():
            ws.cell(row=r, column=1, value=p["SKU"]).font = NORMAL_FONT
            ws.cell(row=r, column=2, value=p["momentum"]).font = NORMAL_FONT
            _growth_cell(ws, r, 3, p["growth"])
            c = ws.cell(row=r, column=4, value=(float(p["profit_share"]) if pd.notna(p["profit_share"]) else None))
            c.font = NORMAL_FONT
            c.number_format = FMT_PCT1
            ws.cell(row=r, column=5, value=p["rec"]).font = NORMAL_FONT
            ws.cell(row=r, column=1).fill = YELLOW_FILL
            r += 1
        if len(alerts) == 0:
            ws.cell(row=r, column=1, value="(tidak ada juara yang menurun — bagus!)").font = NORMAL_FONT

    # --- Sheet 01: focus per SKU ---
    ws2 = wb.create_sheet("01_Fokus_SKU")
    ws2["A1"] = "FOKUS PER SKU — MOMENTUM × ABC"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:J1")
    headers = ["SKU", "ABC", "Momentum", "Qty (recent)", "Qty (prior)", "Growth",
               "Profit (trailing)", "Profit Share", "Rekomendasi", "Saran"]
    widths = [34, 6, 14, 12, 12, 9, 16, 12, 16, 46]
    _style_header(ws2, 3, headers, widths)
    if len(df):
        rr = 4
        for _, p in df.iterrows():
            ws2.cell(row=rr, column=1, value=p["SKU"]).font = NORMAL_FONT
            ws2.cell(row=rr, column=2, value=p["abc"]).font = BOLD_FONT
            ws2.cell(row=rr, column=3, value=p["momentum"]).font = NORMAL_FONT
            c = ws2.cell(row=rr, column=4, value=round(p["last_qty"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=5, value=round(p["prior_qty"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            _growth_cell(ws2, rr, 6, p["growth"])
            c = ws2.cell(row=rr, column=7, value=(round(p["trail_profit"]) if pd.notna(p["trail_profit"]) else None))
            c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=8, value=(float(p["profit_share"]) if pd.notna(p["profit_share"]) else None))
            c.number_format = FMT_PCT1; c.font = NORMAL_FONT
            ws2.cell(row=rr, column=9, value=p["rec"]).font = NORMAL_FONT
            c = ws2.cell(row=rr, column=10, value=p["saran"]); c.font = NORMAL_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if str(p["rec"]).startswith("⭐"):
                ws2.cell(row=rr, column=9).fill = GREEN_FILL
            elif str(p["rec"]).startswith(("⚠", "🚨")):
                ws2.cell(row=rr, column=9).fill = YELLOW_FILL
            elif str(p["rec"]).startswith("✂"):
                ws2.cell(row=rr, column=9).fill = RED_FILL
            rr += 1
        ws2.freeze_panes = "A4"

    # --- Sheet 02: ABC Pareto ---
    ws3 = wb.create_sheet("02_ABC_Pareto")
    ws3["A1"] = "PARETO ABC — KONSENTRASI PROFIT"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:E1")
    _style_header(ws3, 3, ["SKU", "Kelas", "Profit (trailing)", "Profit Share", "Kumulatif"],
                  [34, 7, 16, 13, 12])
    if len(df):
        rr = 4
        for _, p in df.iterrows():
            ws3.cell(row=rr, column=1, value=p["SKU"]).font = NORMAL_FONT
            ws3.cell(row=rr, column=2, value=p["abc"]).font = BOLD_FONT
            c = ws3.cell(row=rr, column=3, value=(round(p["trail_profit"]) if pd.notna(p["trail_profit"]) else None))
            c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=4, value=(float(p["profit_share"]) if pd.notna(p["profit_share"]) else None))
            c.number_format = FMT_PCT1; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=5, value=(float(p["cum_share"]) if pd.notna(p["cum_share"]) else None))
            c.number_format = FMT_PCT1; c.font = NORMAL_FONT
            if rr % 2 == 1:
                for ci in range(1, 6):
                    ws3.cell(row=rr, column=ci).fill = LIGHT_FILL
            rr += 1
        ws3.freeze_panes = "A4"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
