"""Reconcile a list of invoices against the Jual ledger to surface entry bugs.

This is a read-only audit aimed at one question: **how was each order entered into
Jual, compared to the real money it actually made?** For every invoice it pairs

  * the booked Omzet in the itbisa-shop-report-bot ``*Jual*.xlsx`` ledger
    (non-void ``Omzet Barang (Rp)`` across every ``Jual*`` sheet, with the void
    amount kept separately), with
  * the real money received in ``Saldo`` (the full net of every line for the
    invoice) and the ``Fee`` settlement (``Total Penghasilan`` + the refund/fee
    ``Kerugian``),

then classifies the Jual entry:

  BUG: entry hilang          - money came in, but the order is not in Jual
  BUG: harusnya Void         - Omzet booked, but no money (a return left un-voided)
  BUG: salah Void            - voided in Jual, yet money was received
  BUG: omzet != uang         - booked Omzet doesn't match the money received
  OK : cocok / retur / void  - entry already matches reality

Run it with ``python main.py --cek-jual`` (defaults to the built-in list) or
``--cek-jual --invoices my_list.txt`` for any other set, and point
``--jual-dir`` at the bot repo's ``data/``. It writes
``reports/shopee/Cek Jual Shopee.xlsx``.

The invoice list is general: pass a text file (one invoice per line; blank lines and
``#`` comments ignored) to reconcile any set, growing over time.
"""
import glob
import logging
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from laporan.utility.constant import MARKETPLACE_FOLDERS, get_data_dir, get_reports_dir
from laporan.rekonsiliasi.generic import _saldo_files

# Bot Jual ledger column (hand-kept, with a Void flag).
_OMZET = 'Omzet\nBarang\n(Rp)'

# Default invoices to reconcile when no --invoices file is given (extend freely).
DEFAULT_INVOICES = [
    '19080115463288A', '231215A1CX1JNX', '240424JKCDGQEA', '2406057PX7RQ25',
    '2406165YTVPSPF', '24062753NRG1MS', '2406286REGDT4D', '2407096KD753VK',
    '240715N4B89FST', '24080289A8QSPF', '240920F2YS822V', '2410203F0PRAWJ',
    '241115A4659PXA', '241202RJBYJB4P', '2502037PCXKMC6', '250318VKVPVDC0',
    '2504028KT6FM9X', '250918T665Y4PA', '251019G1W6AUMJ', '2510240B6X9K0G',
    '251204EYP3GQ9K', '251205HUS0Q7HN', '251206MY7Y3VHD', '251226D4K0460N',
    '251228HNJ059UG', '260108G7KJTBX7', '260112TH45XY6G', '260125199JXECB',
    '260206074G1DFB', '260403TB2NNS7T', '230408H1TYXK9V', '230606MYVAEE7E',
    '230703230YXGSA', '231011MWAPBMNE', '240629AK24D2Q8', '251103RTPPVFKD',
]

# How close the received money must be to Omzet to count as "cocok" (fees aside).
_MATCH_TOLERANCE_ABS = 2000
_MATCH_TOLERANCE_PCT = 0.25


def load_invoices(path):
    """Read invoices from a text file (one per line; blanks and # comments ignored)."""
    invoices = []
    with open(path, encoding='utf-8') as fh:
        for line in fh:
            inv = line.strip()
            if inv and not inv.startswith('#'):
                invoices.append(inv)
    return invoices


def _jual_detail(invoices, jual_dir=None):
    """Invoice -> {present, omzet (non-void), void_omzet} from the bot Jual files.

    Looks across every ``Jual*`` sheet (so a Shopee order booked under e.g.
    ``JualCoD`` still counts as present).
    """
    inv_set = set(invoices)
    files = []
    for folder in [get_data_dir()] + ([jual_dir] if jual_dir else []):
        files += glob.glob(os.path.join(folder, '**', '*Jual*.xls*'), recursive=True)

    detail = {}
    for path in sorted(set(files)):
        if os.path.basename(path).startswith('~'):
            continue
        try:
            xls = pd.ExcelFile(path)
        except Exception as err:  # noqa: BLE001
            logging.warning("Gagal baca Jual %s: %s", path, err)
            continue
        for sheet in (s for s in xls.sheet_names if s.startswith('Jual')):
            df = xls.parse(sheet)
            if 'Invoice' not in df.columns or _OMZET not in df.columns:
                continue
            df = df[df['Invoice'].astype(str).isin(inv_set)]
            if df.empty:
                continue
            df = df.assign(_om=pd.to_numeric(df[_OMZET], errors='coerce').fillna(0))
            is_void = df['Void'] == True if 'Void' in df.columns else pd.Series(False, index=df.index)  # noqa: E712
            for inv, sub in df.groupby(df['Invoice'].astype(str)):
                rec = detail.setdefault(inv, {'present': True, 'omzet': 0.0, 'void_omzet': 0.0})
                voided = is_void.loc[sub.index]
                rec['omzet'] += sub.loc[~voided, '_om'].sum()
                rec['void_omzet'] += sub.loc[voided, '_om'].sum()
                rec['sheet'] = sheet
    return detail


def _saldo_net(invoices):
    """Invoice -> full Saldo net (every Shopee saldo line for that invoice)."""
    inv_set = set(invoices)
    net = {}

    def add(series):
        for inv, val in series.items():
            if inv in inv_set:
                net[inv] = net.get(inv, 0.0) + val

    for path in _saldo_files('Saldo v2 Shopee'):
        df = pd.read_csv(path, skiprows=6)
        inv = df['Deskripsi'].astype(str).str.extract(r'#(\S+)')[0].str.replace('.', '', regex=False)
        amt = pd.to_numeric(df['Jumlah Dana'], errors='coerce').fillna(0)
        add(pd.DataFrame({'inv': inv, 'a': amt}).dropna(subset=['inv']).groupby('inv')['a'].sum())
    for path in _saldo_files('Saldo v3 Shopee'):
        df = pd.read_excel(path, skiprows=17)
        inv = df['No. Pesanan'].astype(str)
        amt = pd.to_numeric(df['Jumlah'], errors='coerce').fillna(0)
        add(pd.DataFrame({'inv': inv, 'a': amt}).groupby('inv')['a'].sum())
    return net


def _fee_detail(invoices):
    """Invoice -> {tp (Total Penghasilan), ker (refund+fees as Kerugian)} from Fee."""
    inv_set = set(invoices)
    rows = []
    for token in ('Fee v2 Shopee', 'Fee v3 Shopee'):
        for path in _saldo_files(token):
            try:
                df = pd.read_excel(path, sheet_name='Income', skiprows=5)
            except Exception as err:  # noqa: BLE001
                logging.debug("Skip fee file %s: %s", path, err)
                continue
            if 'No. Pesanan' not in df.columns:
                continue
            df = df[df['No. Pesanan'].astype(str).isin(inv_set)]
            if df.empty:
                continue

            def col(name):
                return pd.to_numeric(df[name], errors='coerce').fillna(0) if name in df.columns else 0
            rows.append(pd.DataFrame({
                'Invoice': df['No. Pesanan'].astype(str),
                'tp': col('Total Penghasilan'),
                'ker': -(col('Jumlah Pengembalian Dana ke Pembeli') + col('Biaya Transaksi')
                         + col('Biaya Administrasi') + col('Biaya Layanan (termasuk PPN 11%)'))}))
    if not rows:
        return {}
    # Overlapping fee files (yearly v2 + monthly v3) can list the same order; count once.
    allrows = pd.concat(rows, ignore_index=True).drop_duplicates(subset=['Invoice', 'tp', 'ker'])
    g = allrows.groupby('Invoice').agg(tp=('tp', 'sum'), ker=('ker', 'sum'))
    return {inv: {'tp': r['tp'], 'ker': r['ker']} for inv, r in g.iterrows()}


def _verdict(jual, real, fee):
    """Classify how the Jual entry compares to the real money."""
    present = jual is not None
    omzet = jual['omzet'] if present else 0.0
    void_omzet = jual['void_omzet'] if present else 0.0
    money = real if real is not None else 0.0
    fee_tp = fee['tp'] if fee else 0.0
    has_money = money > 0 or fee_tp > 0

    if not present:
        return 'BUG: entry hilang (ada uang, tidak di Jual)' if has_money \
            else 'OK: retur, tidak dibukukan'
    if omzet > 0 and money <= 0:
        return 'BUG: omzet di Jual tapi tidak ada uang -> Void'
    if omzet <= 0 and void_omzet > 0 and money > 0:
        return 'BUG: di-Void tapi ada uang masuk'
    if omzet > 0 and money > 0:
        if abs(money - omzet) <= max(_MATCH_TOLERANCE_ABS, omzet * _MATCH_TOLERANCE_PCT):
            return 'OK: cocok (omzet = uang)'
        return 'BUG: omzet != uang diterima'
    if omzet <= 0 and void_omzet > 0:
        return 'OK: void (retur)'
    return 'CEK'


def reconcile_invoices(invoices=None, jual_dir=None):
    """Reconcile the given invoices against Jual + Saldo + Fee.

    Writes reports/shopee/Cek Jual Shopee.xlsx and returns the DataFrame.
    """
    invoices = list(invoices) if invoices else list(DEFAULT_INVOICES)
    jual = _jual_detail(invoices, jual_dir)
    net = _saldo_net(invoices)
    fee = _fee_detail(invoices)

    rows = []
    for inv in invoices:
        j = jual.get(inv)
        real = net.get(inv)
        fe = fee.get(inv)
        rows.append({
            'Invoice': inv,
            'Di Jual?': 'Ya' if j else 'Tidak',
            'Omzet (Jual)': round(j['omzet']) if j else 0,
            'Omzet Void': round(j['void_omzet']) if j else 0,
            'Penghasilan (Saldo)': '' if real is None else round(real),
            'Total Penghasilan (Fee)': '' if fe is None else round(fe['tp']),
            'Kerugian (Fee)': '' if fe is None else round(fe['ker']),
            'Verdict': _verdict(j, real, fe),
        })
    df = pd.DataFrame(rows)
    df['_bug'] = df['Verdict'].str.startswith('BUG')
    df = df.sort_values(['_bug', 'Verdict', 'Invoice'], ascending=[False, True, True]).drop(columns='_bug')

    _write(df)
    bugs = int(df['Verdict'].str.startswith('BUG').sum())
    logging.info("Cek Jual: %d invoice, %d BUG ditemukan", len(df), bugs)
    for verdict, count in df['Verdict'].value_counts().items():
        logging.info("   %2d  %s", count, verdict)
    return df


def _write(df):
    folder = os.path.join(get_reports_dir(), MARKETPLACE_FOLDERS['Shopee'])
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, 'Cek Jual Shopee.xlsx')
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cek Jual', index=False)
        sheet = writer.book['Cek Jual']
        for i, width in enumerate([18, 12, 16, 12, 22, 22, 14, 46]):
            sheet.column_dimensions[get_column_letter(i + 1)].width = width
        for cell in sheet[1]:
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sheet.freeze_panes = 'A2'
        verdict_col = list(df.columns).index('Verdict') + 1
        red = PatternFill('solid', fgColor='FFC7CE')
        green = PatternFill('solid', fgColor='C6EFCE')
        for row in range(2, len(df) + 2):
            cell = sheet.cell(row=row, column=verdict_col)
            cell.fill = red if str(cell.value).startswith('BUG') else green
