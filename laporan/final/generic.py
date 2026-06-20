import glob
import logging
import os

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from laporan.utility.constant import MARKETPLACE_FOLDERS, get_reports_dir

FINAL_SHEET = 'Final'

# Money columns rendered as integers (blank cells stay blank, not 0/NaN).
_MONEY_COLUMNS = [
    'Ongkir', 'Asuransi', 'Omzet\nBarang', 'Nominal\nInvoice',
    'Potongan Pembayaran', 'Nominal Remit', 'Keuntungan\nTambahan', 'Kerugian Tambahan',
]

# Remit amount columns summed when an order has several remit entries.
_REMIT_AMOUNT_COLUMNS = ['Potongan Pembayaran', 'Nominal Remit', 'Keuntungan Tambahan', 'Kerugian Tambahan']

# Column groups, colored to show their source: 1 = order side (Invoice/Jual),
# 2 = remit side (Remit), 3 = manual columns.
_GROUP_ORDER = 1
_GROUP_REMIT = 2
_GROUP_MANUAL = 3

# (header, width, group) for each Final column, in order.
_COLUMN_LAYOUT = [
    ('Tanggal\nPesan', 20, _GROUP_ORDER), ('Marketplace', 14, _GROUP_ORDER),
    ('Invoice', 40, _GROUP_ORDER), ('Ongkir', 12, _GROUP_ORDER),
    ('Asuransi', 12, _GROUP_ORDER), ('Omzet\nBarang', 15, _GROUP_ORDER),
    ('Nominal\nInvoice', 16, _GROUP_ORDER),
    ('Tanggal\nRemit', 18, _GROUP_REMIT), ('Potongan Pembayaran', 20, _GROUP_REMIT),
    ('Nominal Remit', 15, _GROUP_REMIT), ('Keuntungan\nTambahan', 14, _GROUP_REMIT),
    ('Kerugian Tambahan', 18, _GROUP_REMIT), ('Cek\nRemit', 12, _GROUP_REMIT),
    ('Untung Lainnya', 16, _GROUP_MANUAL), ('Rugi Lainnya', 14, _GROUP_MANUAL),
    ('Keterangan', 30, _GROUP_MANUAL),
]

# Per-group colors: (header fill, data fill). Header gets a strong fill + white
# bold font; the data rows get a light tint of the same hue.
_GROUP_STYLE = {
    _GROUP_ORDER: ('4472C4', 'D9E1F2'),   # blue
    _GROUP_REMIT: ('548235', 'E2EFDA'),   # green
    _GROUP_MANUAL: ('BF8F00', 'FFF2CC'),  # amber
}
_HEADER_FONT_COLOR = 'FFFFFF'
_MONEY_FORMAT = '#,##0'
_GRID = Side(style='thin', color='D9D9D9')
_BORDER = Border(left=_GRID, right=_GRID, top=_GRID, bottom=_GRID)


def _to_number(series):
    """Coerce an Excel-read column (often stored as text) to numeric."""
    return pd.to_numeric(series, errors='coerce')


def _read(xls, sheet_name):
    """Return a sheet from an open ExcelFile, or None if it is absent."""
    if sheet_name in xls.sheet_names:
        return xls.parse(sheet_name)
    return None


def _list_workbooks(folder):
    files = glob.glob(os.path.join(folder, '*.xlsx'))
    # Skip Excel temp lock files (~$...).
    return sorted(f for f in files if not os.path.basename(f).startswith('~'))


def _combined_remit(workbooks, remit_sheet):
    """Union every workbook's Remit sheet, one aggregated row per Invoice.

    Lets an order be matched to a remit that landed in a different period's
    Laporan (ordered this month, remitted next month).
    """
    frames = []
    for path in workbooks:
        with pd.ExcelFile(path) as xls:
            remit = _read(xls, remit_sheet)
        if remit is None or remit.empty or 'Invoice' not in remit.columns:
            continue
        keep = ['Invoice', 'Tanggal Remit'] + _REMIT_AMOUNT_COLUMNS
        remit = remit[[c for c in keep if c in remit.columns]].copy()
        remit['Invoice'] = remit['Invoice'].astype(str)
        for col in _REMIT_AMOUNT_COLUMNS:
            if col in remit.columns:
                remit[col] = _to_number(remit[col])
        frames.append(remit)

    if not frames:
        return None

    allremit = pd.concat(frames, ignore_index=True)
    agg = {col: 'sum' for col in _REMIT_AMOUNT_COLUMNS if col in allremit.columns}
    if 'Tanggal Remit' in allremit.columns:
        agg['Tanggal Remit'] = 'max'  # latest remit date (ISO strings sort correctly)
    return allremit.groupby('Invoice', as_index=False).agg(agg)


def _build_final(invoice_df, jual_df, remit_idx):
    """Join one workbook's Invoice with its Jual omzet and the remit index."""
    df = invoice_df.copy()
    df['Invoice'] = df['Invoice'].astype(str)
    for col in ['Ongkir', 'Asuransi']:
        df[col] = _to_number(df[col]).fillna(0)

    # Omzet Barang = sum of Jual Omzet per Invoice.
    if jual_df is not None and {'Invoice', 'Omzet'}.issubset(jual_df.columns):
        omzet = jual_df[['Invoice', 'Omzet']].copy()
        omzet['Invoice'] = omzet['Invoice'].astype(str)
        omzet['Omzet'] = _to_number(omzet['Omzet']).fillna(0)
        omzet = omzet.groupby('Invoice', as_index=False)['Omzet'].sum()
        df = df.merge(omzet, on='Invoice', how='left')
    else:
        df['Omzet'] = 0
    df['Omzet'] = df['Omzet'].fillna(0)

    # Nominal Invoice = Omzet Barang + Ongkir + Asuransi.
    df['Nominal Invoice'] = df['Omzet'] + df['Ongkir'] + df['Asuransi']

    # Remit columns (left join; blank when no remit is found in any period).
    if remit_idx is not None:
        df = df.merge(remit_idx, on='Invoice', how='left')
    for col in ['Tanggal Remit'] + _REMIT_AMOUNT_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA

    # Assemble in the requested column order; Cek Remit + the "Lainnya"/Keterangan
    # columns are left blank for manual entry.
    out = pd.DataFrame({
        'Tanggal\nPesan': df['Tanggal'].values,
        'Marketplace': df['Marketplace'].values,
        'Invoice': df['Invoice'].values,
        'Ongkir': df['Ongkir'].values,
        'Asuransi': df['Asuransi'].values,
        'Omzet\nBarang': df['Omzet'].values,
        'Nominal\nInvoice': df['Nominal Invoice'].values,
        'Tanggal\nRemit': df['Tanggal Remit'].values,
        'Potongan Pembayaran': df['Potongan Pembayaran'].values,
        'Nominal Remit': df['Nominal Remit'].values,
        'Keuntungan\nTambahan': df['Keuntungan Tambahan'].values,
        'Kerugian Tambahan': df['Kerugian Tambahan'].values,
        'Cek\nRemit': pd.NA,
        'Untung Lainnya': pd.NA,
        'Rugi Lainnya': pd.NA,
        'Keterangan': pd.NA,
    })

    # Render money as integers; unmatched remit cells stay blank (Int64 <NA>).
    # to_numeric first so an all-blank remit column (object NA) still rounds cleanly.
    for col in _MONEY_COLUMNS:
        out[col] = pd.to_numeric(out[col], errors='coerce').round().astype('Int64')

    return out.sort_values('Invoice').reset_index(drop=True)


def _style_final(sheet, n_rows):
    """Style the Final sheet to match Invoice and color the column groups."""
    last_row = n_rows + 1  # +1 for the header row

    # Index/row-number gutter (column A), like Invoice.
    sheet.column_dimensions['A'].width = 6
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    for row in range(1, last_row + 1):
        sheet.cell(row=row, column=1).border = _BORDER

    for offset, (header, width, group) in enumerate(_COLUMN_LAYOUT):
        col = offset + 2  # data starts at column B (column A is the index)
        sheet.column_dimensions[get_column_letter(col)].width = width
        header_fill, data_fill = _GROUP_STYLE[group]

        head = sheet.cell(row=1, column=col)
        head.fill = PatternFill('solid', fgColor=header_fill)
        head.font = Font(bold=True, color=_HEADER_FONT_COLOR)
        head.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        head.border = _BORDER

        number_format = _MONEY_FORMAT if header in _MONEY_COLUMNS else None
        for row in range(2, last_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill('solid', fgColor=data_fill)
            cell.border = _BORDER
            if number_format:
                cell.number_format = number_format

    sheet.row_dimensions[1].height = 30  # room for the two-line headers
    sheet.freeze_panes = 'B2'  # keep the header row and the row-number gutter in view


def _write_final(out, path, sheet_name=FINAL_SHEET):
    """Append (or replace) the Final sheet in an existing Laporan workbook."""
    out = out.copy()
    out.index = range(1, len(out) + 1)  # 1-based row numbers, like Invoice

    try:
        writer = pd.ExcelWriter(path, mode='a', engine='openpyxl', if_sheet_exists='replace')
    except (FileNotFoundError, ValueError):
        writer = pd.ExcelWriter(path, engine='openpyxl')

    out.to_excel(writer, sheet_name=sheet_name)  # include the row-number index column

    _style_final(writer.book[sheet_name], n_rows=len(out))

    writer.close()


def generate_final(marketplace):
    """Write the Final sheet into every Laporan workbook of one marketplace.

    marketplace: the report token used in the sheet/file names and as a
    MARKETPLACE_FOLDERS key ('Shopee' / 'Tiktok' / 'Tokopedia' / 'Bukalapak').
    """
    subfolder = MARKETPLACE_FOLDERS.get(marketplace)
    if subfolder is None:
        return
    folder = os.path.join(get_reports_dir(), subfolder)
    if not os.path.isdir(folder):
        return
    workbooks = _list_workbooks(folder)
    if not workbooks:
        return

    invoice_sheet = 'Invoice {0}'.format(marketplace)
    jual_sheet = 'Jual {0}'.format(marketplace)
    remit_sheet = 'Remit {0}'.format(marketplace)

    # Build the cross-period remit index once, before writing any Final sheet.
    remit_idx = _combined_remit(workbooks, remit_sheet)

    for path in workbooks:
        with pd.ExcelFile(path) as xls:
            invoice_df = _read(xls, invoice_sheet)
            jual_df = _read(xls, jual_sheet)
        if invoice_df is None or invoice_df.empty or 'Invoice' not in invoice_df.columns:
            continue  # no orders to anchor a Final sheet on
        out = _build_final(invoice_df, jual_df, remit_idx)
        _write_final(out, path)
        logging.info("Generate Final {0} -> {1} ({2} rows)".format(
            marketplace, os.path.basename(path), len(out)))
