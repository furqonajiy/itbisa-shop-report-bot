"""Excel report writer."""
from __future__ import annotations
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    ALERT_TEXT_COLOR, BLUE_FILL_COLOR, FMT_DEC, FMT_NUM, FMT_PCT, FMT_RP,
    FONT_NAME, GREEN_FILL_COLOR, HEADER_BG_COLOR, HEADER_TEXT_COLOR,
    LEAD_TIME_MARKET_MONTHS, LEAD_TIME_PERCENTILE, LEDGER_SHEET_NAME,
    LIGHT_GRAY_COLOR,
    MARKUP_THRESHOLD_KANDIDAT, ORANGE_FILL_COLOR, OVERSTOCK_MONTHS,
    PRICE_SCENARIOS, RED_FILL_COLOR, ROP_SOON_RATIO, ROP_URGENT_RATIO,
    SAFETY_MULT_MODERATE, SAFETY_MULT_STABLE, SAFETY_MULT_VOLATILE,
    SLOW_DEAD_MAX_VELOCITY, TARGET_MARKUP_KOREKSI, TARGET_MONTHS_POST_REORDER,
    TITLE_COLOR, TOP_N_PER_PLATFORM, YELLOW_FILL_COLOR,
)
from tables import build_top_per_platform

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
ALERT_FONT = Font(name=FONT_NAME, italic=True, size=10, color=ALERT_TEXT_COLOR)
RED_FILL = PatternFill("solid", start_color=RED_FILL_COLOR)
ORANGE_FILL = PatternFill("solid", start_color=ORANGE_FILL_COLOR)
GREEN_FILL = PatternFill("solid", start_color=GREEN_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
BLUE_FILL = PatternFill("solid", start_color=BLUE_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def write_headers(ws: Worksheet, row: int, headers: list[str],
                  widths: list[float] | None = None) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 30


def write_data_rows(ws: Worksheet, start_row: int, df: pd.DataFrame,
                    formats: list[str | None] | None = None,
                    negative_highlight_col: int | None = None) -> None:
    for r_idx, (_, row) in enumerate(df.iterrows()):
        for c_idx, val in enumerate(row, start=1):
            if pd.isna(val):
                val = None
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if formats and c_idx - 1 < len(formats) and formats[c_idx - 1]:
                cell.number_format = formats[c_idx - 1]
            if r_idx % 2 == 1:
                cell.fill = LIGHT_FILL
        if negative_highlight_col is not None:
            target = ws.cell(row=start_row + r_idx, column=negative_highlight_col)
            if isinstance(target.value, (int, float)) and target.value < 0:
                for c in range(1, len(row) + 1):
                    ws.cell(row=start_row + r_idx, column=c).fill = RED_FILL


def _build_findings(sku_agg: pd.DataFrame, tables: dict,
                    sku_no_hpp: list[str], oversold: list[str] | None = None) -> list[str]:
    lines: list[str] = []

    rugi = tables["rugi"]
    if len(rugi) > 0:
        lines.append("🔴 BARANG RUGI YANG HARUS SEGERA DINAIKKAN HARGANYA:")
        for _, r in rugi.head(5).iterrows():
            hpp_dasar = r["hpp_pricing"]
            rekom = hpp_dasar * (1 + TARGET_MARKUP_KOREKSI)
            lines.append(
                f"   • {r['SKU']} — terjual {r['qty_terjual']:,.0f} pcs, "
                f"RUGI Rp {r['profit']:,.0f}. HPP Rp {r['hpp_wa']:,.0f}, "
                f"dijual Rp {r['harga_jual_avg']:,.0f}. "
                f"Saran: Rp {rekom:,.0f} (HPP dasar harga Rp {hpp_dasar:,.0f} × {1 + TARGET_MARKUP_KOREKSI:.2f})."
            )
        lines.append("")

    kandidat = tables["kandidat"]
    if len(kandidat) > 0:
        lines.append("🟢 TOP KANDIDAT NAIK HARGA (laris + margin sehat + stok ada):")
        for _, r in kandidat.head(5).iterrows():
            if r.get("harga_baru_flag"):
                tgl = r.get("tgl_naik")
                tgl_str = tgl.strftime("%d %b %Y") if pd.notna(tgl) else "baru-baru ini"
                lines.append(
                    f"   • {r['SKU']} — {r['qty_terjual']:,.0f} pcs, harga BARU naik "
                    f"{tgl_str} (Rp{r['harga_sekarang']:,.0f}) & belum tervalidasi — "
                    f"pantau dulu, jangan naik lagi."
                )
                continue
            note = ""
            if r["restock_di_tahun"] and pd.notna(r["qty_setelah_restock"]) and r["qty_terjual"] > 0:
                pct = r["qty_setelah_restock"] / r["qty_terjual"] * 100
                note = (f", restock tahun ini langsung "
                        f"{r['qty_setelah_restock']:,.0f} pcs terjual ({pct:.0f}%)")
            lines.append(
                f"   • {r['SKU']} — {r['qty_terjual']:,.0f} pcs, "
                f"markup {r['markup_pct']:.1f}%, stok {r['sisa_stok']:,.0f}{note}."
            )
        lines.append("")

    top_qty = sku_agg.nlargest(3, "qty_terjual")
    lines.append("⭐ BARANG PALING DIMINATI:")
    for i, (_, r) in enumerate(top_qty.iterrows(), 1):
        lines.append(
            f"   {i}. {r['SKU']} — {r['qty_terjual']:,.0f} pcs, "
            f"{r['jumlah_transaksi']:,} transaksi "
            f"(rata-rata {r['avg_qty_per_order']:.0f} pcs/order)"
        )
    lines.append("")

    top_profit = sku_agg.nlargest(5, "profit")
    lines.append("💰 PENYUMBANG PROFIT TERBESAR:")
    for i, (_, r) in enumerate(top_profit.iterrows(), 1):
        lines.append(f"   {i}. {r['SKU']} — Rp {r['profit']:,.0f}")
    lines.append("")

    plat = tables["platform"]
    total_omzet_plat = plat["omzet"].sum()
    lines.append("🏪 PER PLATFORM:")
    for _, r in plat.iterrows():
        share = r["omzet"] / total_omzet_plat * 100 if total_omzet_plat else 0
        lines.append(
            f"   • {r['akun_penjual']}: {share:.0f}% omzet, "
            f"margin {r['margin_pct']:.1f}%, biaya admin {r['admin_pct']:.1f}%"
        )
    lines.append("")

    supplier = tables.get("supplier", {})
    comp = supplier.get("comparison", pd.DataFrame())
    volatile = supplier.get("volatile", pd.DataFrame())
    if len(comp) > 0 or len(volatile) > 0:
        lines.append("🏭 SUPPLIER INSIGHT (China direct vs Market):")
        stop_china = comp[comp["rekomendasi"].astype(str).str.startswith("🔴")]
        if len(stop_china) > 0:
            lines.append(f"   • {len(stop_china)} SKU lebih murah dari Market dibanding China — pertimbangkan stop reorder China")
            for _, r in stop_china.head(3).iterrows():
                lines.append(f"     - {r['SKU']}: HPP China Rp {r['hpp_china']:,.0f} vs Market Rp {r['hpp_market']:,.0f}")
        if len(volatile) > 0:
            lines.append(f"   • {len(volatile)} SKU dengan HPP China tidak konsisten (CV > 15%)")
        lines.append("")

    reorder = tables.get("reorder", {})
    if reorder:
        n_stockout = len(reorder.get("stockout", []))
        n_urgent = len(reorder.get("urgent", []))
        n_now = len(reorder.get("now", []))
        n_over = len(reorder.get("overstock", []))
        if n_stockout + n_urgent + n_now + n_over > 0:
            lines.append("📦 REORDER PRIORITAS (lihat sheet 09 untuk detail):")
            if n_stockout > 0:
                lines.append(f"   • 🔴 {n_stockout} SKU sudah STOCKOUT — kehilangan sales sekarang")
            if n_urgent > 0:
                lines.append(f"   • 🔴 {n_urgent} SKU URGENT reorder — stok hampir habis")
            if n_now > 0:
                lines.append(f"   • 🟠 {n_now} SKU perlu reorder minggu ini")
            if n_over > 0:
                lines.append(f"   • 🔵 {n_over} SKU overstock (> {OVERSTOCK_MONTHS:.0f} bulan cadangan) — stop reorder")
            lines.append("")

    dq: list[str] = []
    if oversold:
        dq.append(
            f"{len(oversold)} SKU OVERSOLD (sisa stok tercatat negatif — kemungkinan "
            f"baris pembelian hilang/salah tag, cek data sumber): "
            + ", ".join(oversold[:10]) + ("..." if len(oversold) > 10 else "")
        )
    if sku_no_hpp:
        dq.append(
            f"{len(sku_no_hpp)} SKU dijual tanpa HPP (di-exclude dari profit): "
            + ", ".join(sku_no_hpp[:10]) + ("..." if len(sku_no_hpp) > 10 else "")
        )
    if dq:
        lines.append("🧹 KUALITAS DATA (rapikan di sumber agar laporan makin akurat):")
        for x in dq:
            lines.append(f"   • {x}")
        lines.append("")
    return lines


def _write_summary(ws, year, jual, sku_agg, tables, sku_no_hpp, oversold=None):
    ws["A1"] = f"LAPORAN ANALISA PENJUALAN ITBISA SHOP {year}"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = (f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}  |  "
                f"Periode: {year}  |  Metode HPP: Weighted Average")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:D2")

    # Partial / in-progress year: totals here are NOT a full year — flag so they
    # aren't compared straight against complete years.
    last_date = pd.to_datetime(jual["tanggal_pesan"]).max()
    if year == datetime.now().year and pd.notna(last_date):
        c = ws.cell(row=3, column=1, value=(
            f"⚠ TAHUN BERJALAN — data s/d {last_date.strftime('%d %b %Y')} "
            f"(belum setahun penuh). Total di bawah jangan dibandingkan langsung "
            f"dengan tahun yang sudah penuh."))
        c.font = ALERT_FONT
        ws.merge_cells("A3:D3")

    total_omzet = jual["omzet"].sum()
    total_hpp = jual["hpp_total"].sum()
    total_admin = jual["biaya_admin"].sum()
    total_profit = jual["profit"].sum()
    total_qty = jual["qty_jual"].sum()
    total_trans = jual["Invoice"].nunique()
    margin = total_profit / total_omzet if total_omzet else 0

    ws["A4"] = "RINGKASAN PERFORMA"
    ws["A4"].font = TITLE_FONT
    summary = [
        ("Total Transaksi", total_trans, FMT_NUM),
        ("Total Qty Terjual (pcs)", total_qty, FMT_NUM),
        ("Unique SKU Terjual", len(sku_agg), FMT_NUM),
        ("", "", ""),
        ("Total Omzet", total_omzet, FMT_RP),
        ("Total HPP", total_hpp, FMT_RP),
        ("Total Biaya Admin Platform", total_admin, FMT_RP),
        ("TOTAL PROFIT", total_profit, FMT_RP),
        ("Margin Keseluruhan", margin, FMT_PCT),
        ("", "", ""),
        ("Barang RUGI Total", len(tables["rugi"]), FMT_NUM),
        ("Markup di Bawah Floor 30%", len(tables["borderline"]), FMT_NUM),
        ("Kandidat Naik Harga", len(tables["kandidat"]), FMT_NUM),
    ]
    for i, (lbl, val, fmt) in enumerate(summary, start=6):
        is_bold = "TOTAL" in str(lbl) or "Margin" in str(lbl)
        c1 = ws.cell(row=i, column=1, value=lbl)
        c1.font = BOLD_FONT if is_bold else NORMAL_FONT
        c2 = ws.cell(row=i, column=2, value=val)
        c2.font = BOLD_FONT if is_bold else NORMAL_FONT
        if fmt:
            c2.number_format = fmt
        if "TOTAL PROFIT" in str(lbl):
            c2.fill = GREEN_FILL

    ws["A21"] = "TEMUAN UTAMA & REKOMENDASI"
    ws["A21"].font = TITLE_FONT
    findings = _build_findings(sku_agg, tables, sku_no_hpp, oversold)
    for i, line in enumerate(findings, start=22):
        c = ws.cell(row=i, column=1, value=line)
        c.font = NORMAL_FONT
        if line and any(line.startswith(em) for em in ["🔴", "🟢", "⭐", "💰", "🏪", "🏭", "📦", "🧹"]):
            c.font = BOLD_FONT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22


def _write_diminati(ws, df):
    ws["A1"] = "TABEL 1: BARANG PALING DIMINATI (sort by Qty Terjual)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = "Avg Qty/Order TINGGI = reseller borongan. RENDAH = retail individual."
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:I2")
    write_headers(ws, 4,
                  ["SKU", "Qty Terjual", "Jumlah Transaksi", "Avg Qty/Order",
                   "Harga Jual Avg", "Omzet", "HPP/buah", "Profit", "Margin %"],
                  widths=[38, 13, 13, 13, 15, 18, 13, 18, 12])
    df = df.copy()
    df["margin_frac"] = df["margin_pct"] / 100
    out = df[["SKU", "qty_terjual", "jumlah_transaksi", "avg_qty_per_order",
              "harga_jual_avg", "omzet", "hpp_wa", "profit", "margin_frac"]]
    write_data_rows(ws, 5, out,
                    formats=[None, FMT_NUM, FMT_NUM, FMT_DEC, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_PCT],
                    negative_highlight_col=8)
    ws.freeze_panes = "B5"


def _write_profit(ws, df):
    ws["A1"] = "TABEL 2: BARANG PENYUMBANG PROFIT TERBESAR"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = "Sort by total Profit (Rupiah)."
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:H2")
    write_headers(ws, 4,
                  ["SKU", "Total Profit", "Profit/Buah", "Qty Terjual",
                   "Margin %", "Omzet", "Harga Jual Avg", "HPP/Buah"],
                  widths=[38, 18, 13, 13, 12, 18, 15, 13])
    df = df.copy()
    df["margin_frac"] = df["margin_pct"] / 100
    out = df[["SKU", "profit", "profit_per_buah", "qty_terjual",
              "margin_frac", "omzet", "harga_jual_avg", "hpp_wa"]]
    write_data_rows(ws, 5, out,
                    formats=[None, FMT_RP, FMT_RP, FMT_NUM, FMT_PCT, FMT_RP, FMT_RP, FMT_RP],
                    negative_highlight_col=2)
    ws.freeze_panes = "B5"


def _write_rugi(ws, df):
    ws["A1"] = "TABEL 3: BARANG RUGI / SALAH PASANG HARGA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = ("Selisih = Harga Jual - (HPP_WA + |Admin|). NEGATIF = jual di bawah modal. "
                "Rugi dihitung dari HPP_WA (realisasi); rekomendasi harga pakai HPP dasar harga (lot LN terakhir bila ada).")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:I2")

    if len(df) == 0:
        ws["A4"] = "(tidak ada barang rugi)"
        ws["A4"].font = BOLD_FONT
        return

    write_headers(ws, 4,
                  ["SKU", "Qty Terjual", "HPP/buah", "Admin/buah", "Total Cost/buah",
                   "Harga Jual Avg", "Selisih", "Profit/buah", "Total Profit"],
                  widths=[38, 13, 13, 13, 15, 15, 13, 13, 18])
    out = df[["SKU", "qty_terjual", "hpp_wa", "biaya_admin_per_buah",
              "total_cost_per_buah", "harga_jual_avg", "selisih_harga",
              "profit_per_buah", "profit"]]
    write_data_rows(ws, 5, out,
                    formats=[None, FMT_NUM, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_RP],
                    negative_highlight_col=9)

    row_rec = 5 + len(df) + 2
    ws.cell(row=row_rec, column=1,
            value=f"REKOMENDASI HARGA KOREKSI (HPP dasar harga × {1 + TARGET_MARKUP_KOREKSI:.2f}, markup {TARGET_MARKUP_KOREKSI*100:.0f}%):").font = BOLD_FONT
    for i, (_, r) in enumerate(df.iterrows(), start=row_rec + 1):
        hpp_dasar = r["hpp_pricing"]
        rekom = hpp_dasar * (1 + TARGET_MARKUP_KOREKSI)
        ws.cell(row=i, column=1, value=r["SKU"]).font = NORMAL_FONT
        ws.cell(row=i, column=2, value=f"Sekarang: Rp {r['harga_jual_avg']:,.0f}").font = NORMAL_FONT
        ws.cell(row=i, column=3, value=f"HPP dasar harga: Rp {hpp_dasar:,.0f}").font = NORMAL_FONT
        c = ws.cell(row=i, column=4, value=f"Rekomendasi: Rp {rekom:,.0f}")
        c.font = BOLD_FONT
        c.fill = GREEN_FILL


def _write_borderline(ws, df):
    ws["A1"] = "TABEL 4: MARKUP < 30% (HARGA DI BAWAH FLOOR — PERLU NAIK HARGA)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = ("Floor: harga jual minimal HPP dasar harga × 1.30. HPP dasar harga = harga lot luar negeri "
                "terakhir bila ada, else HPP_WA. SKU di sini di bawah floor — wajib review (kecuali ada alasan kompetitif).")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:I2")

    if len(df) == 0:
        ws["A4"] = "(tidak ada SKU dalam kategori ini)"
        ws["A4"].font = BOLD_FONT
        return

    write_headers(ws, 4,
                  ["SKU", "Qty Terjual", "HPP Dasar Harga", "Admin/buah", "Harga Sekarang",
                   "Markup %", "Margin %", "Profit/buah", "Profit Total"],
                  widths=[38, 13, 14, 13, 15, 11, 11, 13, 18])
    df = df.copy()
    df["markup_frac"] = df["markup_pct"] / 100
    df["margin_frac"] = df["margin_pct"] / 100
    out = df[["SKU", "qty_terjual", "hpp_pricing", "biaya_admin_per_buah", "harga_sekarang",
              "markup_frac", "margin_frac", "profit_per_buah", "profit"]]
    write_data_rows(ws, 5, out,
                    formats=[None, FMT_NUM, FMT_RP, FMT_RP, FMT_RP, FMT_PCT, FMT_PCT, FMT_RP, FMT_RP])

    row_rec = 5 + len(df) + 2
    ws.cell(row=row_rec, column=1,
            value="REKOMENDASI HARGA TARGET (HPP dasar harga × 1.30):").font = BOLD_FONT
    for i, (_, r) in enumerate(df.iterrows(), start=row_rec + 1):
        target = r["hpp_pricing"] * 1.30
        ws.cell(row=i, column=1, value=r["SKU"]).font = NORMAL_FONT
        ws.cell(row=i, column=2, value=f"Sekarang: Rp {r['harga_sekarang']:,.0f}").font = NORMAL_FONT
        ws.cell(row=i, column=3, value=f"HPP dasar harga: Rp {r['hpp_pricing']:,.0f}").font = NORMAL_FONT
        c = ws.cell(row=i, column=4, value=f"Target: Rp {target:,.0f}")
        c.font = BOLD_FONT
        c.fill = GREEN_FILL


def _write_kandidat(ws, df):
    ws["A1"] = "TABEL 5: KANDIDAT NAIK HARGA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:M1")
    ws["A2"] = (f"Kriteria: top 20% qty + markup ≥ {MARKUP_THRESHOLD_KANDIDAT:.0f}% atas HPP dasar harga "
                "(lot LN terakhir bila ada, else HPP_WA) + stok ada. Score = 60% velocity + 40% markup.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:M2")
    ws["A3"] = ('⚠ "Harga Sekarang" = harga satuan TERENDAH di hari penjualan non-CoD terakhir SKU. '
                '"Qty Setelah Restock" tinggi = restock baru cepat habis = harga underpriced. '
                'Baris ⏳ (abu-abu) = harga BARU dinaikkan & belum tervalidasi: kolom Harga +%/Proyeksi '
                'dikosongkan karena qty/profit masih dari harga lama — kumpulkan data dulu, jangan naik lagi.')
    ws["A3"].font = ALERT_FONT
    ws.merge_cells("A3:M3")

    if len(df) == 0:
        ws["A5"] = "(belum ada kandidat memenuhi semua kriteria)"
        ws["A5"].font = BOLD_FONT
        return

    pcts = PRICE_SCENARIOS
    headers = ["SKU", "Score", "Qty Terjual", "Markup %", "Harga Sekarang",
               "Sisa Stok", "Restock Tahun?", "Qty Setelah Restock"]
    headers += [f"Harga +{int(p*100)}%" for p in pcts]
    headers += [f"Proyeksi Profit +{int(pcts[0]*100)}%", "Saran"]
    widths = [35, 8, 12, 10, 14, 11, 13, 14] + [13] * len(pcts) + [18, 35]
    write_headers(ws, 5, headers, widths=widths)

    df = df.copy()
    df["markup_frac"] = df["markup_pct"] / 100
    df["restock_str"] = df["restock_di_tahun"].map({True: "Ya", False: "Tidak"})

    out_cols = ["SKU", "score_total", "qty_terjual", "markup_frac", "harga_sekarang",
                "sisa_stok", "restock_str", "qty_setelah_restock"]
    out_cols += [f"harga_+{int(p*100)}pct" for p in pcts]
    out_cols += [f"proyeksi_profit_+{int(pcts[0]*100)}pct", "saran"]
    out = df[out_cols]

    formats = [None, FMT_DEC, FMT_NUM, FMT_PCT, FMT_RP, FMT_NUM, None, FMT_NUM]
    formats += [FMT_RP] * len(pcts)
    formats += [FMT_RP, None]
    write_data_rows(ws, 6, out, formats=formats)

    # Highlight the top non-held candidates yellow (act on these); a recent,
    # under-validated hike (harga_baru_flag) is greyed out (held, see Saran).
    flags = (df["harga_baru_flag"].tolist() if "harga_baru_flag" in df.columns
             else [False] * len(df))
    yellow_left = 5
    for r_idx in range(len(df)):
        if flags[r_idx]:
            fill = LIGHT_FILL
        elif yellow_left > 0:
            fill = YELLOW_FILL
            yellow_left -= 1
        else:
            continue
        for c in range(1, len(headers) + 1):
            ws.cell(row=6 + r_idx, column=c).fill = fill


def _write_platform(ws, plat_df, jual):
    ws["A1"] = "TABEL 6: BREAKDOWN PER PLATFORM"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = "Lihat margin dan biaya admin per platform."
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:I2")
    write_headers(ws, 4,
                  ["Platform", "Transaksi", "Qty", "Omzet", "HPP",
                   "Biaya Admin", "Profit", "Margin %", "Admin %"],
                  widths=[15, 12, 13, 18, 18, 18, 18, 12, 12])
    plat_df = plat_df.copy()
    plat_df["margin_frac"] = plat_df["margin_pct"] / 100
    plat_df["admin_frac"] = plat_df["admin_pct"] / 100
    out = plat_df[["akun_penjual", "transaksi", "qty", "omzet", "hpp",
                   "biaya_admin", "profit", "margin_frac", "admin_frac"]]
    write_data_rows(ws, 5, out,
                    formats=[None, FMT_NUM, FMT_NUM, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_PCT, FMT_PCT])

    section_row = 5 + len(plat_df) + 2
    ws.cell(row=section_row, column=1, value="TOP SKU BY PROFIT — PER PLATFORM").font = TITLE_FONT

    row_cur = section_row + 2
    for platform in plat_df["akun_penjual"].tolist():
        top = build_top_per_platform(jual, platform, TOP_N_PER_PLATFORM)
        if len(top) == 0:
            continue
        ws.cell(row=row_cur, column=1, value=platform).font = BOLD_FONT
        write_headers(ws, row_cur + 1, ["SKU", "Qty", "Omzet", "Profit"],
                      widths=[38, 12, 18, 18])
        write_data_rows(ws, row_cur + 2, top,
                        formats=[None, FMT_NUM, FMT_RP, FMT_RP],
                        negative_highlight_col=4)
        row_cur += 3 + len(top)


def _write_full_data(ws, sku_agg):
    ws["A1"] = "DATA LENGKAP PER SKU"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:V1")
    cols = ["SKU", "qty_terjual", "jumlah_transaksi", "avg_qty_per_order", "omzet",
            "hpp_wa", "hpp_source", "hpp_pricing", "hpp_pricing_source",
            "hpp_cost", "biaya_admin", "biaya_admin_per_buah",
            "profit", "profit_per_buah", "harga_jual_avg", "harga_sekarang",
            "markup_pct", "margin_pct",
            "total_qty_beli", "sisa_stok", "jumlah_pembelian", "restock_di_tahun"]
    headers = ["SKU", "Qty Terjual", "Trans", "Avg Qty/Order", "Omzet", "HPP/Buah (P&L)",
               "HPP Source", "HPP Dasar Harga", "Sumber HPP Harga", "HPP Total",
               "Biaya Admin", "Admin/Buah", "Profit",
               "Profit/Buah", "Harga Jual Avg", "Harga Sekarang", "Markup %", "Margin %",
               "Total Dibeli", "Sisa Stok", "Lot Beli", "Restock Tahun"]
    widths = [38, 12, 8, 13, 16, 13, 10, 14, 13, 16, 16, 12, 16, 12, 14, 14, 11, 11, 12, 12, 9, 11]
    write_headers(ws, 3, headers, widths=widths)
    df_full = sku_agg[cols].sort_values("profit", ascending=False).copy()
    df_full["markup_pct"] = df_full["markup_pct"] / 100
    df_full["margin_pct"] = df_full["margin_pct"] / 100
    df_full["restock_di_tahun"] = df_full["restock_di_tahun"].map({True: "Ya", False: "Tidak"})
    write_data_rows(ws, 4, df_full,
                    formats=[None, FMT_NUM, FMT_NUM, FMT_DEC, FMT_RP, FMT_RP, None, FMT_RP, None,
                             FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_RP, FMT_PCT, FMT_PCT,
                             FMT_NUM, FMT_NUM, FMT_NUM, None],
                    negative_highlight_col=13)
    ws.freeze_panes = "B4"


def _write_supplier_analysis(ws, supplier_tables):
    ws["A1"] = "TABEL 8: SUPPLIER ANALYSIS — CHINA DIRECT vs MARKET BUY"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = ("China = Ocistok/Martkita/Aliexpress/Jasa Impor (import langsung). "
                "Market = beli ulang dari Shopee/Tokopedia/Bukalapak/Tiktok.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:I2")

    row_cur = 4

    comp = supplier_tables["comparison"]
    ws.cell(row=row_cur, column=1, value="A. PERBANDINGAN CHINA vs MARKET (SKU dengan kedua sumber)").font = TITLE_FONT
    row_cur += 1
    if len(comp) == 0:
        ws.cell(row=row_cur, column=1, value="(tidak ada SKU dengan pembelian dari China DAN Market)").font = NORMAL_FONT
        row_cur += 2
    else:
        write_headers(ws, row_cur,
                      ["SKU", "Qty Terjual", "n China", "HPP China",
                       "n Market", "HPP Market", "Selisih (M−C)", "CV China", "Rekomendasi"],
                      widths=[38, 12, 9, 14, 9, 14, 14, 10, 40])
        out = comp[["SKU", "qty_terjual", "n_china", "hpp_china",
                    "n_market", "hpp_market", "selisih_market_vs_china",
                    "cv_china", "rekomendasi"]]
        write_data_rows(ws, row_cur + 1, out,
                        formats=[None, FMT_NUM, FMT_NUM, FMT_RP, FMT_NUM, FMT_RP, FMT_RP, FMT_PCT, None])
        for i in range(len(comp)):
            if str(comp.iloc[i]["rekomendasi"]).startswith("🔴"):
                for c in range(1, 10):
                    ws.cell(row=row_cur + 1 + i, column=c).fill = RED_FILL
        row_cur += len(comp) + 3

    vol = supplier_tables["volatile"]
    ws.cell(row=row_cur, column=1, value="B. CHINA-ONLY DENGAN HPP TIDAK KONSISTEN (CV > 15%)").font = TITLE_FONT
    row_cur += 1
    if len(vol) == 0:
        ws.cell(row=row_cur, column=1, value="(tidak ada SKU dengan variance HPP tinggi dari China)").font = NORMAL_FONT
        row_cur += 2
    else:
        write_headers(ws, row_cur,
                      ["SKU", "Qty Terjual", "n China", "HPP Min", "HPP Max", "HPP Avg", "CV %", "Rekomendasi"],
                      widths=[38, 12, 9, 13, 13, 13, 10, 40])
        out = vol[["SKU", "qty_terjual", "n_china", "hpp_min_china",
                   "hpp_max_china", "hpp_china", "cv_china", "rekomendasi"]]
        write_data_rows(ws, row_cur + 1, out,
                        formats=[None, FMT_NUM, FMT_NUM, FMT_RP, FMT_RP, FMT_RP, FMT_PCT, None])
        row_cur += len(vol) + 3

    co = supplier_tables["china_only"]
    ws.cell(row=row_cur, column=1, value=f"C. TOP CHINA-ONLY (potensi test market buy)").font = TITLE_FONT
    row_cur += 1
    if len(co) == 0:
        ws.cell(row=row_cur, column=1, value="(tidak ada)").font = NORMAL_FONT
        row_cur += 2
    else:
        write_headers(ws, row_cur,
                      ["SKU", "Qty Terjual", "n China", "HPP China"],
                      widths=[38, 12, 9, 14])
        out = co[["SKU", "qty_terjual", "n_china", "hpp_china"]]
        write_data_rows(ws, row_cur + 1, out,
                        formats=[None, FMT_NUM, FMT_NUM, FMT_RP])
        row_cur += len(co) + 3

    mo = supplier_tables["market_only"]
    ws.cell(row=row_cur, column=1, value=f"D. TOP MARKET-ONLY (potensi test import China)").font = TITLE_FONT
    row_cur += 1
    if len(mo) == 0:
        ws.cell(row=row_cur, column=1, value="(tidak ada)").font = NORMAL_FONT
    else:
        write_headers(ws, row_cur,
                      ["SKU", "Qty Terjual", "n Market", "HPP Market"],
                      widths=[38, 12, 9, 14])
        out = mo[["SKU", "qty_terjual", "n_market", "hpp_market"]]
        write_data_rows(ws, row_cur + 1, out,
                        formats=[None, FMT_NUM, FMT_NUM, FMT_RP])


_REORDER_BUCKET_HEADERS = [
    "SKU", "Sisa Stok", "V 3mo", "V 6mo", "V 12mo",
    "Vel. Pakai", "Basis", "Volatility", "Max 1 Order",
    "Lead (bln)", "ROP", "Bulan Cover", "Suggest Order",
    "Last Purchase",
]
_REORDER_BUCKET_WIDTHS = [38, 11, 10, 10, 10, 11, 7, 10, 11, 10, 11, 11, 13, 13]
_REORDER_BUCKET_COLS = [
    "SKU", "sisa_stok", "v3mo", "v6mo", "v12mo",
    "velocity_used", "velocity_basis", "volatility", "max_single_order",
    "lead_months", "rop_final", "months_cover", "qty_order_suggest",
    "last_purchase",
]
_REORDER_BUCKET_FORMATS = [
    None, FMT_NUM, FMT_DEC, FMT_DEC, FMT_DEC,
    FMT_DEC, None, None, FMT_NUM,
    FMT_DEC, FMT_DEC, FMT_DEC, FMT_NUM,
    "yyyy-mm-dd",
]


def _write_reorder_bucket(ws, row_cur, title, df, fill_color):
    c = ws.cell(row=row_cur, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = fill_color
    row_cur += 1
    if len(df) == 0:
        ws.cell(row=row_cur, column=1, value="(tidak ada SKU dalam kategori ini)").font = NORMAL_FONT
        return row_cur + 2
    write_headers(ws, row_cur, _REORDER_BUCKET_HEADERS, widths=_REORDER_BUCKET_WIDTHS)
    out = df[_REORDER_BUCKET_COLS]
    write_data_rows(ws, row_cur + 1, out, formats=_REORDER_BUCKET_FORMATS)
    return row_cur + len(df) + 3


def _write_reorder_analysis(ws, reorder_tables, today=None):
    if today is None:
        today = datetime.now()

    ws["A1"] = "TABEL 9: ANALISA REORDER — KAPAN & BERAPA BANYAK"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")
    ws["A2"] = (f"Snapshot per {today.strftime('%d %B %Y')}. "
                "Sisa stok dan velocity berdasarkan SEMUA data (lintas tahun), "
                "bukan hanya tahun analisa.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:N2")

    row_cur = 4

    ws.cell(row=row_cur, column=1, value="METODOLOGI").font = TITLE_FONT
    row_cur += 1
    method_lines = [
        f"• Velocity = rata-rata qty terjual per bulan. Window 6mo default, fallback 12mo/24mo.",
        f"• CV (volatility) = std/avg qty bulanan. <{int(0.3*100)}%=Stabil, "
        f"<{int(0.7*100)}%=Moderate, ≥{int(0.7*100)}%=Volatile.",
        f"• Safety multiplier × lead demand: Stabil={SAFETY_MULT_STABLE}×, "
        f"Moderate={SAFETY_MULT_MODERATE}×, Volatile={SAFETY_MULT_VOLATILE}×.",
        f"• Lead time impor: dari riwayat kirim Ocistok/Martkita (Tanggal Bayar→Sampai), "
        f"persentil {int(LEAD_TIME_PERCENTILE*100)} per-SKU (fallback global China). "
        f"Market buy={LEAD_TIME_MARKET_MONTHS} bulan (≈ 1 minggu).",
        "• ROP = MAX dari (a) lead demand × safety multiplier, "
        "(b) lead demand + max 1 order (proteksi bulk buyer).",
        f"• Target setelah reorder: {TARGET_MONTHS_POST_REORDER} bulan cadangan + lead time.",
        f"• SKU dengan velocity < {SLOW_DEAD_MAX_VELOCITY}/bulan diklasifikasi 'Slow/Dead', tidak masuk reorder rule.",
    ]
    for line in method_lines:
        c = ws.cell(row=row_cur, column=1, value=line)
        c.font = NORMAL_FONT
        ws.merge_cells(start_row=row_cur, start_column=1, end_row=row_cur, end_column=14)
        row_cur += 1
    row_cur += 1

    ws.cell(row=row_cur, column=1, value="RINGKASAN STATUS").font = TITLE_FONT
    row_cur += 1
    status_order = [
        ("🔴 STOCKOUT", "stockout", RED_FILL),
        ("🔴 Reorder URGENT", "urgent", RED_FILL),
        ("🟠 Reorder Now", "now", ORANGE_FILL),
        ("🟡 Reorder Soon", "soon", YELLOW_FILL),
        ("🟢 Healthy", "healthy", GREEN_FILL),
        ("🔵 Overstock", "overstock", BLUE_FILL),
        ("💤 Slow/Dead", "slow_dead", LIGHT_FILL),
    ]
    for label, key, fill in status_order:
        n = len(reorder_tables[key])
        lbl = ws.cell(row=row_cur, column=1, value=label)
        lbl.font = NORMAL_FONT
        lbl.fill = fill
        val = ws.cell(row=row_cur, column=2, value=n)
        val.font = BOLD_FONT
        val.number_format = FMT_NUM
        row_cur += 1
    row_cur += 2

    buckets = [
        ("A. 🔴 STOCKOUT — SUDAH HABIS, KEHILANGAN SALES", "stockout", RED_FILL),
        ("B. 🔴 REORDER URGENT — Sisa < ROP × 0.7, RESIKO STOCKOUT SEBELUM BARANG DATANG",
         "urgent", RED_FILL),
        (f"C. 🟠 REORDER NOW — Sisa < ROP, BISA REORDER MINGGU INI",
         "now", ORANGE_FILL),
        (f"D. 🟡 REORDER SOON — Sisa < ROP × {ROP_SOON_RATIO}, MULAI SIAP-SIAP",
         "soon", YELLOW_FILL),
        (f"E. 🔵 OVERSTOCK — Sisa > {OVERSTOCK_MONTHS:.0f} bulan cadangan, STOP REORDER",
         "overstock", BLUE_FILL),
    ]
    for title, key, fill in buckets:
        row_cur = _write_reorder_bucket(ws, row_cur, title, reorder_tables[key], fill)


def _write_reorder_full(ws, reorder_full):
    ws["A1"] = "DATA LENGKAP REORDER PER SKU"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:T1")
    headers = ["SKU", "Status", "Sisa Stok", "V 3mo", "V 6mo", "V 12mo", "V 24mo",
               "Vel. Pakai", "Basis", "CV", "Volatility", "Safety ×",
               "Max 1 Order", "Lead (bln)", "Lead Demand",
               "ROP Safety", "ROP Bulk", "ROP Final",
               "Bulan Cover", "Suggest Order"]
    widths = [38, 22, 11, 10, 10, 10, 10, 11, 7, 8, 10, 9,
              12, 11, 12, 12, 12, 12, 12, 13]
    write_headers(ws, 3, headers, widths=widths)
    cols = ["SKU", "status", "sisa_stok", "v3mo", "v6mo", "v12mo", "v24mo",
            "velocity_used", "velocity_basis", "cv", "volatility", "safety_mult",
            "max_single_order", "lead_months", "lead_demand",
            "rop_safety", "rop_bulk", "rop_final", "months_cover", "qty_order_suggest"]
    formats = [None, None, FMT_NUM, FMT_DEC, FMT_DEC, FMT_DEC, FMT_DEC,
               FMT_DEC, None, FMT_DEC, None, FMT_DEC,
               FMT_NUM, FMT_DEC, FMT_DEC,
               FMT_DEC, FMT_DEC, FMT_DEC, FMT_DEC, FMT_NUM]
    out = reorder_full[cols]
    write_data_rows(ws, 4, out, formats=formats)

    status_to_fill = {
        "🔴 STOCKOUT": RED_FILL, "🔴 Reorder URGENT": RED_FILL,
        "🟠 Reorder Now": ORANGE_FILL, "🟡 Reorder Soon": YELLOW_FILL,
        "🟢 Healthy": GREEN_FILL, "🔵 Overstock": BLUE_FILL,
        "💤 Slow/Dead": LIGHT_FILL,
    }
    for r_idx in range(len(out)):
        st = out.iloc[r_idx]["status"]
        fill = status_to_fill.get(st)
        if fill is not None:
            ws.cell(row=4 + r_idx, column=2).fill = fill
    ws.freeze_panes = "B4"


def _write_stock_ledger(ws, ledger_df, today=None) -> None:
    """Per-(SKU, gudang) on-hand stock — the bot's reproduction of BisaRekapBarang,
    computed from the current workbook (arrived beli - non-void jual + ketemu
    - hilang +/- pindah). This is the source of sisa_stok used elsewhere."""
    ws["A1"] = "REKAP STOK PER GUDANG (rekonsiliasi BisaRekapBarang)"
    ws["A1"].font = TITLE_FONT
    n_cols = len(ledger_df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(3, n_cols))
    asof = today.strftime("%d %B %Y") if today is not None else ""
    ws["A2"] = ("Saldo = Σ beli(sudah sampai) − Σ jual(non-void) + ketemu − hilang "
                "± pindah, dari workbook periode berjalan. "
                + (f"Per {asof}." if asof else ""))
    ws["A2"].font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(3, n_cols))

    gudang_cols = [c for c in ledger_df.columns if c not in ("SKU", "Total")]
    ordered = ["SKU"] + gudang_cols + ["Total"]
    df = ledger_df[ordered].sort_values("SKU").copy()

    headers = ["SKU"] + [str(g) for g in gudang_cols] + ["Total Stok"]
    widths = [40] + [18] * len(gudang_cols) + [14]
    write_headers(ws, 4, headers, widths=widths)
    formats = [None] + [FMT_NUM] * len(gudang_cols) + [FMT_NUM]
    write_data_rows(ws, 5, df, formats=formats,
                    negative_highlight_col=len(ordered))
    ws.freeze_panes = "B5"


def write_report(output_path: Path, year: int, jual: pd.DataFrame,
                 sku_agg: pd.DataFrame, tables: dict, sku_no_hpp: list[str],
                 today=None, ledger_df=None) -> None:
    """Write all sheets to a single Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    # Oversold = SKUs the current-workbook ledger shows at a negative on-hand total
    # (sold more than ever purchased) — surface them as a data-quality note.
    oversold = (sorted(ledger_df.loc[ledger_df["Total"] < 0, "SKU"].astype(str))
                if ledger_df is not None and len(ledger_df) and "Total" in ledger_df.columns
                else [])
    _write_summary(ws, year, jual, sku_agg, tables, sku_no_hpp, oversold=oversold)

    _write_diminati(wb.create_sheet("01_Paling_Diminati"), tables["diminati"])
    _write_profit(wb.create_sheet("02_Profit_Tertinggi"), tables["profit"])
    _write_rugi(wb.create_sheet("03_Barang_Rugi"), tables["rugi"])
    _write_borderline(wb.create_sheet("04_Margin_Borderline"), tables["borderline"])
    _write_kandidat(wb.create_sheet("05_Kandidat_Naik_Harga"), tables["kandidat"])
    _write_platform(wb.create_sheet("06_Per_Platform"), tables["platform"], jual)
    _write_full_data(wb.create_sheet("07_Data_Lengkap_per_SKU"), sku_agg)
    if "supplier" in tables:
        _write_supplier_analysis(wb.create_sheet("08_Supplier_Analysis"), tables["supplier"])
    if "reorder" in tables:
        _write_reorder_analysis(wb.create_sheet("09_Reorder_Analysis"),
                                tables["reorder"], today=today)
        _write_reorder_full(wb.create_sheet("10_Reorder_Data_Lengkap"),
                            tables["reorder"]["full"])

    if ledger_df is not None and len(ledger_df) > 0:
        _write_stock_ledger(wb.create_sheet(LEDGER_SHEET_NAME), ledger_df, today=today)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✓ Menulis laporan ke {output_path}")


def write_reorder_standalone(output_path: Path, reorder_tables: dict,
                             today=None, ledger_df=None) -> None:
    """Standalone reorder report (for --reorder CLI flag)."""
    if today is None:
        today = datetime.now()

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Reorder_Summary"

    ws["A1"] = "LAPORAN ANALISA REORDER — ITBISA SHOP"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:F1")
    full = reorder_tables["full"]
    ws["A2"] = (f"Snapshot per {today.strftime('%d %B %Y %H:%M')}  |  "
                f"Total SKU dianalisa: {len(full):,}")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:F2")

    _write_reorder_analysis(wb.create_sheet("01_Reorder_Action"),
                            reorder_tables, today=today)
    _write_reorder_full(wb.create_sheet("02_Reorder_Data_Lengkap"),
                        reorder_tables["full"])

    ws["A4"] = "Ringkasan status di sheet ini, detail di sheet 01 dan 02."
    ws["A4"].font = NORMAL_FONT
    status_order = [
        ("🔴 STOCKOUT", "stockout", RED_FILL),
        ("🔴 Reorder URGENT", "urgent", RED_FILL),
        ("🟠 Reorder Now", "now", ORANGE_FILL),
        ("🟡 Reorder Soon", "soon", YELLOW_FILL),
        ("🟢 Healthy", "healthy", GREEN_FILL),
        ("🔵 Overstock", "overstock", BLUE_FILL),
        ("💤 Slow/Dead", "slow_dead", LIGHT_FILL),
    ]
    for i, (label, key, fill) in enumerate(status_order, start=6):
        n = len(reorder_tables[key])
        lbl = ws.cell(row=i, column=1, value=label)
        lbl.font = BOLD_FONT
        lbl.fill = fill
        val = ws.cell(row=i, column=2, value=n)
        val.font = BOLD_FONT
        val.number_format = FMT_NUM
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    if ledger_df is not None and len(ledger_df) > 0:
        _write_stock_ledger(wb.create_sheet("03_Rekap_Stok_per_Gudang"),
                            ledger_df, today=today)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✓ Menulis laporan ke {output_path}")
