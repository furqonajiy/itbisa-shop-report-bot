"""Per-SKU channel optimizer (`--channel`).

For each SKU, compares the realized NET margin per unit (after the marketplace's
admin fee) across the channels it actually sold on, and recommends the best
channel — flagging SKUs whose volume is concentrated on a channel that nets less
than another established channel.

  net unit/pcs    = (omzet + admin) / qty   per (SKU, channel)
                    (admin = tambahan + kode_unik from BisaJual, stored negative)
  net margin/pcs  = net unit − HPP_WA       (realized P&L cost)

A SKU is flagged 🔁 only when the best established channel (qty ≥ CHANNEL_MIN_QTY)
beats the dominant-volume channel by ≥ CHANNEL_SHIFT_MIN_GAP × HPP per pcs.
Output: output/Analisa_Channel_per_SKU.xlsx. Zero-config (built from BisaJual).
"""
from __future__ import annotations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    BLUE_FILL_COLOR, CHANNEL_MIN_QTY, CHANNEL_SHIFT_MIN_GAP, FMT_NUM, FMT_RP,
    FONT_NAME, GREEN_FILL_COLOR, HEADER_BG_COLOR, HEADER_TEXT_COLOR,
    LIGHT_GRAY_COLOR, TITLE_COLOR, YELLOW_FILL_COLOR,
)

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
GREEN_FILL = PatternFill("solid", start_color=GREEN_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
BLUE_FILL = PatternFill("solid", start_color=BLUE_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)


def analyze_channels(jual: pd.DataFrame, hpp_agg: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Returns (per_sku_recommendation, sku_channel_matrix).

    per_sku: one row per SKU — dominant channel, best-net channel, the per-pcs gap,
    a recommendation, and the potential uplift if the dominant volume moved.
    matrix: long (SKU, channel) rows with qty, net unit, net margin/unit."""
    needed = {"SKU", "akun_penjual", "qty_jual", "omzet"}
    if jual is None or len(jual) == 0 or not needed.issubset(jual.columns):
        return pd.DataFrame(), pd.DataFrame()

    hpp_wa = hpp_agg.set_index("SKU")["hpp_wa"].to_dict() if len(hpp_agg) else {}
    d = jual.copy()
    d["_admin"] = d.get("tambahan", 0).fillna(0) + d.get("kode_unik", 0).fillna(0)
    d["_plat"] = d["akun_penjual"].astype(str).str.strip()

    g = (d.groupby(["SKU", "_plat"])
           .agg(qty=("qty_jual", "sum"), omzet=("omzet", "sum"), admin=("_admin", "sum"))
           .reset_index())
    g = g[g["qty"] > 0].copy()
    g["net_unit"] = (g["omzet"] + g["admin"]) / g["qty"]
    g["hpp"] = g["SKU"].map(hpp_wa)
    g["net_margin_unit"] = g["net_unit"] - g["hpp"]
    g = g.rename(columns={"_plat": "channel"})

    rows = []
    for sku, sub in g.groupby("SKU"):
        hpp = hpp_wa.get(sku)
        total_qty = float(sub["qty"].sum())
        n_channels = int(sub["channel"].nunique())
        dom = sub.loc[sub["qty"].idxmax()]
        dom_ch, dom_net = dom["channel"], dom["net_margin_unit"]
        pool = sub[sub["qty"] >= CHANNEL_MIN_QTY]
        pool = pool if len(pool) else sub
        valid = pool[pool["net_margin_unit"].notna()]
        if len(valid):                       # need ≥1 channel with a computable net margin
            best = valid.loc[valid["net_margin_unit"].idxmax()]
            best_ch, best_net = best["channel"], best["net_margin_unit"]
        else:                                # no HPP → margins unknown
            best_ch, best_net = dom_ch, np.nan
        gap = float(best_net - dom_net) if pd.notna(best_net) and pd.notna(dom_net) else np.nan

        uplift = 0.0
        if hpp is None or pd.isna(hpp) or hpp <= 0:
            rec, saran = "⚪ Tanpa HPP", "HPP belum ada — net margin tak terhitung."
        elif n_channels < 2:
            rec, saran = "⚪ Satu channel", f"Hanya jual di {dom_ch}."
        elif best_ch == dom_ch:
            rec, saran = "🟢 Sudah optimal", f"Volume terbanyak ({dom_ch}) = net margin terbaik."
        elif pd.notna(gap) and gap >= CHANNEL_SHIFT_MIN_GAP * hpp:
            rec = "🔁 Geser channel"
            saran = (f"Net {best_ch} +Rp{gap:,.0f}/pcs vs {dom_ch}. Dorong volume ke {best_ch}.")
            uplift = gap * float(dom["qty"])
        else:
            rec = "🟢 Selisih tipis"
            saran = f"{best_ch} sedikit lebih baik (Rp{gap:,.0f}/pcs) — tidak signifikan."

        rows.append({
            "SKU": sku, "total_qty": total_qty, "n_channels": n_channels,
            "dom_channel": dom_ch, "dom_net_unit": dom_net,
            "best_channel": best_ch, "best_net_unit": best_net,
            "gap_unit": gap, "rec": rec, "saran": saran, "uplift": uplift,
        })

    per_sku = pd.DataFrame(rows)
    if len(per_sku):
        per_sku = per_sku.sort_values(["uplift", "total_qty"], ascending=False).reset_index(drop=True)
    n_shift = int((per_sku["rec"] == "🔁 Geser channel").sum()) if len(per_sku) else 0
    print(f"✓ Channel optimizer: {len(per_sku):,} SKU dianalisa — {n_shift} disarankan geser channel")
    return per_sku, g.sort_values(["SKU", "qty"], ascending=[True, False]).reset_index(drop=True)


# ---------------------------------------------------------------------------
def _style_header(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 28


def write_channel_report(filepath: Path, per_sku: pd.DataFrame,
                         matrix: pd.DataFrame, today) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Ringkasan"
    ws["A1"] = "OPTIMASI CHANNEL PER SKU — JUAL DI MANA PALING UNTUNG"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:D1")
    n_sku = int(len(per_sku))
    n_shift = int((per_sku["rec"] == "🔁 Geser channel").sum()) if len(per_sku) else 0
    total_uplift = float(per_sku["uplift"].sum()) if len(per_sku) else 0.0
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  Net margin/pcs = (omzet + admin) / qty − HPP_WA, "
                f"per channel.  Rekomendasi geser channel kalau net channel terbaik > channel "
                f"volume-terbanyak ≥ {CHANNEL_SHIFT_MIN_GAP*100:.0f}% HPP/pcs.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:D2")

    r = 4
    for label, val, fmt in [
        ("SKU dianalisa", n_sku, FMT_NUM),
        ("SKU disarankan geser channel", n_shift, FMT_NUM),
        ("Potensi tambahan profit (jika volume digeser)", total_uplift, FMT_RP),
    ]:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = BOLD_FONT
        c.number_format = fmt
        r += 1
    ws.cell(row=r, column=1, value="*) Potensi = perkiraan kalau volume channel dominan dijual "
            "di channel net terbaik; tergantung permintaan di channel itu.").font = SUB_FONT

    # Top shift recommendations
    r += 2
    ws.cell(row=r, column=1, value="Top rekomendasi geser channel").font = TITLE_FONT
    r += 1
    _style_header(ws, r, ["SKU", "Dari", "Ke", "Selisih net/pcs", "Potensi"],
                  [34, 14, 14, 16, 16])
    r += 1
    shifts = per_sku[per_sku["rec"] == "🔁 Geser channel"].head(15) if len(per_sku) else per_sku
    for _, p in shifts.iterrows():
        ws.cell(row=r, column=1, value=p["SKU"]).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=p["dom_channel"]).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=p["best_channel"]).font = NORMAL_FONT
        c = ws.cell(row=r, column=4, value=round(p["gap_unit"]))
        c.font = NORMAL_FONT
        c.number_format = FMT_RP
        c2 = ws.cell(row=r, column=5, value=round(p["uplift"]))
        c2.font = NORMAL_FONT
        c2.number_format = FMT_RP
        r += 1

    # --- Sheet 01: full per-SKU recommendation ---
    ws2 = wb.create_sheet("01_Rekomendasi_Channel")
    ws2["A1"] = "REKOMENDASI CHANNEL PER SKU"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:J1")
    headers = ["SKU", "Total Qty", "#Channel", "Channel Dominan", "Net/pcs Dominan",
               "Channel Terbaik", "Net/pcs Terbaik", "Selisih/pcs", "Rekomendasi", "Saran"]
    widths = [34, 11, 9, 16, 15, 16, 15, 13, 18, 52]
    _style_header(ws2, 3, headers, widths)
    if len(per_sku):
        rr = 4
        for _, p in per_sku.iterrows():
            vals = [p["SKU"], round(p["total_qty"]), p["n_channels"], p["dom_channel"],
                    (round(p["dom_net_unit"]) if pd.notna(p["dom_net_unit"]) else None),
                    p["best_channel"],
                    (round(p["best_net_unit"]) if pd.notna(p["best_net_unit"]) else None),
                    (round(p["gap_unit"]) if pd.notna(p["gap_unit"]) else None),
                    p["rec"], p["saran"]]
            fmts = [None, FMT_NUM, FMT_NUM, None, FMT_RP, None, FMT_RP, FMT_RP, None, None]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
                c = ws2.cell(row=rr, column=ci, value=v)
                c.font = NORMAL_FONT
                c.alignment = Alignment(vertical="top", wrap_text=(ci == len(vals)))
                if fmt:
                    c.number_format = fmt
                if rr % 2 == 1:
                    c.fill = LIGHT_FILL
            if str(p["rec"]).startswith("🔁"):
                ws2.cell(row=rr, column=9).fill = YELLOW_FILL
            rr += 1
        ws2.freeze_panes = "A4"

    # --- Sheet 02: SKU × channel detail ---
    ws3 = wb.create_sheet("02_SKU_x_Channel")
    ws3["A1"] = "DETAIL NET MARGIN PER SKU × CHANNEL"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:E1")
    _style_header(ws3, 3, ["SKU", "Channel", "Qty", "Net/pcs", "Net Margin/pcs"],
                  [34, 16, 11, 14, 16])
    if len(matrix):
        rr = 4
        for _, m in matrix.iterrows():
            vals = [m["SKU"], m["channel"], round(m["qty"]),
                    (round(m["net_unit"]) if pd.notna(m["net_unit"]) else None),
                    (round(m["net_margin_unit"]) if pd.notna(m["net_margin_unit"]) else None)]
            fmts = [None, None, FMT_NUM, FMT_RP, FMT_RP]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
                c = ws3.cell(row=rr, column=ci, value=v)
                c.font = NORMAL_FONT
                if fmt:
                    c.number_format = fmt
                if rr % 2 == 1:
                    c.fill = LIGHT_FILL
            rr += 1
        ws3.freeze_panes = "A4"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
