"""Read-only reconciliation of the raw BisaSaldo / BisaFee inputs against what the
generator captures into BisaLaporan.

This module never changes the generated BisaInvoice/BisaJual/BisaRemit/BisaBonus
numbers. It re-reads the raw marketplace exports the same way the processors do,
classifies every BisaSaldo row into one of four buckets (Remit / Bonus /
Withdrawal / Uncaptured) using the exact keyword lists from keywordchecker/, and
writes a separate ``Rekonsiliasi <Marketplace>.xlsx`` per marketplace:

  * Ringkasan           - per-period tie-out (net = Remit + Bonus + Withdrawal +
                          Uncaptured) with an Uncaptured/fee flag.
  * Saldo Tidak Tercatat- every uncaptured BisaSaldo row, with the reason it
                          fell through (this is where e.g. the SPinjam loan rows
                          and Bukalapak's "Need to Check" rows show up).
  * BisaFee Tidak Cocok - (Shopee) fee rows that find no matching remit, and
                          remit invoices that get no fee breakdown.
"""
import glob
import logging
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from keywordchecker import bukalapak as kw_bukalapak
from keywordchecker import shopee as kw_shopee
from keywordchecker import tokopedia as kw_tokopedia
from utility.constant import MARKETPLACE_FOLDERS, get_data_dir, get_reports_dir

# Bucket labels (Bahasa Indonesia, user-facing).
_REMIT = 'Remit'
_BONUS = 'Bonus'
_WITHDRAWAL = 'Penarikan/Transfer'
_UNCAPTURED = 'Tidak Tercatat'

_HEADER_FILL = PatternFill('solid', fgColor='4472C4')
_FLAG_FILL = PatternFill('solid', fgColor='FFC7CE')   # red-ish: needs attention
_OK_FILL = PatternFill('solid', fgColor='C6EFCE')     # green: clean


def _contains_any(text, keywords):
    """Boolean Series: does each row's text contain any of the keywords?

    Mirrors the generators' `str.contains('|'.join(...))` (regex, unescaped).
    """
    if not keywords:
        return pd.Series(False, index=text.index)
    return text.fillna('').str.contains('|'.join(keywords), regex=True)


# --- per-marketplace BisaSaldo adapters -------------------------------------

def _read_shopee_v2(path):
    df = pd.read_csv(path, skiprows=6)
    return pd.DataFrame({
        'Tanggal': df.get('Tanggal'),
        'Deskripsi': df['Deskripsi'].fillna('').astype(str),
        'Nominal': pd.to_numeric(df['Jumlah Dana'], errors='coerce').fillna(0),
    })


def _read_shopee_v3(path):
    df = pd.read_excel(path, skiprows=17)
    no_pesanan = df['No. Pesanan'].astype(str)
    out = pd.DataFrame({
        'Tanggal': df.get('Tanggal Transaksi'),
        'Deskripsi': df['Deskripsi'].fillna('').astype(str),
        'Nominal': pd.to_numeric(df['Jumlah'], errors='coerce').fillna(0),
    })
    # v3 selects remit rows by a real order id (No. Pesanan without a '-').
    out['_remit_filter'] = ~no_pesanan.str.contains('-')
    out['_invoice'] = no_pesanan
    return out


def _read_tokopedia_v2(path):
    df = pd.read_excel(path, skiprows=6)
    return pd.DataFrame({
        'Tanggal': df.get('Date'),
        'Deskripsi': df['Description'].fillna('').astype(str),
        'Nominal': pd.to_numeric(df['Nominal (Rp)'], errors='coerce').fillna(0),
    })


def _read_bukalapak_v2(path):
    df = pd.read_csv(path)
    return pd.DataFrame({
        'Tanggal': df.get('Waktu'),
        'Deskripsi': df['Keterangan'].fillna('').astype(str),
        'Nominal': pd.to_numeric(df['Mutasi'], errors='coerce').fillna(0),
    })


def _components(kw):
    """Union of the keyword lists whose rows carry a remit amount."""
    return (list(getattr(kw, 'VALID_NOMINAL_REMIT_KEYWORD', []))
            + list(getattr(kw, 'VALID_POTONGAN_PEMBAYARAN_KEYWORD', []))
            + list(getattr(kw, 'VALID_KEUNTUNGAN_TAMBAHAN_KEYWORD', []))
            + list(getattr(kw, 'VALID_KERUGIAN_TAMBAHAN_KEYWORD', [])))


def _not_used(kw):
    """The 'Not Used' keywords = VALID_SALDO_KEYWORD minus every captured list."""
    captured = set(_components(kw)) | set(getattr(kw, 'VALID_BONUS_KEYWORD', []))
    return [k for k in kw.VALID_SALDO_KEYWORD if k not in captured]


def _keyword_categories(kw):
    """(category label, keyword list) pairs, in the order a row is matched."""
    return [
        ('Nominal Remit', list(getattr(kw, 'VALID_NOMINAL_REMIT_KEYWORD', []))),
        ('Potongan Pembayaran', list(getattr(kw, 'VALID_POTONGAN_PEMBAYARAN_KEYWORD', []))),
        ('Keuntungan Tambahan', list(getattr(kw, 'VALID_KEUNTUNGAN_TAMBAHAN_KEYWORD', []))),
        ('Kerugian Tambahan', list(getattr(kw, 'VALID_KERUGIAN_TAMBAHAN_KEYWORD', []))),
        ('Bonus', list(getattr(kw, 'VALID_BONUS_KEYWORD', []))),
        ('Tidak Digunakan', _not_used(kw)),
    ]


def _annotate_keyword(df, cfg):
    """Add the matched keyword + its category to each row (first match wins)."""
    desc = df['Deskripsi'].fillna('')
    keyword = pd.Series('', index=df.index)
    kategori = pd.Series('', index=df.index)
    for label, kws in _keyword_categories(cfg['kw']):
        for kw in kws:
            hit = (keyword == '') & desc.str.contains(kw, regex=True)
            keyword[hit] = kw
            kategori[hit] = label
    kategori[keyword == ''] = 'Tidak Cocok Keyword'
    df['Keyword'] = keyword
    df['Kategori'] = kategori
    return df


# Each marketplace: its reports folder, the keyword module, whether a BisaBonus
# sheet is actually generated, how it decides a row is a remit row, and the
# (token -> reader) BisaSaldo sources.
_MARKETPLACES = {
    'Shopee': {
        'kw': kw_shopee, 'has_bonus': True,
        'remit_rule': lambda df: _contains_any(df['Deskripsi'], ['#', 'Penambahan dana']),
        'sources': [('BisaSaldo v2 Shopee', _read_shopee_v2),
                    ('BisaSaldo v3 Shopee', _read_shopee_v3)],
    },
    'Tokopedia': {
        'kw': kw_tokopedia, 'has_bonus': True,
        'remit_rule': lambda df: _contains_any(df['Deskripsi'], ['INV']),
        'sources': [('BisaSaldo v2 Tokopedia', _read_tokopedia_v2)],
    },
    'Bukalapak': {
        'kw': kw_bukalapak, 'has_bonus': False,  # no BisaBonus generator exists
        'remit_rule': lambda df: _contains_any(df['Deskripsi'], ['#']),
        'sources': [('BisaSaldo v2 Bukalapak', _read_bukalapak_v2)],
    },
}


def _classify(df, cfg):
    """Tag each BisaSaldo row with a bucket and (for uncaptured rows) a reason."""
    kw = cfg['kw']
    desc = df['Deskripsi']
    comp = _contains_any(desc, _components(kw))
    bonus = _contains_any(desc, getattr(kw, 'VALID_BONUS_KEYWORD', []))
    notused = _contains_any(desc, _not_used(kw))
    remit_filter = df['_remit_filter'] if '_remit_filter' in df.columns else cfg['remit_rule'](df)

    bucket = pd.Series(_UNCAPTURED, index=df.index)
    reason = pd.Series('', index=df.index)

    is_remit = remit_filter & comp
    is_bonus = bonus & cfg['has_bonus']
    bucket[is_remit] = _REMIT
    bucket[~is_remit & is_bonus] = _BONUS
    bucket[(bucket == _UNCAPTURED) & notused] = _WITHDRAWAL

    # Reasons for whatever is still uncaptured.
    left = bucket == _UNCAPTURED
    reason[left & comp & ~remit_filter] = 'Cocok keyword remit tetapi dikecualikan filter invoice'
    reason[left & remit_filter & ~comp] = 'Baris invoice tetapi tidak ada keyword remit'
    reason[left & bonus & ~cfg['has_bonus']] = 'Keyword bonus tetapi BisaBonus tidak dibuat'
    reason[left & (reason == '')] = 'Tidak cocok kategori manapun'

    df = df.copy()
    df['Bucket'] = bucket
    df['Alasan'] = reason
    return _annotate_keyword(df, cfg)


def _period(path):
    base = os.path.basename(path)
    return base.replace('.csv', '').replace('.xlsx', '')


def _saldo_files(token):
    files = glob.glob(os.path.join(get_data_dir(), '**', '*'), recursive=True)
    return sorted(f for f in files
                  if token in os.path.basename(f) and '~' not in os.path.basename(f))


# --- Shopee BisaFee <-> remit match -----------------------------------------

def _shopee_fee_invoices():
    """Invoice -> Total Penghasilan from every Shopee BisaFee Income sheet."""
    rows = []
    for token in ('BisaFee v2 Shopee', 'BisaFee v3 Shopee'):
        for path in _saldo_files(token):
            try:
                fdf = pd.read_excel(path, sheet_name='Income', skiprows=5)
            except Exception as err:  # noqa: BLE001 - tolerate odd fee files
                logging.debug("Skip fee file %s: %s", path, err)
                continue
            if 'No. Pesanan' not in fdf.columns or 'Total Penghasilan' not in fdf.columns:
                continue
            rows.append(pd.DataFrame({
                'Invoice': fdf['No. Pesanan'].astype(str),
                'Total Penghasilan': pd.to_numeric(fdf['Total Penghasilan'], errors='coerce').fillna(0),
                'File': os.path.basename(path),
            }))
    if not rows:
        return None
    fee = pd.concat(rows, ignore_index=True)
    # Fee files overlap (a yearly v2 file and a monthly v3 file can both list the
    # same order). Count each (Invoice, Total Penghasilan) once so the same fee
    # isn't summed twice.
    return fee.drop_duplicates(subset=['Invoice', 'Total Penghasilan'])


def _shopee_remit_amounts(classified_frames):
    """Invoice -> summed Nominal Remit from the classified Shopee saldo rows."""
    nominal_kw = kw_shopee.VALID_NOMINAL_REMIT_KEYWORD
    parts = []
    for df in classified_frames:
        if '_invoice' in df.columns:
            inv = df['_invoice']
        else:
            inv = df['Deskripsi'].str.extract(r'#(\S+)')[0].str.replace('.', '', regex=False)
        is_nominal = (df['Bucket'] == _REMIT) & _contains_any(df['Deskripsi'], nominal_kw)
        sub = pd.DataFrame({'Invoice': inv, 'Nominal Remit': df['Nominal']})[is_nominal]
        parts.append(sub)
    if not parts:
        return pd.DataFrame(columns=['Invoice', 'Nominal Remit'])
    allp = pd.concat(parts, ignore_index=True).dropna(subset=['Invoice'])
    return allp.groupby('Invoice', as_index=False)['Nominal Remit'].sum()


def _build_fee_mismatch(fee, remit):
    """Rows describing fee<->remit reconciliation problems (Shopee)."""
    if fee is None:
        return pd.DataFrame()
    fee_g = fee.groupby('Invoice', as_index=False)['Total Penghasilan'].sum()
    merged = fee_g.merge(remit, on='Invoice', how='outer', indicator=True)
    out = []
    for _, r in merged.iterrows():
        inv = r['Invoice']
        tp = r['Total Penghasilan']
        nr = r['Nominal Remit']
        if r['_merge'] == 'left_only':
            out.append((inv, tp, None, 'Fee tanpa remit (fee hilang dari laporan)'))
        elif r['_merge'] == 'right_only':
            out.append((inv, None, nr, 'Remit tanpa fee (tidak ada rincian biaya)'))
        elif abs((tp or 0) - (nr or 0)) > 0.5:
            out.append((inv, tp, nr, 'Nominal beda -> join (Invoice, Nominal Remit) gagal, fee jadi 0'))
    return pd.DataFrame(out, columns=['Invoice', 'Total Penghasilan (Fee)', 'Nominal Remit', 'Masalah'])


def _shopee_saldo_vs_fee(fee, remit):
    """Per-invoice side-by-side: BisaSaldo Nominal Remit vs BisaFee Total Penghasilan.

    Lists EVERY invoice (not only mismatches) so the remit can be checked against
    both sources at a glance.
    """
    fee_g = (fee.groupby('Invoice', as_index=False)['Total Penghasilan'].sum()
             if fee is not None else pd.DataFrame(columns=['Invoice', 'Total Penghasilan']))
    m = remit.merge(fee_g, on='Invoice', how='outer', indicator=True)

    status = []
    for _, r in m.iterrows():
        if r['_merge'] == 'left_only':
            status.append('Hanya BisaSaldo')
        elif r['_merge'] == 'right_only':
            status.append('Hanya BisaFee')
        elif abs((r['Nominal Remit'] or 0) - (r['Total Penghasilan'] or 0)) <= 0.5:
            status.append('Cocok')
        else:
            status.append('Beda')

    out = pd.DataFrame({
        'Invoice': m['Invoice'],
        'Remit dari BisaSaldo': m['Nominal Remit'],
        'Total Penghasilan (BisaFee)': m['Total Penghasilan'],
        'Selisih (Saldo - Fee)': m['Nominal Remit'].fillna(0) - m['Total Penghasilan'].fillna(0),
        'Status': status,
    })
    for col in ['Remit dari BisaSaldo', 'Total Penghasilan (BisaFee)', 'Selisih (Saldo - Fee)']:
        out[col] = out[col].round().astype('Int64')
    rank = {'Beda': 0, 'Hanya BisaSaldo': 1, 'Hanya BisaFee': 2, 'Cocok': 3}
    out['_r'] = out['Status'].map(rank)
    return out.sort_values(['_r', 'Invoice']).drop(columns='_r').reset_index(drop=True)


# --- Shopee Omzet (BisaJual) vs BisaFee / BisaSaldo per invoice ---------------

def _price(series):
    return pd.to_numeric(series.astype(str).str.replace('Rp ', '', regex=False)
                         .str.replace('.', '', regex=False), errors='coerce').fillna(0)


def _shopee_omzet():
    """Invoice -> Omzet, re-derived from raw BisaTransaksi (same status filter as BisaJual)."""
    parts = []
    for path in _saldo_files('BisaTransaksi v2 Shopee'):
        df = pd.read_excel(path, dtype={'Harga Setelah Diskon': str, 'Alasan Pembatalan': str})
        keep = (~df['Status Pesanan'].astype(str).str.contains('Batal|Dibatalkan')
                | df['Alasan Pembatalan'].astype(str).str.contains('Paket hilang', na=False))
        df = df[keep]
        parts.append(pd.DataFrame({
            'Invoice': df['No. Pesanan'].astype(str),
            'Omzet': pd.to_numeric(df['Jumlah'], errors='coerce').fillna(0) * _price(df['Harga Setelah Diskon'])}))
    for path in _saldo_files('BisaTransaksi v3 Shopee'):
        df = pd.read_excel(path, dtype={'Harga Setelah Diskon': str, 'Alasan Pembatalan': str})
        j = pd.to_numeric(df['Jumlah'], errors='coerce')
        rq = pd.to_numeric(df.get('Returned quantity'), errors='coerce')
        sv = ['Belum Bayar', 'Batal', 'Dibatalkan', 'Selesai']
        keep = (~df['Status Pesanan'].astype(str).str.contains('|'.join(sv))
                | df['Alasan Pembatalan'].astype(str).str.contains('Paket hilang', na=False)
                | (df['Status Pesanan'].astype(str).str.contains('Selesai') & (j != rq)))
        df = df[keep]
        parts.append(pd.DataFrame({
            'Invoice': df['No. Pesanan'].astype(str),
            'Omzet': pd.to_numeric(df['Jumlah'], errors='coerce').fillna(0) * _price(df['Harga Setelah Diskon'])}))
    if not parts:
        return pd.DataFrame(columns=['Invoice', 'Omzet'])
    return pd.concat(parts, ignore_index=True).groupby('Invoice', as_index=False)['Omzet'].sum()


def _shopee_fee_detail():
    """Invoice -> Total Penghasilan and the Kerugian (refund + fees) from BisaFee."""
    rows = []
    for token in ('BisaFee v2 Shopee', 'BisaFee v3 Shopee'):
        for path in _saldo_files(token):
            try:
                fdf = pd.read_excel(path, sheet_name='Income', skiprows=5)
            except Exception as err:  # noqa: BLE001
                logging.debug("Skip fee file %s: %s", path, err)
                continue
            if 'No. Pesanan' not in fdf.columns:
                continue

            def col(name):
                return pd.to_numeric(fdf[name], errors='coerce').fillna(0) if name in fdf.columns else 0
            kerugian = -(col('Jumlah Pengembalian Dana ke Pembeli') + col('Biaya Transaksi')
                         + col('Biaya Administrasi') + col('Biaya Layanan (termasuk PPN 11%)'))
            rows.append(pd.DataFrame({
                'Invoice': fdf['No. Pesanan'].astype(str),
                'Total Penghasilan': col('Total Penghasilan'),
                'Kerugian (Fee)': kerugian}))
    if not rows:
        return None
    fee = pd.concat(rows, ignore_index=True).drop_duplicates(subset=['Invoice', 'Total Penghasilan'])
    return fee.groupby('Invoice', as_index=False).agg({'Total Penghasilan': 'sum', 'Kerugian (Fee)': 'sum'})


def _shopee_income_all(classified_frames):
    """Invoice -> the full BisaSaldo net (every line for that invoice)."""
    parts = []
    for df in classified_frames:
        inv = (df['_invoice'] if '_invoice' in df.columns
               else df['Deskripsi'].str.extract(r'#(\S+)')[0].str.replace('.', '', regex=False))
        parts.append(pd.DataFrame({'Invoice': inv, 'Penghasilan': df['Nominal']}))
    if not parts:
        return pd.DataFrame(columns=['Invoice', 'Penghasilan'])
    allp = pd.concat(parts, ignore_index=True).dropna(subset=['Invoice'])
    return allp.groupby('Invoice', as_index=False)['Penghasilan'].sum()


def _build_omzet_vs_fee(classified_frames):
    """Per-invoice: Omzet (BisaJual) vs real income (BisaSaldo) vs BisaFee.

    Flags orders booked with Omzet that did not actually settle - including
    returns whose loss only lives in BisaFee (so it is dropped from BisaRemit).
    """
    omz = _shopee_omzet()
    inc = _shopee_income_all(classified_frames)
    fee = _shopee_fee_detail()
    if fee is None:
        fee = pd.DataFrame(columns=['Invoice', 'Total Penghasilan', 'Kerugian (Fee)'])

    fee_invoices = set(fee['Invoice']) if fee is not None and len(fee) else set()
    m = omz.merge(inc, on='Invoice', how='outer').merge(fee, on='Invoice', how='outer')
    for c in ['Omzet', 'Penghasilan', 'Total Penghasilan', 'Kerugian (Fee)']:
        m[c] = pd.to_numeric(m[c], errors='coerce').fillna(0)

    def status(r):
        om, pen, ker = r['Omzet'], r['Penghasilan'], r['Kerugian (Fee)']
        in_fee = r['Invoice'] in fee_invoices
        if om > 0 and pen <= 0:
            if in_fee and ker >= om * 0.99:
                return 'Retur - rugi = omzet (dari BisaFee)'   # confirmed loss
            if in_fee:
                return 'Omzet tidak settle (ada fee, cek)'     # fee exists but nothing received
            return 'Belum ada penghasilan (pending / cek)'     # no saldo & no fee yet
        if om == 0 and pen > 0:
            return 'Penghasilan tanpa Omzet'
        return 'OK'
    m['Status'] = m.apply(status, axis=1)

    flagged = m[m['Status'] != 'OK'].copy()
    for c in ['Omzet', 'Penghasilan', 'Total Penghasilan', 'Kerugian (Fee)']:
        flagged[c] = flagged[c].round().astype('int64')
    flagged = flagged.rename(columns={'Penghasilan': 'Penghasilan (BisaSaldo)',
                                      'Total Penghasilan': 'Total Penghasilan (Fee)'})
    rank = {'Retur - rugi = omzet (dari BisaFee)': 0, 'Omzet tidak settle (ada fee, cek)': 1,
            'Belum ada penghasilan (pending / cek)': 2, 'Penghasilan tanpa Omzet': 3}
    flagged['_r'] = flagged['Status'].map(rank).fillna(4)
    flagged = flagged.sort_values(['_r', 'Omzet'], ascending=[True, False]).drop(columns='_r')
    return flagged[['Invoice', 'Omzet', 'Penghasilan (BisaSaldo)',
                    'Total Penghasilan (Fee)', 'Kerugian (Fee)', 'Status']].reset_index(drop=True)


# --- workbook writer ---------------------------------------------------------

_CATEGORY_ORDER = ['Nominal Remit', 'Potongan Pembayaran', 'Keuntungan Tambahan',
                   'Kerugian Tambahan', 'Bonus', 'Tidak Digunakan', 'Tidak Cocok Keyword']


def _by_description(allrows):
    """Roll every BisaSaldo row up to its matched keyword: count, total, example."""
    g = (allrows.groupby(['Kategori', 'Keyword', 'Bucket'], dropna=False)
         .agg(Baris=('Nominal', 'size'),
              Total=('Nominal', 'sum'),
              Contoh=('Deskripsi', 'first'))
         .reset_index())
    g['Total'] = g['Total'].round().astype('int64')
    g['_ord'] = g['Kategori'].map({c: i for i, c in enumerate(_CATEGORY_ORDER)}).fillna(99)
    g = g.sort_values(['_ord', 'Keyword']).drop(columns='_ord').reset_index(drop=True)
    g.columns = ['Kategori', 'Keyword (Deskripsi)', 'Bucket', 'Jumlah Baris',
                 'Total Nominal', 'Contoh Deskripsi']
    return g


def _style_header(sheet, widths):
    for i, w in enumerate(widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = w
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.freeze_panes = 'A2'


def _write_workbook(path, summary, by_desc, detail, uncaptured, fee_mismatch,
                    saldo_vs_fee=None, omzet_vs_fee=None):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='Ringkasan', index=False)
        by_desc.to_excel(writer, sheet_name='Rincian per Deskripsi', index=False)
        detail.to_excel(writer, sheet_name='Rincian Saldo', index=False)
        uncaptured.to_excel(writer, sheet_name='Saldo Tidak Tercatat', index=False)
        if omzet_vs_fee is not None:
            omzet_vs_fee.to_excel(writer, sheet_name='Cek Omzet vs Fee', index=False)
        if saldo_vs_fee is not None:
            saldo_vs_fee.to_excel(writer, sheet_name='Cek Remit Saldo vs Fee', index=False)
        if fee_mismatch is not None:
            fee_mismatch.to_excel(writer, sheet_name='BisaFee Tidak Cocok', index=False)

        _style_header(writer.book['Ringkasan'], [34, 10, 16, 14, 16, 16, 14, 16, 8])
        _style_header(writer.book['Rincian per Deskripsi'], [20, 44, 18, 12, 16, 70])
        _style_header(writer.book['Rincian Saldo'], [34, 20, 64, 16, 20, 18])
        _style_header(writer.book['Saldo Tidak Tercatat'], [34, 20, 64, 16, 52])
        if omzet_vs_fee is not None:
            _style_header(writer.book['Cek Omzet vs Fee'], [18, 14, 20, 20, 16, 34])
            sheet = writer.book['Cek Omzet vs Fee']
            status_col = omzet_vs_fee.columns.get_loc('Status') + 1
            for row in range(2, len(omzet_vs_fee) + 2):
                sheet.cell(row=row, column=status_col).fill = _FLAG_FILL
        if saldo_vs_fee is not None:
            _style_header(writer.book['Cek Remit Saldo vs Fee'], [40, 20, 24, 20, 18])
            sheet = writer.book['Cek Remit Saldo vs Fee']
            status_col = saldo_vs_fee.columns.get_loc('Status') + 1
            for row in range(2, len(saldo_vs_fee) + 2):
                cell = sheet.cell(row=row, column=status_col)
                cell.fill = _OK_FILL if cell.value == 'Cocok' else _FLAG_FILL
        if fee_mismatch is not None:
            _style_header(writer.book['BisaFee Tidak Cocok'], [40, 20, 16, 48])

        # Flag periods that need attention in the Ringkasan.
        sheet = writer.book['Ringkasan']
        flag_col = summary.columns.get_loc('Perlu Dicek') + 1
        for row in range(2, len(summary) + 2):
            cell = sheet.cell(row=row, column=flag_col)
            cell.fill = _FLAG_FILL if cell.value == 'YA' else _OK_FILL

        # Highlight the uncaptured ('Tidak Tercatat') keyword groups in red.
        sheet = writer.book['Rincian per Deskripsi']
        bucket_col = by_desc.columns.get_loc('Bucket') + 1
        for row in range(2, len(by_desc) + 2):
            cell = sheet.cell(row=row, column=bucket_col)
            if cell.value == _UNCAPTURED:
                cell.fill = _FLAG_FILL


def generate_reconciliation(marketplaces=None):
    """Write a Rekonsiliasi <Marketplace>.xlsx for each selected marketplace."""
    names = marketplaces or list(_MARKETPLACES.keys())
    for name in names:
        cfg = _MARKETPLACES.get(name)
        if cfg is None:
            logging.info("Lewati rekonsiliasi %s (tidak ada BisaSaldo)", name)
            continue

        summary_rows = []
        uncaptured_rows = []
        classified_frames = []
        for token, reader in cfg['sources']:
            for path in _saldo_files(token):
                try:
                    raw = reader(path)
                except Exception as err:  # noqa: BLE001 - report, don't crash the run
                    logging.warning("Gagal membaca %s: %s", path, err)
                    continue
                if raw.empty:
                    continue
                period = _period(path)
                cls = _classify(raw, cfg)
                cls['Periode'] = period
                classified_frames.append(cls)

                net = cls['Nominal'].sum()
                by = cls.groupby('Bucket')['Nominal'].agg(['count', 'sum'])
                def amt(b):
                    return float(by.loc[b, 'sum']) if b in by.index else 0.0
                def cnt(b):
                    return int(by.loc[b, 'count']) if b in by.index else 0
                unc_amt = amt(_UNCAPTURED)
                summary_rows.append({
                    'Periode': period,
                    'Baris': len(cls),
                    'Net Saldo': round(net),
                    'Remit': round(amt(_REMIT)),
                    'Bonus': round(amt(_BONUS)),
                    'Penarikan/Transfer': round(amt(_WITHDRAWAL)),
                    'Tidak Tercatat': round(unc_amt),
                    'Baris Tidak Tercatat': cnt(_UNCAPTURED),
                    'Perlu Dicek': 'YA' if cnt(_UNCAPTURED) > 0 else '-',
                })
                unc = cls[cls['Bucket'] == _UNCAPTURED]
                for _, r in unc.iterrows():
                    uncaptured_rows.append({
                        'Periode': period,
                        'Tanggal': r.get('Tanggal'),
                        'Deskripsi': r['Deskripsi'],
                        'Nominal': round(r['Nominal']),
                        'Alasan': r['Alasan'],
                    })

        if not summary_rows:
            logging.info("Tidak ada BisaSaldo %s untuk direkonsiliasi", name)
            continue

        summary = pd.DataFrame(summary_rows).sort_values('Periode').reset_index(drop=True)
        uncaptured = (pd.DataFrame(uncaptured_rows)
                      if uncaptured_rows else
                      pd.DataFrame(columns=['Periode', 'Tanggal', 'Deskripsi', 'Nominal', 'Alasan']))

        allrows = pd.concat(classified_frames, ignore_index=True)
        by_desc = _by_description(allrows)
        detail = (allrows[['Periode', 'Tanggal', 'Deskripsi', 'Nominal', 'Kategori', 'Bucket']]
                  .copy())
        detail['Nominal'] = detail['Nominal'].round().astype('int64')
        detail = detail.sort_values(['Periode', 'Kategori']).reset_index(drop=True)

        fee_mismatch = None
        saldo_vs_fee = None
        omzet_vs_fee = None
        if name == 'Shopee':
            fee = _shopee_fee_invoices()
            remit = _shopee_remit_amounts(classified_frames)
            fee_mismatch = _build_fee_mismatch(fee, remit)
            saldo_vs_fee = _shopee_saldo_vs_fee(fee, remit)
            omzet_vs_fee = _build_omzet_vs_fee(classified_frames)

        folder = os.path.join(get_reports_dir(), MARKETPLACE_FOLDERS[name])
        os.makedirs(folder, exist_ok=True)
        out_path = os.path.join(folder, 'Rekonsiliasi {0}.xlsx'.format(name))
        _write_workbook(out_path, summary, by_desc, detail, uncaptured, fee_mismatch,
                        saldo_vs_fee, omzet_vs_fee)

        flagged = int((summary['Perlu Dicek'] == 'YA').sum())
        logging.info("Rekonsiliasi %s -> %s (%d periode, %d perlu dicek, %d baris tidak tercatat)",
                     name, os.path.basename(out_path), len(summary), flagged, len(uncaptured))
