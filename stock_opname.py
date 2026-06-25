"""Stock-opname reconciliation → BisaHilang.

Compares a physical stock count (``data/stock_opname.xlsx``) against the bot's
computed on-hand (the current-workbook ledger, the same ``sisa_stok`` used
everywhere) and produces the **BisaHilang** adjustment rows: per SKU,
**Banyak Hilang** when the book stock is higher than the physical count
(shrinkage) or **Banyak Ketemu** when the physical count is higher (found).

The found/lost value (``Nilai Ketemu`` / ``Nilai Hilang``) is ``qty x HPP``
where HPP is the per-SKU cost from ``STOCK_OPNAME_VALUE_BASIS`` — by default
``hpp_wa`` (weighted-average cost), the realized inventory cost and the standard
basis for a stock write-off. FIFO is not supported (no purchase layers are kept).

Paste the generated ``BisaHilang`` sheet into the Google Sheets BisaHilang tab,
re-export ``BisaStok``, and the bot's ``sisa_stok`` lines up with the count.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    BISAHILANG_OUTPUT_FILENAME, COL_BH_NILAI_HILANG, COL_BH_NILAI_KETEMU,
    COL_BH_NOTE, COL_BH_TANGGAL, COL_HILANG_GUDANG, COL_HILANG_HILANG,
    COL_HILANG_KETEMU, COL_HILANG_SKU, COL_SO_FISIK, COL_SO_GUDANG, COL_SO_NOTE,
    COL_SO_SKU, COL_SO_TANGGAL, HEADER_BG_COLOR, HEADER_TEXT_COLOR,
    STOCK_OPNAME_DEFAULT_GUDANG, STOCK_OPNAME_SHEET, STOCK_OPNAME_VALUE_BASIS,
)

_HEADER_FONT = Font(name="Arial", bold=True, color=HEADER_TEXT_COLOR, size=10)
_HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG_COLOR)
_NORMAL_FONT = Font(name="Arial", size=10)
_BOLD_FONT = Font(name="Arial", bold=True, size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)
_FMT_RP = '#,##0'


def create_template(filepath: Path) -> None:
    """Create data/stock_opname.xlsx with headers, a worked example, and a guide."""
    wb = Workbook()
    ws = wb.active
    ws.title = STOCK_OPNAME_SHEET
    headers = [COL_SO_SKU, COL_SO_FISIK, COL_SO_GUDANG, COL_SO_TANGGAL, COL_SO_NOTE]
    widths = [42, 12, 18, 16, 40]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 28
    example = ["ITBISA-SERVO-SG90-180DEG", 1, "KYAI KONTONG", None,
               "contoh — hitung fisik 1 pcs"]
    for i, v in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=v).font = _NORMAL_FONT

    guide = wb.create_sheet("Cara_Pakai")
    guide.column_dimensions["A"].width = 100
    lines = [
        "Cara isi sheet 'StockOpname' (satu baris = satu SKU yang dihitung fisik):",
        f"• {COL_SO_SKU}: persis sama dengan SKU di Stok/Jual.",
        f"• {COL_SO_FISIK}: jumlah fisik hasil hitung (total semua channel).",
        f"• {COL_SO_GUDANG}: gudang tempat mencatat selisih (mis. KYAI KONTONG). "
        "Kosongkan → otomatis pakai gudang dominan SKU itu di ledger.",
        f"• {COL_SO_TANGGAL}: tanggal pengecekan (kosong → hari ini).",
        f"• {COL_SO_NOTE}: catatan bebas (kosong → 'Stock opname <tgl>').",
        "",
        "Lalu jalankan:  python main.py --stock-opname",
        "Hasil: output/" + BISAHILANG_OUTPUT_FILENAME + " — sheet 'BisaHilang' siap "
        "di-copy ke tab BisaHilang di Google Sheets.",
        "",
        "Selisih = stok buku (sisa_stok bot) − stok fisik. >0 = Hilang, <0 = Ketemu.",
        "Nilai = qty × HPP rata-rata tertimbang (HPP_WA) — biaya realisasi persediaan.",
    ]
    for j, n in enumerate(lines, start=1):
        guide.cell(row=j, column=1, value=n).font = _BOLD_FONT if j == 1 else _NORMAL_FONT

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Template stock-opname dibuat: {filepath}")


def load_stock_opname(filepath: Path) -> pd.DataFrame:
    """Read the physical-count template; normalize to columns SKU/fisik/gudang/tanggal/note."""
    df = pd.read_excel(filepath, sheet_name=STOCK_OPNAME_SHEET)
    df = df.rename(columns={COL_SO_SKU: "SKU", COL_SO_FISIK: "fisik",
                            COL_SO_GUDANG: "gudang", COL_SO_TANGGAL: "tanggal",
                            COL_SO_NOTE: "note"})
    for c in ("gudang", "tanggal", "note"):
        if c not in df.columns:
            df[c] = None
    df = df[df["SKU"].notna()].copy()
    df["SKU"] = df["SKU"].astype(str).str.upper().str.strip()
    df["fisik"] = pd.to_numeric(df["fisik"], errors="coerce")
    df = df[df["fisik"].notna()].copy()
    df["tanggal"] = pd.to_datetime(df["tanggal"], errors="coerce")
    return df[["SKU", "fisik", "gudang", "tanggal", "note"]]


def _dominant_gudang(ledger_df: pd.DataFrame):
    """Return a function SKU -> gudang holding the most of that SKU (ledger), else
    the first gudang column / '' when the SKU isn't in the ledger."""
    gudang_cols = [c for c in ledger_df.columns if c not in ("SKU", "Total")]
    led = ledger_df.set_index("SKU") if "SKU" in ledger_df.columns else ledger_df
    # Fallback for a SKU with no book stock anywhere: the busiest warehouse overall
    # (most likely where a found item physically sits), not an alphabetical guess.
    fallback = (ledger_df[gudang_cols].sum().idxmax() if gudang_cols else "")

    def pick(sku):
        # Largest ABSOLUTE holding: where the SKU actually sits, even when an
        # oversold gudang shows a negative (a 0 gudang must not win over a −2).
        if gudang_cols and sku in led.index:
            row = led.loc[sku, gudang_cols].astype(float)
            return row.abs().idxmax() if row.abs().sum() > 0 else fallback
        return fallback
    return pick


def analyze_stock_opname(opname: pd.DataFrame, sisa_by_sku: pd.Series,
                         ledger_df: pd.DataFrame, hpp_agg: pd.DataFrame,
                         today: pd.Timestamp,
                         value_basis: str = STOCK_OPNAME_VALUE_BASIS,
                         default_gudang: str = STOCK_OPNAME_DEFAULT_GUDANG
                         ) -> pd.DataFrame:
    """Reconcile physical count vs the bot ledger. One row per opname SKU with
    Banyak/Nilai Ketemu/Hilang, the booked gudang, date and note.
    `selisih = book − fisik`; >0 → Hilang, <0 → Ketemu. Value = qty × HPP
    (`value_basis` column of hpp_agg, default hpp_wa)."""
    basis_col = value_basis if value_basis in hpp_agg.columns else "hpp_wa"
    hpp = hpp_agg.set_index("SKU")[basis_col].to_dict()
    pick_gudang = _dominant_gudang(ledger_df)

    rows = []
    for _, r in opname.iterrows():
        sku = r["SKU"]
        fisik = float(r["fisik"])
        book = float(sisa_by_sku.get(sku, 0.0))
        selisih = book - fisik
        ketemu = max(0.0, -selisih)
        hilang = max(0.0, selisih)
        h = hpp.get(sku)
        h = float(h) if h is not None and pd.notna(h) else 0.0

        gud = r.get("gudang")
        if gud is None or (isinstance(gud, float) and pd.isna(gud)) or str(gud).strip() == "":
            gud = default_gudang or pick_gudang(sku)
        tgl = r.get("tanggal")
        tgl = today if pd.isna(tgl) else pd.Timestamp(tgl)
        note = r.get("note")
        if note is None or (isinstance(note, float) and pd.isna(note)) or str(note).strip() == "":
            note = f"Stock opname {tgl.strftime('%Y-%m-%d')}"

        rows.append({
            "tanggal": tgl.normalize(),
            "SKU": sku,
            "stok_buku": book,
            "stok_fisik": fisik,
            "selisih": selisih,
            "banyak_ketemu": ketemu,
            "nilai_ketemu": round(ketemu * h),
            "banyak_hilang": hilang,
            "nilai_hilang": round(hilang * h),
            "gudang": str(gud),
            "keterangan": str(note),
            "ada_selisih": bool(ketemu or hilang),
        })

    df = pd.DataFrame(rows)
    n_h = int((df["banyak_hilang"] > 0).sum()) if len(df) else 0
    n_k = int((df["banyak_ketemu"] > 0).sum()) if len(df) else 0
    n_ok = len(df) - n_h - n_k
    th = int(df["nilai_hilang"].sum()) if len(df) else 0
    tk = int(df["nilai_ketemu"].sum()) if len(df) else 0
    print(f"✓ Stock opname: {len(df)} SKU dicek — {n_h} Hilang, {n_k} Ketemu, {n_ok} cocok "
          f"(basis nilai: {basis_col})")
    print(f"  → Nilai Hilang Rp {th:,} | Nilai Ketemu Rp {tk:,} | net Rp {th - tk:,}")
    return df


def write_bisahilang_report(filepath: Path, recon: pd.DataFrame,
                            today: pd.Timestamp) -> Path:
    """Write output/BisaHilang_Rekonsiliasi.xlsx: a pasteable 'BisaHilang' sheet
    (discrepancy rows only) + a 'Ringkasan' audit sheet (every counted SKU)."""
    wb = Workbook()

    # --- Sheet 1: BisaHilang (discrepancies only, exact tab format) ---
    ws = wb.active
    ws.title = "BisaHilang"
    headers = [COL_BH_TANGGAL, COL_HILANG_SKU, COL_HILANG_KETEMU, COL_BH_NILAI_KETEMU,
               COL_HILANG_HILANG, COL_BH_NILAI_HILANG, COL_HILANG_GUDANG, COL_BH_NOTE]
    widths = [14, 42, 10, 13, 10, 13, 16, 42]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _THIN
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 46

    disc = recon[recon["ada_selisih"]] if len(recon) else recon
    r = 2
    for _, row in disc.iterrows():
        ws.cell(row=r, column=1, value=row["tanggal"]).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=row["SKU"])
        ws.cell(row=r, column=3, value=int(row["banyak_ketemu"]))
        ws.cell(row=r, column=4, value=int(row["nilai_ketemu"])).number_format = _FMT_RP
        ws.cell(row=r, column=5, value=int(row["banyak_hilang"]))
        ws.cell(row=r, column=6, value=int(row["nilai_hilang"])).number_format = _FMT_RP
        ws.cell(row=r, column=7, value=row["gudang"])
        ws.cell(row=r, column=8, value=row["keterangan"])
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = _THIN
            ws.cell(row=r, column=c).font = _NORMAL_FONT
        r += 1
    if len(disc) == 0:
        ws.cell(row=2, column=1, value="(tidak ada selisih — semua cocok)").font = _BOLD_FONT
    ws.freeze_panes = "A2"

    # --- Sheet 2: Ringkasan (audit of every counted SKU) ---
    rk = wb.create_sheet("Ringkasan")
    th = int(recon["nilai_hilang"].sum()) if len(recon) else 0
    tk = int(recon["nilai_ketemu"].sum()) if len(recon) else 0
    n_h = int((recon["banyak_hilang"] > 0).sum()) if len(recon) else 0
    n_k = int((recon["banyak_ketemu"] > 0).sum()) if len(recon) else 0
    n_ok = len(recon) - n_h - n_k
    rk["A1"] = f"REKONSILIASI STOCK OPNAME — {today.strftime('%d %B %Y')}"
    rk["A1"].font = Font(name="Arial", bold=True, size=12)
    summ = [
        ("SKU dicek", len(recon)),
        ("Hilang (buku > fisik)", n_h),
        ("Ketemu (fisik > buku)", n_k),
        ("Cocok", n_ok),
        ("Total Nilai Hilang", th),
        ("Total Nilai Ketemu", tk),
        ("Net write-off (Hilang − Ketemu)", th - tk),
    ]
    for j, (k, v) in enumerate(summ, start=3):
        rk.cell(row=j, column=1, value=k).font = _BOLD_FONT
        cell = rk.cell(row=j, column=2, value=v)
        if "Nilai" in k or "write-off" in k:
            cell.number_format = _FMT_RP
    rk.column_dimensions["A"].width = 34
    rk.column_dimensions["B"].width = 16

    hdr_row = 11
    audit_cols = ["Tanggal", "SKU", "Stok Buku", "Stok Fisik", "Selisih",
                  "Banyak Ketemu", "Banyak Hilang", "Gudang", "Status"]
    audit_w = [12, 42, 11, 11, 9, 13, 13, 16, 10]
    for i, h in enumerate(audit_cols, start=1):
        c = rk.cell(row=hdr_row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        rk.column_dimensions[get_column_letter(i)].width = audit_w[i - 1]
    rr = hdr_row + 1
    for _, row in recon.iterrows():
        status = ("Hilang" if row["banyak_hilang"] > 0
                  else ("Ketemu" if row["banyak_ketemu"] > 0 else "Cocok"))
        rk.cell(row=rr, column=1, value=row["tanggal"]).number_format = "yyyy-mm-dd"
        rk.cell(row=rr, column=2, value=row["SKU"])
        rk.cell(row=rr, column=3, value=int(row["stok_buku"]))
        rk.cell(row=rr, column=4, value=int(row["stok_fisik"]))
        rk.cell(row=rr, column=5, value=int(row["selisih"]))
        rk.cell(row=rr, column=6, value=int(row["banyak_ketemu"]))
        rk.cell(row=rr, column=7, value=int(row["banyak_hilang"]))
        rk.cell(row=rr, column=8, value=row["gudang"])
        rk.cell(row=rr, column=9, value=status)
        rr += 1
    rk.freeze_panes = "A12"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ BisaHilang rekonsiliasi: {filepath} ({len(disc)} baris selisih)")
    return filepath
