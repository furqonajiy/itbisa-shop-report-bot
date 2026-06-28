"""A/B testing analyzer: compare sales metrics before vs after price changes."""
from __future__ import annotations
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_loader import resolve_sheet
from config import (
    AB_BULK_CONCENTRATION, AB_LOW_STOCK_MONTHS_COVER,
    AB_MIN_DAYS_POST, AB_MIN_TRANS_POST, AB_MIN_TRANS_PRE,
    AB_MIN_VALID_DAYS,
    AB_PRE_WINDOW_DAYS, AB_TESTS_SHEET, ALERT_TEXT_COLOR, COL_AB_CATATAN,
    COL_AB_NAMA, COL_AB_SKU, COL_AB_TANGGAL, FMT_DEC, FMT_NUM, FMT_PCT, FMT_RP,
    FONT_NAME, GREEN_FILL_COLOR, HEADER_BG_COLOR, HEADER_TEXT_COLOR,
    LIGHT_GRAY_COLOR, RED_FILL_COLOR, STATUS_SUDAH_DIPESAN, TITLE_COLOR, YELLOW_FILL_COLOR,
)

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)
GREEN_FILL = PatternFill("solid", start_color=GREEN_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
RED_FILL = PatternFill("solid", start_color=RED_FILL_COLOR)


def load_ab_tests(filepath: Path) -> pd.DataFrame:
    """Load A/B test config. Returns empty DataFrame if file missing."""
    if not filepath.exists():
        print(f"  ⚠ File {filepath.name} tidak ditemukan, A/B test skipped")
        return pd.DataFrame()

    df = pd.read_excel(filepath, sheet_name=resolve_sheet(filepath, AB_TESTS_SHEET))
    df = df.rename(columns={
        COL_AB_SKU: "sku",
        COL_AB_TANGGAL: "tanggal_perubahan",
        COL_AB_NAMA: "nama_test",
        COL_AB_CATATAN: "catatan",
    })
    df["tanggal_perubahan"] = pd.to_datetime(df["tanggal_perubahan"], errors="coerce")
    df = df[df["sku"].notna() & df["tanggal_perubahan"].notna()].copy()
    print(f"✓ Loaded {len(df):,} test config dari {filepath.name}")
    return df


def load_ab_change_dates(ab_config_path: Path) -> pd.Series:
    """Per-SKU latest logged change date (SKU normalized to UPPER().strip() to match
    the jual data) from ab_tests.xlsx — the authoritative signal for the Kandidat
    recent-increase guard. Empty Series if the file is missing or has no valid rows."""
    if not ab_config_path.exists():
        return pd.Series(dtype="datetime64[ns]")
    df = load_ab_tests(ab_config_path)
    if len(df) == 0:
        return pd.Series(dtype="datetime64[ns]")
    df = df.copy()
    df["sku"] = df["sku"].astype(str).str.upper().str.strip()
    return df.groupby("sku")["tanggal_perubahan"].max()


def _compute_period_metrics(jual_sub: pd.DataFrame, hpp_wa: float) -> dict:
    """Compute metrics for a slice of jual transactions."""
    if len(jual_sub) == 0:
        return {
            "n_trans": 0, "qty": 0, "omzet": 0.0, "profit": 0.0,
            "avg_price": float("nan"), "markup_pct": float("nan"),
            "margin_pct": float("nan"), "days": 0,
            "qty_per_day": 0.0, "omzet_per_day": 0.0, "profit_per_day": 0.0,
            "admin_per_day": 0.0, "margin_per_unit": float("nan"),
            "max_single_order": 0.0, "first_date": None, "last_date": None,
        }

    qty = jual_sub["qty_jual"].sum()
    omzet = jual_sub["omzet"].sum()
    biaya_admin = (jual_sub["tambahan"] + jual_sub["kode_unik"]).sum()
    hpp_total = hpp_wa * qty if pd.notna(hpp_wa) else 0
    profit = omzet - hpp_total + biaya_admin

    first_date = jual_sub["tanggal_pesan"].min()
    last_date = jual_sub["tanggal_pesan"].max()
    days = max(1, (last_date - first_date).days + 1)

    avg_price = omzet / qty if qty > 0 else float("nan")
    markup_pct = ((avg_price - hpp_wa) / hpp_wa * 100) if pd.notna(hpp_wa) and hpp_wa > 0 else float("nan")
    margin_pct = (profit / omzet * 100) if omzet > 0 else float("nan")
    # Gross margin per unit (sebelum admin) — basis bridge & break-even.
    margin_per_unit = (avg_price - hpp_wa) if pd.notna(hpp_wa) and qty > 0 else float("nan")

    return {
        "n_trans": jual_sub["Invoice"].nunique(),
        "qty": qty,
        "omzet": omzet,
        "profit": profit,
        "avg_price": avg_price,
        "markup_pct": markup_pct,
        "margin_pct": margin_pct,
        "days": days,
        "qty_per_day": qty / days,
        "omzet_per_day": omzet / days,
        "profit_per_day": profit / days,
        "admin_per_day": biaya_admin / days,
        "margin_per_unit": margin_per_unit,
        "max_single_order": float(jual_sub["qty_jual"].max()),
        "first_date": first_date,
        "last_date": last_date,
    }


def _pct_change(old: float, new: float) -> float:
    if pd.isna(old) or pd.isna(new) or old == 0:
        return float("nan")
    return (new - old) / abs(old) * 100


def _profit_bridge(pre_m: dict, post_m: dict) -> dict:
    """Decompose Δ(profit/hari) into price, volume, interaction and admin effects.

    Gross margin/day = margin_per_unit × qty/day, so its change splits exactly into:
      efek_harga  = Δmargin_per_unit × qty/day_pre   (margin naik, volume tetap)
      efek_volume = Δqty/day × margin_per_unit_pre   (volume berubah, margin tetap)
      interaksi   = Δmargin_per_unit × Δqty/day
    Δprofit/day = (efek_harga + efek_volume + interaksi) + efek_admin.
    Answers directly: did the price move (efek_harga, +) get cancelled by lost
    volume (efek_volume, −)? Break-even = how much volume could drop first."""
    mu_pre, mu_post = pre_m["margin_per_unit"], post_m["margin_per_unit"]
    q_pre, q_post = pre_m["qty_per_day"], post_m["qty_per_day"]
    if pd.isna(mu_pre) or pd.isna(mu_post):
        return {k: float("nan") for k in
                ("efek_harga", "efek_volume", "interaksi", "efek_admin",
                 "break_even_drop_pct", "headroom_pct", "elasticity")}

    d_mu, d_q = mu_post - mu_pre, q_post - q_pre
    efek_harga = d_mu * q_pre
    efek_volume = d_q * mu_pre
    interaksi = d_mu * d_q
    efek_admin = post_m["admin_per_day"] - pre_m["admin_per_day"]

    # Volume boleh turun s/d ini sebelum gross margin/day balik ke level pre.
    break_even = (1 - mu_pre / mu_post) * 100 if mu_post > 0 else float("nan")
    d_qty_pct = _pct_change(q_pre, q_post)
    headroom = (d_qty_pct + break_even) if pd.notna(d_qty_pct) and pd.notna(break_even) else float("nan")
    d_price_pct = _pct_change(pre_m["avg_price"], post_m["avg_price"])
    elasticity = (d_qty_pct / d_price_pct) if pd.notna(d_qty_pct) and pd.notna(d_price_pct) and d_price_pct != 0 else float("nan")

    return {
        "efek_harga": efek_harga, "efek_volume": efek_volume,
        "interaksi": interaksi, "efek_admin": efek_admin,
        "break_even_drop_pct": break_even, "headroom_pct": headroom,
        "elasticity": elasticity,
    }


def _confound_flags(pre_m: dict, post_m: dict, bridge: dict) -> list[str]:
    """Reasons the causal claim (price → profit) may be unreliable."""
    flags = []
    if post_m["days"] < AB_MIN_DAYS_POST:
        flags.append(f"post baru {post_m['days']} hari")
    if post_m["n_trans"] < AB_MIN_TRANS_POST:
        flags.append(f"post cuma {post_m['n_trans']} transaksi")
    if pre_m["n_trans"] < AB_MIN_TRANS_PRE:
        flags.append(f"baseline pre tipis ({pre_m['n_trans']} trx)")
    if post_m["qty"] > 0 and post_m["max_single_order"] / post_m["qty"] > AB_BULK_CONCENTRATION:
        share = post_m["max_single_order"] / post_m["qty"] * 100
        flags.append(f"qty post didominasi 1 order grosir ({share:.0f}%)")
    el = bridge["elasticity"]
    if pd.notna(el) and el > 0:
        flags.append("qty & harga sama-sama naik (elastisitas +) → ada faktor lain, efek harga tak terisolasi")
    return flags


def _build_stock_lookup(reorder_df: pd.DataFrame | None) -> dict[str, dict]:
    if reorder_df is None or len(reorder_df) == 0 or "SKU" not in reorder_df.columns:
        return {}
    rows = reorder_df.drop_duplicates("SKU", keep="last").to_dict("records")
    lookup = {}
    for row in rows:
        sku = str(row.get("SKU", "")).strip()
        if sku:
            lookup[sku] = row
            lookup[sku.upper()] = row
    return lookup


def _fmt_qty(value: float) -> str:
    if pd.isna(value):
        return "?"
    value = float(value)
    if value.is_integer():
        return f"{int(value):,}"
    return f"{value:,.1f}"


def _stock_pause_reason(stock: dict | None) -> str:
    if not stock:
        return ""

    status = str(stock.get("status") or "")
    sisa = stock.get("sisa_stok", float("nan"))
    months_cover = stock.get("months_cover", float("nan"))
    cover_value = float("nan")
    if pd.notna(months_cover):
        cover_value = float(months_cover)

    is_stockout = "STOCKOUT" in status or (pd.notna(sisa) and sisa <= 0)
    is_low_cover = (
        pd.notna(cover_value)
        and cover_value != float("inf")
        and cover_value <= AB_LOW_STOCK_MONTHS_COVER
    )
    if not (is_stockout or is_low_cover):
        return ""

    details = []
    if status:
        details.append(status)
    if pd.notna(sisa):
        details.append(f"sisa {_fmt_qty(sisa)} pcs")
    if pd.notna(cover_value) and cover_value != float("inf"):
        details.append(f"cover {cover_value:.1f} bln")
    eta = stock.get("est_arrival")
    if status == STATUS_SUDAH_DIPESAN and eta is not None and pd.notna(eta):
        details.append(f"ETA {pd.Timestamp(eta).strftime('%Y-%m-%d')}")

    suffix = f" ({', '.join(details)})" if details else ""
    return f"stok hampir habis{suffix} → tunggu restock sebelum baca A/B"


def _verdict(bridge: dict, delta_profit: float, delta_qty: float,
             confounds: list[str], stock_pause: str = "",
             in_progress: str = "") -> str:
    """Descriptive verdict on profit, with break-even framing.

    A test younger than AB_MIN_VALID_DAYS (~2 months) is NOT yet conclusive:
    `in_progress` overrides everything so the verdict reads "⏳ In Progress"
    instead of a (premature) Effective/Bad/Mixed call."""
    if in_progress:
        return in_progress
    if stock_pause:
        return "⏸️ Pending — stok hampir habis, A/B ditunda sampai restock"
    if pd.isna(delta_profit):
        return "⚪ Inconclusive — data kurang"
    be = bridge["break_even_drop_pct"]
    strong_confound = any("didominasi" in c or "elastisitas" in c or "post baru" in c
                          for c in confounds)
    if delta_profit < -5:
        if pd.notna(delta_qty) and pd.notna(be) and delta_qty < -be:
            return "🔴 Bad — volume turun melebihi break-even, harga naik menurunkan profit"
        return "🔴 Bad — profit/hari turun"
    if delta_profit > 5:
        base = "✅ Effective — profit/hari naik"
        if pd.notna(delta_qty) and delta_qty < 0:
            base = "✅ Effective — margin menutup penurunan volume"
        if strong_confound:
            return base.replace("✅", "🟡") + " (atribusi lemah, cek catatan)"
        return base
    return "⚪ No significant change"


def analyze_ab_tests(ab_tests: pd.DataFrame, jual_full_clean: pd.DataFrame,
                     hpp_agg: pd.DataFrame, today: datetime,
                     reorder_df: pd.DataFrame | None = None) -> pd.DataFrame:
    """For each test config, compare a comparable pre-window vs post change date,
    decompose Δprofit (price vs volume), compute break-even volume drop & confounds."""
    if len(ab_tests) == 0:
        return pd.DataFrame()

    hpp_lookup = hpp_agg.set_index("SKU")["hpp_wa"].to_dict()
    stock_lookup = _build_stock_lookup(reorder_df)
    results = []

    for _, t in ab_tests.iterrows():
        sku = t["sku"]
        change_date = t["tanggal_perubahan"]
        hpp = hpp_lookup.get(sku, float("nan"))

        sku_jual = jual_full_clean[jual_full_clean["SKU"] == sku].copy()
        if len(sku_jual) == 0:
            print(f"  ⚠ {sku}: tidak ada transaksi, skip")
            continue

        pre_start = change_date - pd.Timedelta(days=AB_PRE_WINDOW_DAYS)
        pre = sku_jual[(sku_jual["tanggal_pesan"] >= pre_start)
                       & (sku_jual["tanggal_pesan"] < change_date)]
        post = sku_jual[sku_jual["tanggal_pesan"] >= change_date]

        pre_m = _compute_period_metrics(pre, hpp)
        post_m = _compute_period_metrics(post, hpp)
        bridge = _profit_bridge(pre_m, post_m)
        confounds = _confound_flags(pre_m, post_m, bridge)
        stock = stock_lookup.get(str(sku).strip()) or stock_lookup.get(str(sku).upper().strip())
        stock_pause = _stock_pause_reason(stock)
        if stock_pause:
            confounds.insert(0, stock_pause)

        # Hard 2-month validity gate: below AB_MIN_VALID_DAYS post-change the test
        # is still running — force an "In Progress" verdict, no conclusive call.
        days_since_change = max(0, (today - change_date).days)
        in_progress = ""
        if days_since_change < AB_MIN_VALID_DAYS:
            in_progress = (f"⏳ In Progress (<2 bln) — baru {days_since_change} hari "
                           f"sejak perubahan, butuh ≥{AB_MIN_VALID_DAYS} hari")
            confounds.insert(0, f"masa uji {days_since_change}/{AB_MIN_VALID_DAYS} hari "
                                f"— hasil belum valid (<2 bln)")

        delta_qty = _pct_change(pre_m["qty_per_day"], post_m["qty_per_day"])
        delta_omzet = _pct_change(pre_m["omzet_per_day"], post_m["omzet_per_day"])
        delta_profit = _pct_change(pre_m["profit_per_day"], post_m["profit_per_day"])
        delta_price = _pct_change(pre_m["avg_price"], post_m["avg_price"])

        results.append({
            "sku": sku,
            "nama_test": t.get("nama_test", "") or "",
            "tanggal_perubahan": change_date,
            "days_since_change": days_since_change,
            "hpp_wa": hpp,
            "pre_window_days": AB_PRE_WINDOW_DAYS,
            "pre_first": pre_m["first_date"],
            "pre_last": pre_m["last_date"],
            "pre_days": pre_m["days"] if pre_m["n_trans"] > 0 else 0,
            "pre_n_trans": pre_m["n_trans"],
            "pre_qty": pre_m["qty"],
            "pre_qty_per_day": pre_m["qty_per_day"],
            "pre_omzet_per_day": pre_m["omzet_per_day"],
            "pre_profit_per_day": pre_m["profit_per_day"],
            "pre_avg_price": pre_m["avg_price"],
            "pre_markup_pct": pre_m["markup_pct"],
            "pre_margin_pct": pre_m["margin_pct"],
            "post_first": post_m["first_date"],
            "post_last": post_m["last_date"],
            "post_days": post_m["days"] if post_m["n_trans"] > 0 else 0,
            "post_n_trans": post_m["n_trans"],
            "post_qty": post_m["qty"],
            "post_qty_per_day": post_m["qty_per_day"],
            "post_omzet_per_day": post_m["omzet_per_day"],
            "post_profit_per_day": post_m["profit_per_day"],
            "post_avg_price": post_m["avg_price"],
            "post_markup_pct": post_m["markup_pct"],
            "post_margin_pct": post_m["margin_pct"],
            "sisa_stok": stock.get("sisa_stok", float("nan")) if stock else float("nan"),
            "stock_months_cover": stock.get("months_cover", float("nan")) if stock else float("nan"),
            "reorder_status": stock.get("status", "") if stock else "",
            "delta_qty_pct": delta_qty,
            "delta_omzet_pct": delta_omzet,
            "delta_profit_pct": delta_profit,
            "delta_price_pct": delta_price,
            "efek_harga_per_hari": bridge["efek_harga"],
            "efek_volume_per_hari": bridge["efek_volume"],
            "interaksi_per_hari": bridge["interaksi"],
            "efek_admin_per_hari": bridge["efek_admin"],
            "break_even_drop_pct": bridge["break_even_drop_pct"],
            "headroom_pct": bridge["headroom_pct"],
            "elasticity": bridge["elasticity"],
            "verdict": _verdict(bridge, delta_profit, delta_qty, confounds,
                                stock_pause, in_progress),
            "warning": "; ".join(confounds),
            "catatan": t.get("catatan", "") or "",
        })

    return pd.DataFrame(results)


def write_ab_test_report(output_path: Path, results: pd.DataFrame,
                          today: datetime) -> None:
    """Write A/B test analysis to a standalone Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"

    ws["A1"] = "LAPORAN ANALISA A/B TEST — ITBISA SHOP"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (f"Generated: {today.strftime('%d %B %Y %H:%M')}  |  "
                f"Pre = {AB_PRE_WINDOW_DAYS} hari sebelum perubahan vs Post (daily-rate). "
                f"Profit bridge memisah efek harga vs volume. "
                f"Hasil baru valid setelah ≥{AB_MIN_VALID_DAYS} hari (±2 bln) — "
                f"di bawah itu ⏳ In Progress.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:F2")

    if len(results) == 0:
        ws["A4"] = "Tidak ada test config. Isi data/ab_tests.xlsx untuk track perubahan harga."
        ws["A4"].font = BOLD_FONT
        ws.column_dimensions["A"].width = 70
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"✓ Menulis laporan ke {output_path} (kosong)")
        return

    ws["A4"] = "RINGKASAN"
    ws["A4"].font = TITLE_FONT
    summary = [
        ("Total Test Tracked", len(results)),
        ("In Progress (<2 bln, belum valid)", (results["verdict"].str.startswith("⏳")).sum()),
        ("Effective (profit naik)", (results["verdict"].str.startswith("✅")).sum()),
        ("Mixed", (results["verdict"].str.startswith("🟡")).sum()),
        ("Bad (profit turun)", (results["verdict"].str.startswith("🔴")).sum()),
        ("Pending restock (A/B ditunda)", (results["verdict"].str.startswith("⏸")).sum()),
        ("Inconclusive / No change", (results["verdict"].str.startswith("⚪")).sum()),
    ]
    for i, (lbl, val) in enumerate(summary, start=5):
        ws.cell(row=i, column=1, value=lbl).font = NORMAL_FONT
        c = ws.cell(row=i, column=2, value=val)
        c.font = BOLD_FONT
        c.number_format = FMT_NUM

    list_title_row = 5 + len(summary) + 1
    ws.cell(row=list_title_row, column=1, value="DAFTAR TEST").font = TITLE_FONT
    for i, (_, r) in enumerate(results.iterrows(), start=list_title_row + 1):
        ws.cell(row=i, column=1, value=r["sku"]).font = NORMAL_FONT
        ws.cell(row=i, column=2, value=r["nama_test"]).font = NORMAL_FONT
        c = ws.cell(row=i, column=3, value=r["verdict"])
        c.font = BOLD_FONT
        if r["warning"]:
            wc = ws.cell(row=i, column=4, value=r["warning"])
            wc.font = Font(name=FONT_NAME, italic=True, size=10, color=ALERT_TEXT_COLOR)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 40

    _write_details_sheet(wb.create_sheet("01_Test_Results"), results)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✓ Menulis laporan ke {output_path}")


def _write_details_sheet(ws, results: pd.DataFrame) -> None:
    ws["A1"] = "DETAIL A/B TEST — PRE vs POST + PROFIT BRIDGE"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:Z1")
    ws["A2"] = ("Pre = window setara sebelum perubahan (bukan all-time). Bridge: Δprofit/hari "
                "= Efek Harga + Efek Volume + Interaksi + Efek Admin. Break-even = batas turun "
                "volume sebelum profit balik ke level pre; Headroom = jarak qty aktual ke batas itu "
                "(+ = aman). Stok hampir habis/stockout = A/B ditunda sampai restock. "
                "Masa Uji < 2 bln (60 hari) = ⏳ In Progress, hasil belum valid.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:Z2")

    headers = [
        "SKU", "Nama Test", "Tgl Perubahan", "Masa Uji (hari)", "HPP/Buah",
        "Pre Qty/Day", "Pre Profit/Day", "Pre Avg Price",
        "Post Qty/Day", "Post Profit/Day", "Post Avg Price",
        "Sisa Stok", "Cover Stok (bln)", "Status Reorder",
        "Δ Qty/Day", "Δ Profit/Day", "Δ Price",
        "Efek Harga/hr", "Efek Volume/hr", "Interaksi/hr", "Efek Admin/hr",
        "Break-even Turun Qty", "Headroom Qty", "Elastisitas",
        "Verdict", "Catatan (confound)",
    ]
    widths = [38, 22, 13, 13, 11,
              12, 14, 13,
              12, 14, 13,
              10, 12, 22,
              11, 11, 11,
              14, 14, 13, 13,
              16, 13, 11,
              46, 52]

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[4].height = 30

    field_map = {
        "SKU": "sku", "Nama Test": "nama_test", "Tgl Perubahan": "tanggal_perubahan",
        "Masa Uji (hari)": "days_since_change",
        "HPP/Buah": "hpp_wa",
        "Pre Qty/Day": "pre_qty_per_day", "Pre Profit/Day": "pre_profit_per_day",
        "Pre Avg Price": "pre_avg_price",
        "Post Qty/Day": "post_qty_per_day", "Post Profit/Day": "post_profit_per_day",
        "Post Avg Price": "post_avg_price",
        "Sisa Stok": "sisa_stok", "Cover Stok (bln)": "stock_months_cover",
        "Status Reorder": "reorder_status",
        "Δ Qty/Day": "delta_qty_pct", "Δ Profit/Day": "delta_profit_pct", "Δ Price": "delta_price_pct",
        "Efek Harga/hr": "efek_harga_per_hari", "Efek Volume/hr": "efek_volume_per_hari",
        "Interaksi/hr": "interaksi_per_hari", "Efek Admin/hr": "efek_admin_per_hari",
        "Break-even Turun Qty": "break_even_drop_pct", "Headroom Qty": "headroom_pct",
        "Elastisitas": "elasticity",
        "Verdict": "verdict", "Catatan (confound)": "warning",
    }
    pct_headers = {"Δ Qty/Day", "Δ Profit/Day", "Δ Price", "Break-even Turun Qty", "Headroom Qty"}
    rp_headers = {"HPP/Buah", "Pre Profit/Day", "Pre Avg Price", "Post Profit/Day",
                  "Post Avg Price", "Efek Harga/hr", "Efek Volume/hr", "Interaksi/hr", "Efek Admin/hr"}
    num_headers = {"Pre Qty/Day", "Post Qty/Day", "Sisa Stok"}
    dec_headers = {"Elastisitas", "Cover Stok (bln)"}
    int_headers = {"Masa Uji (hari)"}
    verdict_col = headers.index("Verdict") + 1

    for r_idx, (_, row) in enumerate(results.iterrows()):
        verdict = row["verdict"]
        for c_idx, h in enumerate(headers, start=1):
            val = row[field_map[h]]
            if pd.isna(val):
                val = None
            elif h in pct_headers and val is not None:
                val = val / 100
            cell = ws.cell(row=5 + r_idx, column=c_idx, value=val)
            cell.font = NORMAL_FONT
            if h in pct_headers:
                cell.number_format = FMT_PCT
            elif h in rp_headers:
                cell.number_format = FMT_RP
            elif h in num_headers:
                cell.number_format = FMT_DEC
            elif h in dec_headers:
                cell.number_format = FMT_DEC
            elif h in int_headers:
                cell.number_format = FMT_NUM
            if r_idx % 2 == 1:
                cell.fill = LIGHT_FILL

        vcell = ws.cell(row=5 + r_idx, column=verdict_col)
        if verdict.startswith("✅"):
            vcell.fill = GREEN_FILL
        elif verdict.startswith("🔴"):
            vcell.fill = RED_FILL
        elif verdict.startswith("🟡") or verdict.startswith("⏸") or verdict.startswith("⏳"):
            vcell.fill = YELLOW_FILL

    ws.freeze_panes = "B5"


def create_template(filepath: Path, example_sku: str = "ITBISA-IC-NE555P-DIP8") -> None:
    """Create the ab_tests.xlsx template with sample data."""
    wb = Workbook()
    ws = wb.active
    ws.title = AB_TESTS_SHEET

    headers = [COL_AB_SKU, COL_AB_TANGGAL, COL_AB_NAMA, COL_AB_CATATAN]
    widths = [38, 18, 30, 50]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 25

    example = [
        example_sku,
        datetime(2026, 5, 17),
        "NE555P Price Bump May 2026",
        "Harga naik dari 599 ke 699 (single), 50pcs ke 689, 1000pcs ke 679",
    ]
    for i, v in enumerate(example, start=1):
        c = ws.cell(row=2, column=i, value=v)
        c.font = NORMAL_FONT
        if i == 2:
            c.number_format = "yyyy-mm-dd"

    ws.cell(row=4, column=1, value="Format kolom:").font = BOLD_FONT
    notes = [
        "SKU: persis sama dengan SKU di data Jual (case-sensitive)",
        "Tanggal Perubahan: tanggal harga berubah (YYYY-MM-DD)",
        "Nama Test: nama bebas untuk identifikasi (opsional)",
        "Catatan: detail perubahan harga, alasan, dll. (opsional)",
        "",
        "Tambah baris baru untuk setiap perubahan harga yang ingin dilacak.",
    ]
    for i, n in enumerate(notes, start=5):
        c = ws.cell(row=i, column=1, value=n)
        c.font = SUB_FONT

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Template dibuat: {filepath}")
