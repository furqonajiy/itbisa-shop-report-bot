"""Sales trend & seasonality report (`--trend`).

A cross-year view that the per-year files can't give on their own:

  - **Yearly trend** — omzet / qty / profit / orders per year + YoY growth (the
    current year is flagged partial).
  - **Monthly trend** — the full omzet/profit time series across all years.
  - **Seasonality** — a per-calendar-month index = how that month compares to its
    own year's typical month, averaged over the COMPLETE years (the partial current
    year is excluded so it can't bias the pattern). Index > 1 = consistently strong.

`profit = omzet + admin − HPP_WA × qty` (admin = `tambahan + kode_unik`, negative;
SKUs without HPP contribute omzet but not profit). Output:
output/Analisa_Tren_Musiman.xlsx. Zero-config (built from BisaJual + HPP).
"""
from __future__ import annotations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    BLUE_FILL_COLOR, FMT_NUM, FMT_RP, FONT_NAME, GREEN_FILL_COLOR,
    HEADER_BG_COLOR, HEADER_TEXT_COLOR, LIGHT_GRAY_COLOR, RED_FILL_COLOR,
    TITLE_COLOR, TREND_LOW_INDEX, TREND_PEAK_INDEX, TREND_SEASONAL_MIN_YEARS,
    YELLOW_FILL_COLOR,
)

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
GREEN_FILL = PatternFill("solid", start_color=GREEN_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
RED_FILL = PatternFill("solid", start_color=RED_FILL_COLOR)
BLUE_FILL = PatternFill("solid", start_color=BLUE_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)
FMT_PCT1 = '+0.0%;-0.0%'
FMT_IDX = '0.00'

_BULAN = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
          "Agustus", "September", "Oktober", "November", "Desember"]


def analyze_trend(jual: pd.DataFrame, hpp_agg: pd.DataFrame, today) -> dict:
    """Returns {'monthly', 'yearly', 'seasonal', 'headline'} DataFrames/dict."""
    needed = {"SKU", "tanggal_pesan", "qty_jual", "omzet"}
    if jual is None or len(jual) == 0 or not needed.issubset(jual.columns):
        return {}

    hpp_wa = hpp_agg.set_index("SKU")["hpp_wa"].to_dict() if len(hpp_agg) else {}
    j = jual[(jual["qty_jual"] > 0) & jual["tanggal_pesan"].notna()].copy()
    j["_admin"] = j.get("tambahan", 0).fillna(0) + j.get("kode_unik", 0).fillna(0)
    j["_hpp"] = j["SKU"].map(hpp_wa)
    j["_profit"] = j["omzet"] + j["_admin"] - j["_hpp"] * j["qty_jual"]   # NaN if no HPP
    j["year"] = j["tanggal_pesan"].dt.year.astype(int)
    j["month"] = j["tanggal_pesan"].dt.month.astype(int)
    j["ym"] = j["tanggal_pesan"].dt.to_period("M")
    has_inv = "Invoice" in j.columns

    def _orders(d):
        return d["Invoice"].nunique() if has_inv else len(d)

    # --- monthly time series (full calendar range, gaps filled with 0) ---
    rows = []
    for ym, d in j.groupby("ym"):
        rows.append({"ym": str(ym), "year": ym.year, "month": ym.month,
                     "omzet": float(d["omzet"].sum()),
                     "qty": float(d["qty_jual"].sum()),
                     "profit": float(d["_profit"].sum(skipna=True)),
                     "orders": int(_orders(d))})
    monthly = pd.DataFrame(rows).sort_values("ym").reset_index(drop=True)
    if len(monthly):
        full = pd.period_range(j["ym"].min(), j["ym"].max(), freq="M")
        idx = pd.DataFrame({"ym": [str(p) for p in full],
                            "year": [p.year for p in full], "month": [p.month for p in full]})
        monthly = idx.merge(monthly.drop(columns=["year", "month"]), on="ym", how="left")
        for c in ("omzet", "qty", "profit", "orders"):
            monthly[c] = monthly[c].fillna(0.0)

    # --- yearly trend + YoY ---
    yr = (j.groupby("year").agg(omzet=("omzet", "sum"), qty=("qty_jual", "sum"),
                                profit=("_profit", "sum")).reset_index())
    yr["orders"] = [int(_orders(j[j["year"] == y])) for y in yr["year"]]
    yr = yr.sort_values("year").reset_index(drop=True)
    yr["yoy_omzet"] = yr["omzet"].pct_change()
    cur_year = int(today.year)
    yr["partial"] = yr["year"] >= cur_year

    # --- seasonality (complete years only, normalized per year) ---
    complete_years = [y for y in yr["year"] if y < cur_year]
    seas_src = monthly[monthly["year"].isin(complete_years)].copy() if len(monthly) else monthly
    seasonal = pd.DataFrame()
    if len(seas_src):
        year_avg = seas_src.groupby("year")["omzet"].transform("mean")
        seas_src = seas_src.assign(ratio=np.where(year_avg > 0, seas_src["omzet"] / year_avg, np.nan))
        agg = (seas_src.groupby("month").agg(index=("ratio", "mean"), n_years=("year", "nunique"))
               .reindex(range(1, 13)))
        agg["bulan"] = [_BULAN[m - 1] for m in agg.index]
        seasonal = agg.reset_index().rename(columns={"month": "month_no"})

    # --- headline metrics ---
    total_omzet = float(j["omzet"].sum())
    total_profit = float(j["_profit"].sum(skipna=True))
    # year-to-date vs same period last year
    cur_m = int(today.month)
    ytd = float(j[(j["year"] == cur_year) & (j["month"] <= cur_m)]["omzet"].sum())
    prev_ytd = float(j[(j["year"] == cur_year - 1) & (j["month"] <= cur_m)]["omzet"].sum())
    ytd_growth = (ytd / prev_ytd - 1.0) if prev_ytd > 0 else np.nan
    peak = trough = None
    if len(seasonal) and seasonal["index"].notna().any():
        peak = seasonal.loc[seasonal["index"].idxmax()]
        trough = seasonal.loc[seasonal["index"].idxmin()]
    headline = {"total_omzet": total_omzet, "total_profit": total_profit,
                "cur_year": cur_year, "cur_month": cur_m, "ytd": ytd, "prev_ytd": prev_ytd,
                "ytd_growth": ytd_growth,
                "peak": (peak["bulan"], float(peak["index"])) if peak is not None and pd.notna(peak["index"]) else None,
                "trough": (trough["bulan"], float(trough["index"])) if trough is not None and pd.notna(trough["index"]) else None}

    g = f"{ytd_growth:+.0%}" if pd.notna(ytd_growth) else "n/a"
    print(f"✓ Tren & musiman: {len(monthly)} bulan, {len(yr)} tahun — "
          f"YTD {cur_year} vs {cur_year-1}: {g}"
          + (f"; puncak {headline['peak'][0]}" if headline['peak'] else ""))
    return {"monthly": monthly, "yearly": yr, "seasonal": seasonal, "headline": headline}


# ---------------------------------------------------------------------------
def _style_header(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 28


def _season_label(idx: float) -> str:
    if pd.isna(idx):
        return "—"
    if idx >= TREND_PEAK_INDEX:
        return "🔥 Puncak"
    if idx >= 1.0:
        return "🟢 Di atas rata2"
    if idx > TREND_LOW_INDEX:
        return "⚪ Normal"
    return "🔻 Sepi"


def write_trend_report(filepath: Path, data: dict, today) -> None:
    monthly = data.get("monthly", pd.DataFrame())
    yearly = data.get("yearly", pd.DataFrame())
    seasonal = data.get("seasonal", pd.DataFrame())
    h = data.get("headline", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Ringkasan"
    ws["A1"] = "TREN & MUSIMAN PENJUALAN — LINTAS TAHUN"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  Profit = omzet + admin − HPP_WA×qty "
                f"(SKU tanpa HPP tetap masuk omzet, tidak profit). Musiman dihitung dari "
                f"tahun penuh saja (tahun berjalan dikecualikan).")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:E2")

    r = 4
    lines = [("Total omzet (semua tahun)", h.get("total_omzet", 0), FMT_RP),
             ("Total profit (semua tahun)", h.get("total_profit", 0), FMT_RP)]
    for label, val, fmt in lines:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=r, column=2, value=float(val)); c.font = BOLD_FONT; c.number_format = fmt
        r += 1
    cy, pm = h.get("cur_year"), h.get("cur_month")
    ws.cell(row=r, column=1, value=f"Omzet YTD {cy} (s/d bln {pm})").font = BOLD_FONT
    c = ws.cell(row=r, column=2, value=float(h.get("ytd", 0))); c.font = BOLD_FONT; c.number_format = FMT_RP
    r += 1
    ws.cell(row=r, column=1, value=f"Periode sama {cy-1 if cy else ''}").font = NORMAL_FONT
    c = ws.cell(row=r, column=2, value=float(h.get("prev_ytd", 0))); c.font = NORMAL_FONT; c.number_format = FMT_RP
    r += 1
    yg = h.get("ytd_growth")
    ws.cell(row=r, column=1, value="Pertumbuhan YTD (vs tahun lalu)").font = BOLD_FONT
    if yg is not None and pd.notna(yg):
        c = ws.cell(row=r, column=2, value=float(yg)); c.number_format = FMT_PCT1
        c.font = Font(name=FONT_NAME, bold=True, size=10, color=("1A7A1A" if yg >= 0 else "B00000"))
    else:
        ws.cell(row=r, column=2, value="n/a").font = BOLD_FONT
    r += 2

    if h.get("peak"):
        ws.cell(row=r, column=1, value=f"📈 Bulan terkuat: {h['peak'][0]} (indeks {h['peak'][1]:.2f})").font = BOLD_FONT
        r += 1
    if h.get("trough"):
        ws.cell(row=r, column=1, value=f"📉 Bulan terlemah: {h['trough'][0]} (indeks {h['trough'][1]:.2f})").font = BOLD_FONT
        r += 1

    # --- Sheet 01: yearly ---
    ws1 = wb.create_sheet("01_Tren_Tahunan")
    ws1["A1"] = "TREN TAHUNAN + PERTUMBUHAN YoY"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:F1")
    _style_header(ws1, 3, ["Tahun", "Omzet", "Qty", "Profit", "#Order", "YoY Omzet"],
                  [12, 18, 12, 18, 12, 12])
    if len(yearly):
        rr = 4
        for _, y in yearly.iterrows():
            label = f"{int(y['year'])}" + (" (berjalan)" if y["partial"] else "")
            ws1.cell(row=rr, column=1, value=label).font = (BOLD_FONT if y["partial"] else NORMAL_FONT)
            c = ws1.cell(row=rr, column=2, value=float(y["omzet"])); c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws1.cell(row=rr, column=3, value=int(y["qty"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            c = ws1.cell(row=rr, column=4, value=float(y["profit"])); c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws1.cell(row=rr, column=5, value=int(y["orders"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            if pd.notna(y["yoy_omzet"]):
                c = ws1.cell(row=rr, column=6, value=float(y["yoy_omzet"])); c.number_format = FMT_PCT1
                c.font = Font(name=FONT_NAME, size=10, color=("1A7A1A" if y["yoy_omzet"] >= 0 else "B00000"))
            if y["partial"]:
                ws1.cell(row=rr, column=1).fill = YELLOW_FILL
            rr += 1
        ws1.freeze_panes = "A4"

    # --- Sheet 02: monthly ---
    ws2 = wb.create_sheet("02_Tren_Bulanan")
    ws2["A1"] = "TREN BULANAN (SERI WAKTU)"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:E1")
    _style_header(ws2, 3, ["Bulan", "Omzet", "Qty", "Profit", "#Order"], [12, 18, 12, 18, 12])
    if len(monthly):
        rr = 4
        for _, m in monthly.iterrows():
            ws2.cell(row=rr, column=1, value=m["ym"]).font = NORMAL_FONT
            c = ws2.cell(row=rr, column=2, value=float(m["omzet"])); c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=3, value=int(m["qty"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=4, value=float(m["profit"])); c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=5, value=int(m["orders"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            if rr % 2 == 1:
                for ci in range(1, 6):
                    ws2.cell(row=rr, column=ci).fill = LIGHT_FILL
            rr += 1
        ws2.freeze_panes = "A4"

    # --- Sheet 03: seasonality ---
    ws3 = wb.create_sheet("03_Musiman")
    ws3["A1"] = "POLA MUSIMAN PER BULAN (TAHUN PENUH)"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:D1")
    ws3["A2"] = ("Indeks = omzet bulan ini ÷ rata-rata bulanan tahunnya, dirata-rata lintas tahun "
                 "penuh. >1 = konsisten di atas rata-rata (waktu bagus untuk dorong stok & iklan).")
    ws3["A2"].font = SUB_FONT
    ws3.merge_cells("A2:D2")
    _style_header(ws3, 3, ["Bulan", "Indeks Musiman", "Pola", "#Tahun"], [16, 16, 18, 10])
    if len(seasonal):
        rr = 4
        for _, s in seasonal.iterrows():
            idx = s["index"]
            n_years = int(s["n_years"]) if pd.notna(s["n_years"]) else 0
            thin = n_years < TREND_SEASONAL_MIN_YEARS
            ws3.cell(row=rr, column=1, value=s["bulan"]).font = NORMAL_FONT
            if pd.notna(idx):
                c = ws3.cell(row=rr, column=2, value=float(idx)); c.number_format = FMT_IDX; c.font = NORMAL_FONT
            else:
                ws3.cell(row=rr, column=2, value="—").font = NORMAL_FONT
            label = _season_label(idx) + (" (data tipis)" if thin else "")
            ws3.cell(row=rr, column=3, value=label).font = NORMAL_FONT
            ws3.cell(row=rr, column=4, value=n_years).font = NORMAL_FONT
            ws3.cell(row=rr, column=4).number_format = FMT_NUM
            if pd.notna(idx) and not thin:
                if idx >= TREND_PEAK_INDEX:
                    ws3.cell(row=rr, column=3).fill = GREEN_FILL
                elif idx <= TREND_LOW_INDEX:
                    ws3.cell(row=rr, column=3).fill = RED_FILL
            rr += 1
        ws3.freeze_panes = "A4"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
