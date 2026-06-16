"""Dead-stock / capital-release report (`--deadstock`).

Quantifies the capital frozen in slow-moving stock and recommends how to free it.
For every SKU flagged `🔵 Overstock` or `💤 Slow/Dead` by the reorder analysis:

  held value    = sisa_stok × HPP_WA            (capital tied up at cost)
  excess units  = max(0, sisa_stok − target_qty_post_reorder)
  freeable      = excess units × HPP_WA         (the actionable opportunity)

Recommendation:
  - 🧹 Likuidasi   — no demand (velocity ≈ 0 or no sale in `DEADSTOCK_DEAD_DAYS`)
  - 🏷️ Markdown    — slow turnover, cut price to speed it up
  - ⛔ Stop reorder — healthy demand but far above target; stop buying / bundle to clear

Output: output/Analisa_Modal_Beku.xlsx. Built from the reorder metrics — zero-config.
"""
from __future__ import annotations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cashflow import compute_primary_shop
from config import (
    BLUE_FILL_COLOR, DEADSTOCK_DEAD_DAYS, FMT_NUM, FMT_RP, FONT_NAME,
    HEADER_BG_COLOR, HEADER_TEXT_COLOR, LIGHT_GRAY_COLOR, RED_FILL_COLOR,
    TITLE_COLOR, YELLOW_FILL_COLOR,
)

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
RED_FILL = PatternFill("solid", start_color=RED_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
BLUE_FILL = PatternFill("solid", start_color=BLUE_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)


def _cover_txt(cover: float) -> str:
    if cover is None or not np.isfinite(cover):
        return "∞"
    return f"{cover:.0f}"


def analyze_deadstock(reorder_df: pd.DataFrame, hpp_agg: pd.DataFrame,
                      jual: pd.DataFrame, stok: pd.DataFrame, today) -> pd.DataFrame:
    """One row per Overstock / Slow / Dead SKU with held & freeable capital + action."""
    if reorder_df is None or len(reorder_df) == 0:
        return pd.DataFrame()

    hpp_wa = hpp_agg.set_index("SKU")["hpp_wa"].to_dict() if len(hpp_agg) else {}
    last_sale = (jual.groupby("SKU")["tanggal_pesan"].max().to_dict()
                 if jual is not None and "tanggal_pesan" in jual.columns else {})
    primary_shop = compute_primary_shop(stok)

    rows = []
    for _, r in reorder_df.iterrows():
        status = str(r["status"])
        is_over = "Overstock" in status
        is_dead = ("Slow" in status) or ("Dead" in status)
        if not (is_over or is_dead):
            continue
        sisa = float(r["sisa_stok"])
        if sisa <= 0:
            continue

        hpp = hpp_wa.get(r["SKU"], np.nan)
        held = sisa * hpp if pd.notna(hpp) else np.nan
        target = float(r["target_qty_post_reorder"])
        excess = max(0.0, sisa - target)
        freeable = excess * hpp if pd.notna(hpp) else np.nan
        vel = float(r["velocity_used"])
        cover = float(r["months_cover"]) if pd.notna(r["months_cover"]) else np.inf
        ls = last_sale.get(r["SKU"])
        days_idle = (today - ls).days if ls is not None and pd.notna(ls) else None

        no_recent_sale = days_idle is not None and days_idle > DEADSTOCK_DEAD_DAYS
        if vel <= 0 or no_recent_sale:
            rec = "🧹 Likuidasi"
            saran = ("Nyaris tanpa demand — clearance/markdown agresif, keluarkan dari katalog aktif "
                     "agar modal & gudang bebas.")
        elif is_dead:
            rec = "🏷️ Markdown"
            saran = f"Perputaran lambat (~{_cover_txt(cover)} bln cover) — turunkan harga untuk percepat jual."
        else:  # overstock with healthy demand
            rec = "⛔ Stop reorder"
            saran = (f"Demand sehat tapi stok ~{_cover_txt(cover)} bln — stop beli sampai turun ke "
                     f"target ({target:.0f}); pertimbangkan bundle untuk percepat.")

        rows.append({
            "SKU": r["SKU"], "status": status, "supplier": str(primary_shop.get(r["SKU"], "—")) or "—",
            "sisa_stok": sisa, "velocity": vel, "months_cover": cover, "days_idle": days_idle,
            "hpp": hpp, "held_value": held, "target": target, "excess_units": excess,
            "freeable": freeable, "rec": rec, "saran": saran,
        })

    df = pd.DataFrame(rows)
    if len(df):
        df = df.sort_values("freeable", ascending=False, na_position="last").reset_index(drop=True)
    total_held = float(df["held_value"].sum(skipna=True)) if len(df) else 0.0
    total_free = float(df["freeable"].sum(skipna=True)) if len(df) else 0.0
    print(f"✓ Modal beku: {len(df):,} SKU lambat/mati — Rp {total_held:,.0f} tertahan, "
          f"Rp {total_free:,.0f} bisa dibebaskan")
    return df


# ---------------------------------------------------------------------------
def _style_header(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 28


def write_deadstock_report(filepath: Path, df: pd.DataFrame, today) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Ringkasan"
    ws["A1"] = "MODAL BEKU — KAPITAL YANG TERTAHAN DI STOK LAMBAT/MATI"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:E1")
    total_held = float(df["held_value"].sum(skipna=True)) if len(df) else 0.0
    total_free = float(df["freeable"].sum(skipna=True)) if len(df) else 0.0
    n_over = int(df["status"].str.contains("Overstock").sum()) if len(df) else 0
    n_dead = int(df["status"].str.contains("Slow|Dead").sum()) if len(df) else 0
    n_liq = int((df["rec"] == "🧹 Likuidasi").sum()) if len(df) else 0
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  Nilai = qty × HPP_WA (modal di harga beli).  "
                f"'Bisa dibebaskan' = kelebihan di atas target reorder × HPP. Hanya SKU Overstock + Slow/Dead.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:E2")

    r = 4
    for label, val, fmt in [
        ("Total modal tertahan (held)", total_held, FMT_RP),
        ("Modal yang bisa dibebaskan (excess)", total_free, FMT_RP),
        ("SKU Overstock", n_over, FMT_NUM),
        ("SKU Slow/Dead", n_dead, FMT_NUM),
        ("SKU disarankan likuidasi", n_liq, FMT_NUM),
    ]:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = BOLD_FONT
        c.number_format = fmt
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Top peluang pembebasan modal").font = TITLE_FONT
    r += 1
    _style_header(ws, r, ["SKU", "Status", "Bisa dibebaskan", "Cover (bln)", "Aksi"],
                  [34, 16, 16, 11, 16])
    r += 1
    for _, p in (df.head(15).iterrows() if len(df) else iter(())):
        ws.cell(row=r, column=1, value=p["SKU"]).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=p["status"]).font = NORMAL_FONT
        c = ws.cell(row=r, column=3, value=(round(p["freeable"]) if pd.notna(p["freeable"]) else None))
        c.font = NORMAL_FONT
        c.number_format = FMT_RP
        ws.cell(row=r, column=4, value=_cover_txt(p["months_cover"])).font = NORMAL_FONT
        ws.cell(row=r, column=5, value=p["rec"]).font = NORMAL_FONT
        r += 1
    if len(df) == 0:
        ws.cell(row=r, column=1, value="(tidak ada stok lambat/mati — bagus!)").font = BOLD_FONT

    # --- Sheet 01: per-SKU detail ---
    ws2 = wb.create_sheet("01_Modal_Beku_per_SKU")
    ws2["A1"] = "MODAL BEKU PER SKU"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:M1")
    headers = ["SKU", "Status", "Supplier", "Sisa Stok", "Velocity/bln", "Cover (bln)",
               "Idle (hari)", "HPP/pcs", "Nilai Tertahan", "Target", "Excess (unit)",
               "Bisa Dibebaskan", "Aksi"]
    widths = [34, 15, 16, 10, 11, 10, 10, 12, 15, 9, 11, 15, 15]
    _style_header(ws2, 3, headers, widths)
    if len(df):
        rr = 4
        for _, p in df.iterrows():
            vals = [p["SKU"], p["status"], p["supplier"], round(p["sisa_stok"]),
                    round(p["velocity"], 1), _cover_txt(p["months_cover"]),
                    (int(p["days_idle"]) if pd.notna(p["days_idle"]) else None),
                    (round(p["hpp"]) if pd.notna(p["hpp"]) else None),
                    (round(p["held_value"]) if pd.notna(p["held_value"]) else None),
                    round(p["target"]), round(p["excess_units"]),
                    (round(p["freeable"]) if pd.notna(p["freeable"]) else None), p["rec"]]
            fmts = [None, None, None, FMT_NUM, FMT_NUM, None, FMT_NUM, FMT_RP, FMT_RP,
                    FMT_NUM, FMT_NUM, FMT_RP, None]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
                c = ws2.cell(row=rr, column=ci, value=v)
                c.font = NORMAL_FONT
                if fmt:
                    c.number_format = fmt
                if rr % 2 == 1:
                    c.fill = LIGHT_FILL
            if str(p["rec"]).startswith("🧹"):
                ws2.cell(row=rr, column=13).fill = RED_FILL
            elif str(p["rec"]).startswith("🏷️"):
                ws2.cell(row=rr, column=13).fill = YELLOW_FILL
            rr += 1
        ws2.freeze_panes = "A4"

    # --- Sheet 02: per supplier ---
    ws3 = wb.create_sheet("02_Per_Supplier")
    ws3["A1"] = "MODAL BEKU PER SUPPLIER — SIAPA YANG STOKNYA NUMPUK"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:D1")
    _style_header(ws3, 3, ["Supplier", "Jumlah SKU", "Nilai Tertahan", "Bisa Dibebaskan"],
                  [26, 12, 18, 18])
    if len(df):
        by_sup = (df.groupby("supplier")
                    .agg(n=("SKU", "count"), held=("held_value", "sum"), free=("freeable", "sum"))
                    .reset_index().sort_values("free", ascending=False))
        rr = 4
        for _, s in by_sup.iterrows():
            ws3.cell(row=rr, column=1, value=str(s["supplier"])).font = NORMAL_FONT
            c = ws3.cell(row=rr, column=2, value=int(s["n"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=3, value=float(s["held"])); c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=4, value=float(s["free"])); c.number_format = FMT_RP; c.font = NORMAL_FONT
            if rr % 2 == 1:
                for ci in range(1, 5):
                    ws3.cell(row=rr, column=ci).fill = LIGHT_FILL
            rr += 1
        ws3.freeze_panes = "A4"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
