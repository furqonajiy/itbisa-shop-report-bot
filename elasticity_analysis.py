"""Price-elasticity miner (`--elasticity`).

Estimates each SKU's price elasticity of demand from its own sales history and
recommends a pricing direction. For every SKU with enough monthly observations
and real price movement, it fits a log-log OLS:

    ln(qty_month) = a + b · ln(price_month)        →  b = price elasticity

  - |b| < `ELASTICITY_INELASTIC_THRESHOLD` (inelastic) → demand barely reacts to
    price → 🔼 raising price lifts revenue.
  - |b| ≥ threshold (elastic) → demand is price-sensitive → 🔽 raising risks volume,
    a cut can grow it.
  - b ≥ 0 (qty & price move together) → ↔ inconclusive (other factors).

Observational elasticity is confounded (seasonality, stock-outs, promos, competitor
moves), so low-`R²` / thin fits are flagged as low confidence and NOT surfaced as
"raise" candidates. Output: output/Analisa_Elastisitas_Harga.xlsx. Zero-config.
"""
from __future__ import annotations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    BLUE_FILL_COLOR, ELASTICITY_INELASTIC_THRESHOLD, ELASTICITY_MIN_MONTHS,
    ELASTICITY_MIN_PRICE_CV, ELASTICITY_MIN_R2, ELASTICITY_PRICE_NUDGE, FMT_NUM,
    FMT_RP, FONT_NAME, GREEN_FILL_COLOR, HEADER_BG_COLOR, HEADER_TEXT_COLOR,
    LIGHT_GRAY_COLOR, TITLE_COLOR, YELLOW_FILL_COLOR,
)

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
GREEN_FILL = PatternFill("solid", start_color=GREEN_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
BLUE_FILL = PatternFill("solid", start_color=BLUE_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)
FMT_PCT1 = '0.0%'
FMT_E = '0.00'


def analyze_elasticity(jual: pd.DataFrame, hpp_agg: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Returns (per_sku, monthly_detail).

    per_sku: one row per SKU — elasticity, R², confidence, classification, current
    price/margin, the modeled +nudge price scenario, and a recommendation.
    monthly_detail: the (SKU, month, price, qty) points used to fit analyzable SKUs."""
    needed = {"SKU", "tanggal_pesan", "qty_jual", "omzet"}
    if jual is None or len(jual) == 0 or not needed.issubset(jual.columns):
        return pd.DataFrame(), pd.DataFrame()

    hpp_wa = hpp_agg.set_index("SKU")["hpp_wa"].to_dict() if len(hpp_agg) else {}
    j = jual[(jual["qty_jual"] > 0) & (jual["omzet"] > 0) & jual["tanggal_pesan"].notna()].copy()
    j["m"] = j["tanggal_pesan"].dt.to_period("M")
    g = (j.groupby(["SKU", "m"]).agg(qty=("qty_jual", "sum"), omzet=("omzet", "sum"))
           .reset_index())
    g["price"] = g["omzet"] / g["qty"]

    nudge = ELASTICITY_PRICE_NUDGE
    rows, detail_skus = [], []
    for sku, sub in g.groupby("SKU"):
        sub = sub.sort_values("m")
        n = len(sub)
        prices = sub["price"].to_numpy(dtype=float)
        qtys = sub["qty"].to_numpy(dtype=float)
        price_cv = float(prices.std() / prices.mean()) if prices.mean() > 0 else 0.0
        last = sub.iloc[-1]
        cur_price = float(last["price"])
        recent = sub.tail(min(3, n))
        avg_recent_omzet = float(recent["omzet"].mean())
        hpp = hpp_wa.get(sku, np.nan)

        base = {"SKU": sku, "n_months": n, "price_cv": price_cv, "cur_price": cur_price,
                "hpp": hpp, "elasticity": np.nan, "r2": np.nan, "confidence": "—",
                "qty_chg": np.nan, "rev_chg": np.nan, "est_uplift": np.nan,
                "avg_recent_omzet": avg_recent_omzet}

        x = np.log(prices)
        if n < ELASTICITY_MIN_MONTHS or price_cv < ELASTICITY_MIN_PRICE_CV or x.std() == 0:
            base.update(rec="⚪ Data kurang",
                        saran=f"Belum cukup variasi harga / bulan (n={n}, CV harga={price_cv:.0%}) untuk diukur.")
            rows.append(base)
            continue

        y = np.log(qtys)
        b, a = np.polyfit(x, y, 1)
        yhat = a + b * x
        ss_res = float(((y - yhat) ** 2).sum())
        ss_tot = float(((y - y.mean()) ** 2).sum())
        r2 = (1 - ss_res / ss_tot) if ss_tot > 0 else 0.0

        conf = ("Tinggi" if (r2 >= 0.6 and n >= ELASTICITY_MIN_MONTHS + 3)
                else "Sedang" if r2 >= ELASTICITY_MIN_R2 else "Rendah")
        qty_chg = (1 + nudge) ** b - 1.0          # modeled qty change for +nudge price
        rev_chg = (1 + nudge) ** (1 + b) - 1.0     # revenue ∝ price^(1+b)

        if b >= 0:
            rec = "↔ Tak konklusif"
            saran = "Qty & harga bergerak searah — ada faktor lain (musim/stok); elastisitas tak terisolasi."
            est_uplift = np.nan
        elif abs(b) < ELASTICITY_INELASTIC_THRESHOLD:
            rec = "🔼 Naikkan harga"
            saran = (f"Inelastis (e={b:.2f}): +{nudge:.0%} harga ≈ {qty_chg:+.0%} qty, "
                     f"{rev_chg:+.0%} omzet → ada ruang naik.")
            est_uplift = rev_chg * avg_recent_omzet if conf != "Rendah" else np.nan
        else:
            rec = "🔽 Sensitif harga"
            saran = (f"Elastis (e={b:.2f}): +{nudge:.0%} harga ≈ {qty_chg:+.0%} qty → hati-hati naik; "
                     f"turun harga bisa tambah volume.")
            est_uplift = np.nan

        if conf == "Rendah" and rec == "🔼 Naikkan harga":
            saran += " (confidence rendah — validasi dulu sebelum naik)."

        base.update(elasticity=float(b), r2=float(r2), confidence=conf, qty_chg=float(qty_chg),
                    rev_chg=float(rev_chg), est_uplift=est_uplift, rec=rec, saran=saran)
        rows.append(base)
        sd = sub[["m", "price", "qty", "omzet"]].copy()
        sd.insert(0, "SKU", sku)
        detail_skus.append(sd)

    per_sku = pd.DataFrame(rows)
    if len(per_sku):
        # raise-candidates (high confidence) first, by modeled uplift, then everyone else
        per_sku["_rank"] = np.where(per_sku["rec"] == "🔼 Naikkan harga",
                                    per_sku["est_uplift"].fillna(0.0), -1.0)
        per_sku = per_sku.sort_values(["_rank", "n_months"], ascending=False).drop(columns="_rank").reset_index(drop=True)
    detail = (pd.concat(detail_skus, ignore_index=True) if detail_skus else pd.DataFrame())
    if len(detail):
        detail["m"] = detail["m"].astype(str)

    n_fit = int((per_sku["rec"] != "⚪ Data kurang").sum()) if len(per_sku) else 0
    n_raise = int((per_sku["rec"] == "🔼 Naikkan harga").sum()) if len(per_sku) else 0
    print(f"✓ Elastisitas harga: {len(per_sku):,} SKU ({n_fit} bisa diukur) — "
          f"{n_raise} kandidat naik harga (inelastis)")
    return per_sku, detail


# ---------------------------------------------------------------------------
def _style_header(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 28


def write_elasticity_report(filepath: Path, per_sku: pd.DataFrame,
                            detail: pd.DataFrame, today) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Ringkasan"
    ws["A1"] = "ELASTISITAS HARGA — DI MANA ADA RUANG NAIK / TURUN HARGA"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  e = elastisitas (slope log-log qty vs harga). "
                f"|e|<1 inelastis (naikkan), |e|>1 elastis (hati-hati). Data observasional → "
                f"confidence Rendah tidak dijadikan rekomendasi naik.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:E2")

    r = 4
    if len(per_sku):
        n_fit = int((per_sku["rec"] != "⚪ Data kurang").sum())
        n_raise = int((per_sku["rec"] == "🔼 Naikkan harga").sum())
        n_elastic = int((per_sku["rec"] == "🔽 Sensitif harga").sum())
        n_low = int((per_sku["confidence"] == "Rendah").sum())
        uplift = float(per_sku["est_uplift"].sum(skipna=True))
        for label, val, fmt in [
            ("SKU dianalisa", len(per_sku), FMT_NUM),
            ("SKU bisa diukur (cukup data)", n_fit, FMT_NUM),
            ("🔼 Kandidat naik harga (inelastis)", n_raise, FMT_NUM),
            ("🔽 Sensitif harga (elastis)", n_elastic, FMT_NUM),
            ("Est. tambahan omzet/bln (modeled, conf≥Sedang)", uplift, FMT_RP),
        ]:
            ws.cell(row=r, column=1, value=label).font = BOLD_FONT
            c = ws.cell(row=r, column=2, value=val)
            c.font = BOLD_FONT
            c.number_format = fmt
            r += 1
        ws.cell(row=r, column=1, value=f"*) {n_low} SKU confidence Rendah — perlu validasi, "
                "bukan rekomendasi pasti. Elastisitas observasional bisa terpengaruh musim/stok/kompetitor.").font = SUB_FONT
        r += 2

        ws.cell(row=r, column=1, value="Top kandidat naik harga (inelastis, confidence ≥ Sedang)").font = TITLE_FONT
        r += 1
        _style_header(ws, r, ["SKU", "Elastisitas", "Confidence", "+10% → omzet", "Est. omzet/bln"],
                      [34, 11, 11, 13, 15])
        r += 1
        top = per_sku[(per_sku["rec"] == "🔼 Naikkan harga") & (per_sku["confidence"] != "Rendah")].head(15)
        for _, p in top.iterrows():
            ws.cell(row=r, column=1, value=p["SKU"]).font = NORMAL_FONT
            c = ws.cell(row=r, column=2, value=round(p["elasticity"], 2)); c.number_format = FMT_E; c.font = NORMAL_FONT
            ws.cell(row=r, column=3, value=p["confidence"]).font = NORMAL_FONT
            c = ws.cell(row=r, column=4, value=float(p["rev_chg"])); c.number_format = FMT_PCT1; c.font = NORMAL_FONT
            c = ws.cell(row=r, column=5, value=(round(p["est_uplift"]) if pd.notna(p["est_uplift"]) else None))
            c.number_format = FMT_RP; c.font = NORMAL_FONT
            ws.cell(row=r, column=1).fill = GREEN_FILL
            r += 1
        if len(top) == 0:
            ws.cell(row=r, column=1, value="(belum ada kandidat naik harga dengan confidence cukup)").font = NORMAL_FONT

    # --- Sheet 01: per-SKU recommendation ---
    ws2 = wb.create_sheet("01_Rekomendasi_Harga")
    ws2["A1"] = "REKOMENDASI HARGA PER SKU (ELASTISITAS)"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:L1")
    headers = ["SKU", "#Bulan", "CV Harga", "Harga Skrg", "Elastisitas", "R²", "Confidence",
               "+10% → Qty", "+10% → Omzet", "Est. Omzet/bln", "Rekomendasi", "Saran"]
    widths = [34, 8, 9, 12, 11, 7, 10, 11, 12, 14, 16, 50]
    _style_header(ws2, 3, headers, widths)
    if len(per_sku):
        rr = 4
        for _, p in per_sku.iterrows():
            vals = [p["SKU"], int(p["n_months"]), float(p["price_cv"]),
                    (round(p["cur_price"]) if pd.notna(p["cur_price"]) else None),
                    (round(p["elasticity"], 2) if pd.notna(p["elasticity"]) else None),
                    (round(p["r2"], 2) if pd.notna(p["r2"]) else None), p["confidence"],
                    (float(p["qty_chg"]) if pd.notna(p["qty_chg"]) else None),
                    (float(p["rev_chg"]) if pd.notna(p["rev_chg"]) else None),
                    (round(p["est_uplift"]) if pd.notna(p["est_uplift"]) else None),
                    p["rec"], p["saran"]]
            fmts = [None, FMT_NUM, FMT_PCT1, FMT_RP, FMT_E, FMT_E, None, FMT_PCT1, FMT_PCT1,
                    FMT_RP, None, None]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
                c = ws2.cell(row=rr, column=ci, value=v)
                c.font = NORMAL_FONT
                c.alignment = Alignment(vertical="top", wrap_text=(ci == len(vals)))
                if fmt:
                    c.number_format = fmt
                if rr % 2 == 1:
                    c.fill = LIGHT_FILL
            if str(p["rec"]).startswith("🔼") and p["confidence"] != "Rendah":
                ws2.cell(row=rr, column=11).fill = GREEN_FILL
            elif str(p["rec"]).startswith("🔽"):
                ws2.cell(row=rr, column=11).fill = YELLOW_FILL
            rr += 1
        ws2.freeze_panes = "A4"

    # --- Sheet 02: monthly points used to fit ---
    ws3 = wb.create_sheet("02_Data_Bulanan")
    ws3["A1"] = "DATA BULANAN (HARGA & QTY) — DASAR PERHITUNGAN"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:E1")
    _style_header(ws3, 3, ["SKU", "Bulan", "Harga rata2", "Qty", "Omzet"],
                  [34, 12, 14, 11, 16])
    if len(detail):
        rr = 4
        for _, d in detail.iterrows():
            ws3.cell(row=rr, column=1, value=d["SKU"]).font = NORMAL_FONT
            ws3.cell(row=rr, column=2, value=str(d["m"])).font = NORMAL_FONT
            c = ws3.cell(row=rr, column=3, value=round(float(d["price"]))); c.number_format = FMT_RP; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=4, value=int(d["qty"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=5, value=round(float(d["omzet"]))); c.number_format = FMT_RP; c.font = NORMAL_FONT
            if rr % 2 == 1:
                for ci in range(1, 6):
                    ws3.cell(row=rr, column=ci).fill = LIGHT_FILL
            rr += 1
        ws3.freeze_panes = "A4"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
