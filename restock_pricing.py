"""Restock price evaluator.

Given a SKU you're considering restocking (offered price + Toko + competitor
price range), this:
  1) predicts the landed HPP in Rupiah (from a raw RMB price, calibrated on your
     own import history; or uses a given HPP IDR directly),
  2) judges the cost vs the SKU's historical HPP (too high / reasonable / cheaper),
  3) recommends the selling price per marketplace (Shopee/Tokopedia/Tiktok) that
     yields a target net profit AFTER the platform fee, and
  4) decides — restock & sell, sell-but-thin, or don't sell — against the
     competitor price range.
Input: data/restock_check.xlsx (auto-template if missing). Output:
output/Analisa_Restock_Check.xlsx.
"""
from __future__ import annotations
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    COL_RC_HPP, COL_RC_KMAX, COL_RC_KMIN, COL_RC_NOTE, COL_RC_RMB, COL_RC_SKU,
    COL_RC_TOKO, COL_STOK_QTY, COL_STOK_TOTAL_HPP, FMT_NUM, FMT_RP, FONT_NAME,
    GREEN_FILL_COLOR, HEADER_BG_COLOR, HEADER_TEXT_COLOR, LIGHT_GRAY_COLOR,
    OCISTOK_KEYWORDS, PLATFORM_FEE_FALLBACK, RED_FILL_COLOR, RESTOCK_CHECK_SHEET,
    RESTOCK_COST_TOL, RESTOCK_PLATFORMS, RESTOCK_RMB_MIN_LOTS,
    RESTOCK_TARGET_NET_MARKUP, RMB_SPOT_FX_IDR, RMB_TO_IDR_FALLBACK, STOK_SHEET,
    TITLE_COLOR, YELLOW_FILL_COLOR,
)
from data_loader import resolve_sheet

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
GREEN_FILL = PatternFill("solid", start_color=GREEN_FILL_COLOR)
YELLOW_FILL = PatternFill("solid", start_color=YELLOW_FILL_COLOR)
RED_FILL = PatternFill("solid", start_color=RED_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)


def _norm_sku(s) -> str:
    return str(s).upper().strip()


def _num(x) -> float:
    v = pd.to_numeric(x, errors="coerce")
    return float(v) if pd.notna(v) else float("nan")


# ---------------------------------------------------------------------------
# Input template / loading
# ---------------------------------------------------------------------------
def create_restock_template(filepath: Path) -> None:
    """Create data/restock_check.xlsx with headers, a worked example, and notes."""
    wb = Workbook()
    ws = wb.active
    ws.title = RESTOCK_CHECK_SHEET
    headers = [COL_RC_SKU, COL_RC_TOKO, COL_RC_RMB, COL_RC_HPP,
               COL_RC_KMIN, COL_RC_KMAX, COL_RC_NOTE]
    widths = [38, 18, 12, 12, 15, 15, 40]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 28
    # worked example
    example = ["ITBISA-SERVO-SG90-180DEG", "Ocistok/Martkita", 3.5, None, 17000, 20000,
               "Contoh: beli 3.5 RMB, harga pesaing 17-20rb"]
    for i, v in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=v).font = NORMAL_FONT

    # Instructions go on a SEPARATE sheet so the data sheet stays pristine
    # (every row in RestockCheck is a real SKU to analyze).
    guide = wb.create_sheet("Cara_Pakai")
    guide.column_dimensions["A"].width = 100
    lines = [
        "Cara isi sheet 'RestockCheck' (satu baris = satu SKU yang mau di-restock):",
        f"• {COL_RC_SKU}: persis sama dengan SKU di Stok/Jual.",
        f"• {COL_RC_TOKO}: calon supplier (Ocistok/Martkita, AliExpress, Shopee, Tokopedia, dst).",
        f"• {COL_RC_RMB}: harga supplier per pcs dalam RMB (Yuan) — untuk impor; HPP IDR diprediksi otomatis.",
        f"• {COL_RC_HPP}: HPP per pcs dalam Rupiah (sudah termasuk ongkir/impor) — isi kalau sudah tahu; menimpa prediksi RMB.",
        f"• {COL_RC_KMIN} / {COL_RC_KMAX}: rentang harga jual kompetitor di marketplace (Rupiah).",
        "",
        "Minimal isi salah satu dari Harga RMB ATAU HPP IDR. Lalu jalankan:  python main.py --restock-check",
    ]
    for j, n in enumerate(lines, start=1):
        guide.cell(row=j, column=1, value=n).font = BOLD_FONT if j == 1 else NORMAL_FONT

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Template restock dibuat: {filepath}")


def load_restock_check(filepath: Path) -> pd.DataFrame:
    """Load restock_check.xlsx → normalized rows. Empty DataFrame if file missing."""
    if not filepath.exists():
        return pd.DataFrame()
    df = pd.read_excel(filepath, sheet_name=RESTOCK_CHECK_SHEET)
    df = df[df[COL_RC_SKU].notna()].copy()
    if len(df) == 0:
        return df
    out = pd.DataFrame({
        "sku": df[COL_RC_SKU].map(_norm_sku),
        "toko": df.get(COL_RC_TOKO, "").astype(str).str.strip(),
        "rmb": df.get(COL_RC_RMB).map(_num) if COL_RC_RMB in df else np.nan,
        "hpp_idr": df.get(COL_RC_HPP).map(_num) if COL_RC_HPP in df else np.nan,
        "kmin": df.get(COL_RC_KMIN).map(_num) if COL_RC_KMIN in df else np.nan,
        "kmax": df.get(COL_RC_KMAX).map(_num) if COL_RC_KMAX in df else np.nan,
        "note": df.get(COL_RC_NOTE, "").astype(str).replace("nan", "").str.strip(),
    })
    # A real check needs a price (RMB or HPP IDR) — drops any stray/notes rows.
    out = out[(out["rmb"] > 0) | (out["hpp_idr"] > 0)].reset_index(drop=True)
    print(f"✓ Loaded {len(out):,} baris restock-check dari {filepath.name}")
    return out


# ---------------------------------------------------------------------------
# Calibration from history
# ---------------------------------------------------------------------------
def load_rmb_hpp_history(stok_files: list[Path]) -> pd.DataFrame:
    """Per-lot (SKU, rmb_unit, hpp_idr_pc) from Stok: rmb from the
    Keterangan "(x RMB)" note, landed HPP/pc = Total HPP (Rp) ÷ qty."""
    rows = []
    for fp in stok_files:
        df = pd.read_excel(fp, sheet_name=resolve_sheet(fp, STOK_SHEET), header=1)
        ket_cols = [c for c in df.columns if str(c).strip().startswith("Keterangan")]
        ket = ket_cols[0] if ket_cols else None
        tok_cols = [c for c in df.columns if str(c).strip().lower().startswith("toko")]
        tok = tok_cols[0] if tok_cols else None
        df = df[df["SKU"].notna()]
        for _, r in df.iterrows():
            qty = _num(r.get(COL_STOK_QTY))
            thp = _num(r.get(COL_STOK_TOTAL_HPP))
            hpp_pc = thp / qty if (qty and qty > 0 and pd.notna(thp)) else float("nan")
            rmb = float("nan")
            if ket is not None:
                m = re.search(r"\(([\d.]+)\s*RMB\)", str(r.get(ket)))
                if m:
                    rmb = float(m.group(1))
            rows.append((_norm_sku(r["SKU"]), str(r.get(tok)) if tok else "", rmb, hpp_pc))
    return pd.DataFrame(rows, columns=["SKU", "toko", "rmb", "hpp_idr_pc"])


def compute_rmb_factor(hist: pd.DataFrame) -> tuple[pd.Series, float]:
    """(per_sku_factor, global_factor) = FINAL landed IDR per 1 RMB. Calibrated on
    the Ocistok/Martkita forwarder channel only — so it embeds that forwarder's
    margin + shipping + import (it's realized HPP/pc ÷ the (x RMB) price)."""
    d = hist[(hist["rmb"] > 0) & (hist["hpp_idr_pc"] > 0)].copy()
    if len(d) == 0:
        return pd.Series(dtype=float), float(RMB_TO_IDR_FALLBACK)
    tl = d["toko"].astype(str).str.lower()
    oci = d[tl.apply(lambda t: any(k in t for k in OCISTOK_KEYWORDS))]
    d = oci if len(oci) else d            # prefer the forwarder channel; else all RMB lots
    d["factor"] = d["hpp_idr_pc"] / d["rmb"]
    glob = float(d["factor"].median())
    g = d.groupby("SKU")["factor"]
    per_sku = g.median()[g.count() >= RESTOCK_RMB_MIN_LOTS]
    return per_sku, glob


def compute_platform_fees(jual: pd.DataFrame) -> dict[str, float]:
    """Per-platform fee = |admin| / omzet from all-time Jual (admin = tambahan +
    kode_unik, stored negative). Fallback to PLATFORM_FEE_FALLBACK for thin platforms."""
    fees = {}
    if jual is not None and len(jual):
        admin = jual["tambahan"].fillna(0) + jual["kode_unik"].fillna(0)
        j = jual.assign(_admin=admin)
        for p in RESTOCK_PLATFORMS:
            sub = j[j["akun_penjual"].astype(str).str.strip().str.lower() == p.lower()]
            omz = sub["omzet"].sum()
            if omz > 0 and len(sub) >= 10:
                fees[p] = float(-sub["_admin"].sum() / omz)
    for p in RESTOCK_PLATFORMS:
        fees.setdefault(p, PLATFORM_FEE_FALLBACK.get(p, 0.13))
    return fees


# ---------------------------------------------------------------------------
# Per-SKU analysis
# ---------------------------------------------------------------------------
def _sell_for_target(hpp: float, fee: float, markup: float) -> float:
    """Selling price so that net profit (after fee) = markup × hpp."""
    return hpp * (1 + markup) / (1 - fee)


def analyze_restock(checks: pd.DataFrame, hpp_agg: pd.DataFrame,
                    hist: pd.DataFrame, fees: dict[str, float]) -> pd.DataFrame:
    per_sku_factor, global_factor = compute_rmb_factor(hist)
    hpp_wa = hpp_agg.set_index("SKU")["hpp_wa"].to_dict() if len(hpp_agg) else {}
    rmb_hist = (hist[hist["rmb"] > 0].groupby("SKU")["rmb"].apply(list).to_dict())
    markup = RESTOCK_TARGET_NET_MARKUP
    best_platform = min(fees, key=fees.get)  # lowest fee

    out = []
    for _, r in checks.iterrows():
        sku = r["sku"]
        rec = {"SKU": sku, "Toko": r["toko"], "kmin": r["kmin"], "kmax": r["kmax"],
               "note": r["note"]}

        # 1) landed HPP
        if pd.notna(r["hpp_idr"]) and r["hpp_idr"] > 0:
            hpp = float(r["hpp_idr"]); rec["input"] = f"HPP IDR diberikan Rp{hpp:,.0f}"
            rec["factor"] = np.nan
        elif pd.notna(r["rmb"]) and r["rmb"] > 0:
            factor = float(per_sku_factor.get(sku, global_factor))
            hpp = r["rmb"] * factor
            src = "histori SKU" if sku in per_sku_factor.index else "global"
            rec["input"] = (f"{r['rmb']:g} RMB × {factor:,.0f} landed ({src}; "
                            f"sudah termasuk margin Martkita + ongkir + impor)")
            rec["factor"] = factor
        else:
            rec.update(hpp=np.nan, verdict="⚠ Tidak ada Harga RMB / HPP IDR",
                       keputusan="—", saran="Isi Harga RMB atau HPP IDR.")
            for p in RESTOCK_PLATFORMS:
                rec[f"jual_{p}"] = np.nan
            out.append(rec); continue
        rec["hpp"] = round(hpp)

        # 2) cost verdict vs historical HPP
        wa = hpp_wa.get(sku)
        rec["hpp_histori"] = round(wa) if wa else np.nan
        if wa and wa > 0:
            ratio = hpp / wa
            if ratio < 1 - RESTOCK_COST_TOL:
                rec["verdict"] = f"🟢 Lebih murah dari biasanya ({(1-ratio)*100:.0f}% di bawah HPP histori)"
            elif ratio <= 1 + RESTOCK_COST_TOL:
                rec["verdict"] = "🟢 Wajar (≈ HPP histori)"
            else:
                rec["verdict"] = f"🔴 Lebih mahal {(ratio-1)*100:.0f}% dari HPP histori"
        else:
            rec["verdict"] = "⚪ SKU baru / tanpa histori HPP"

        # 3) sell price per platform for the target net markup
        for p in RESTOCK_PLATFORMS:
            rec[f"jual_{p}"] = round(_sell_for_target(hpp, fees[p], markup))

        # 4) decision vs competitor range
        kmin, kmax = r["kmin"], r["kmax"]
        bf = fees[best_platform]
        sell_min_best = _sell_for_target(hpp, bf, markup)
        if pd.isna(kmax):
            rec["keputusan"] = "🟢 Jual (tanpa data kompetitor)"
            rec["saran"] = (f"Tidak ada harga kompetitor. Jual minimal Rp{round(sell_min_best):,} "
                            f"di {best_platform} untuk net {markup*100:.0f}%.")
        else:
            net_at_kmax = kmax * (1 - bf) - hpp
            if sell_min_best <= kmax:
                # can hit target within the market
                pos = "di BAWAH harga pasar termurah" if (pd.notna(kmin) and sell_min_best <= kmin) else "masih dalam rentang pasar"
                head = f" (net @ pasar terendah Rp{round((kmin*(1-bf)-hpp)/hpp*100) if pd.notna(kmin) else 0}%)" if pd.notna(kmin) else ""
                netpct_kmin = ((kmin * (1 - bf) - hpp) / hpp * 100) if pd.notna(kmin) else None
                rec["keputusan"] = "🟢 Restock & jual"
                rec["saran"] = (f"Untung. Jual ≥ Rp{round(sell_min_best):,} ({best_platform}) "
                                f"untuk net {markup*100:.0f}% — {pos}." +
                                (f" Di harga pasar terendah Rp{kmin:,.0f}, net ≈ {netpct_kmin:.0f}% HPP." if netpct_kmin is not None else ""))
            elif net_at_kmax >= 0:
                rec["keputusan"] = "🟡 Tipis — keputusan Anda"
                rec["saran"] = (f"Tidak bisa capai net {markup*100:.0f}% di harga pasar. "
                                f"Di harga pasar tertinggi Rp{kmax:,.0f}, net hanya ≈ {net_at_kmax/hpp*100:.0f}% HPP "
                                f"(Rp{round(net_at_kmax):,}/pcs) di {best_platform}.")
            else:
                rec["keputusan"] = "🔴 Jangan jual"
                rec["saran"] = (f"Rugi walau di harga pasar tertinggi Rp{kmax:,.0f}: "
                                f"net ≈ Rp{round(net_at_kmax):,}/pcs di {best_platform}. "
                                f"Cari supplier lebih murah atau lewati item ini.")
        out.append(rec)

    df = pd.DataFrame(out)
    n = {"🟢": 0, "🟡": 0, "🔴": 0}
    for k in df.get("keputusan", []):
        for e in n:
            if str(k).startswith(e):
                n[e] += 1
    print(f"✓ Restock-check: {len(df)} SKU — 🟢 {n['🟢']} jual, 🟡 {n['🟡']} tipis, 🔴 {n['🔴']} jangan jual")
    return df


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------
def write_restock_report(filepath: Path, results: pd.DataFrame,
                         fees: dict[str, float], factor_global: float, today) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Restock_Check"
    ws["A1"] = "ANALISA HARGA RESTOCK — BELI & JUAL"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:K1")
    feetxt = ", ".join(f"{p} {fees[p]*100:.0f}%" for p in RESTOCK_PLATFORMS)
    uplift = (factor_global / RMB_SPOT_FX_IDR - 1) * 100 if RMB_SPOT_FX_IDR else 0
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  Target net ≥ {RESTOCK_TARGET_NET_MARKUP*100:.0f}% HPP "
                f"setelah fee  |  Fee marketplace (dari data): {feetxt}  |  "
                f"Prediksi HPP landed: 1 RMB ≈ Rp{factor_global:,.0f} "
                f"(kurs spot ≈ Rp{RMB_SPOT_FX_IDR:,.0f}, +{uplift:.0f}% = margin Martkita + ongkir + impor)")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:K2")
    ws["A3"] = ("Landed HPP = biaya final per pcs (sudah termasuk margin forwarder + ongkir + impor), dikalibrasi "
                "dari histori lot Ocistok/Martkita. Verdict = vs HPP histori SKU. 'Jual …' = harga jual minimal "
                "agar net ≥ target setelah fee marketplace. Keputusan = vs rentang harga kompetitor.")
    ws["A3"].font = SUB_FONT
    ws.merge_cells("A3:K3")

    if results is None or len(results) == 0:
        ws["A5"] = "(belum ada baris di restock_check.xlsx)"
        ws["A5"].font = BOLD_FONT
        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        print(f"✓ Menulis laporan ke {filepath} (kosong)")
        return

    headers = ["SKU", "Toko", "Input / Prediksi HPP", "Landed HPP", "HPP Histori",
               "Verdict (vs histori)", "Kompetitor"]
    headers += [f"Jual {p}" for p in RESTOCK_PLATFORMS]
    headers += ["Keputusan", "Saran"]
    widths = [34, 16, 26, 12, 12, 30, 16] + [12] * len(RESTOCK_PLATFORMS) + [22, 60]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[5].height = 30

    fill_for = {"🟢": GREEN_FILL, "🟡": YELLOW_FILL, "🔴": RED_FILL}
    for ridx, (_, r) in enumerate(results.iterrows()):
        row = 6 + ridx
        komp = (f"Rp{r['kmin']:,.0f}–Rp{r['kmax']:,.0f}" if pd.notna(r.get("kmin")) and pd.notna(r.get("kmax"))
                else (f"≤Rp{r['kmax']:,.0f}" if pd.notna(r.get("kmax")) else "—"))
        vals = [r["SKU"], r.get("Toko", ""), r.get("input", ""),
                r.get("hpp"), r.get("hpp_histori"), r.get("verdict", ""), komp]
        vals += [r.get(f"jual_{p}") for p in RESTOCK_PLATFORMS]
        vals += [r.get("keputusan", ""), r.get("saran", "")]
        fmts = [None, None, None, FMT_RP, FMT_RP, None, None] + [FMT_RP] * len(RESTOCK_PLATFORMS) + [None, None]
        for cidx, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=row, column=cidx, value=(None if (isinstance(v, float) and pd.isna(v)) else v))
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(cidx == len(vals)))
            if fmt:
                cell.number_format = fmt
            if ridx % 2 == 1:
                cell.fill = LIGHT_FILL
        # color the Keputusan cell
        kcell = ws.cell(row=row, column=len(headers) - 1)
        for e, fill in fill_for.items():
            if str(r.get("keputusan", "")).startswith(e):
                kcell.fill = fill
    ws.freeze_panes = "A6"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
