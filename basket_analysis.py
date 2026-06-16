"""Bundle / cross-sell market-basket analysis (`--bundle`).

Finds SKU pairs frequently bought in the same order (grouped by `Invoice`) so you
can build bundles / "frequently bought together" offers. For each pair:

  support      = number of orders containing BOTH SKUs
  confidence   = P(B in order | A in order) = support / orders-with-A  (both ways)
  lift         = support × N_orders / (orders-with-A × orders-with-B)
                 (> 1 = bought together more than chance)

Output: output/Analisa_Bundle_CrossSell.xlsx. Zero-config (built from BisaJual).
"""
from __future__ import annotations
from collections import Counter
from itertools import combinations
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    BASKET_MIN_PAIR_SUPPORT, BASKET_TOP_N, FMT_NUM, FONT_NAME, HEADER_BG_COLOR,
    HEADER_TEXT_COLOR, LIGHT_GRAY_COLOR, TITLE_COLOR,
)

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)

FMT_PCT = '0.0%'
FMT_X = '0.0"x"'


def analyze_baskets(jual: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """Returns (pairs_df, cross_sell_df, stats).

    pairs_df: SKU pairs with support / confidence (both directions) / lift.
    cross_sell_df: per SKU, its best co-bought partner (by confidence).
    stats: {orders, multi_orders, pairs}."""
    if jual is None or len(jual) == 0 or "Invoice" not in jual.columns:
        return pd.DataFrame(), pd.DataFrame(), {"orders": 0, "multi_orders": 0, "pairs": 0}

    d = jual[jual["Invoice"].notna() & jual["SKU"].notna()]
    baskets = d.groupby("Invoice")["SKU"].apply(lambda s: sorted(set(s)))
    n_orders = int(len(baskets))

    sku_orders = Counter()          # orders containing each SKU
    pair_count = Counter()          # orders containing each unordered pair
    multi_orders = 0
    for skus in baskets:
        for s in skus:
            sku_orders[s] += 1
        if len(skus) > 1:
            multi_orders += 1
            for a, b in combinations(skus, 2):
                pair_count[(a, b)] += 1

    rows = []
    for (a, b), c in pair_count.items():
        if c < BASKET_MIN_PAIR_SUPPORT:
            continue
        ca, cb = sku_orders[a], sku_orders[b]
        lift = (c * n_orders) / (ca * cb) if ca and cb else 0.0
        rows.append({
            "sku_a": a, "sku_b": b, "support": c,
            "conf_ab": c / ca if ca else 0.0,    # buyers of A who also buy B
            "conf_ba": c / cb if cb else 0.0,    # buyers of B who also buy A
            "lift": lift,
        })
    pairs = pd.DataFrame(rows)
    if len(pairs):
        pairs = pairs.sort_values(["support", "lift"], ascending=False).reset_index(drop=True)

    # Per-SKU best partner (directed: source → partner, ranked by confidence then support)
    directed = []
    for _, p in pairs.iterrows():
        directed.append((p["sku_a"], p["sku_b"], p["conf_ab"], p["support"], p["lift"]))
        directed.append((p["sku_b"], p["sku_a"], p["conf_ba"], p["support"], p["lift"]))
    cross = pd.DataFrame(directed, columns=["SKU", "partner", "confidence", "support", "lift"])
    if len(cross):
        cross = (cross.sort_values(["SKU", "confidence", "support"], ascending=[True, False, False])
                      .groupby("SKU", as_index=False).head(1).reset_index(drop=True))
        cross = cross.sort_values(["support", "confidence"], ascending=False).reset_index(drop=True)

    stats = {"orders": n_orders, "multi_orders": multi_orders, "pairs": int(len(pairs))}
    print(f"✓ Bundle/cross-sell: {n_orders:,} order ({multi_orders:,} multi-SKU); "
          f"{len(pairs):,} pasangan ≥ {BASKET_MIN_PAIR_SUPPORT} order")
    return pairs.head(BASKET_TOP_N), cross, stats


# ---------------------------------------------------------------------------
def _style_header(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 28


def write_basket_report(filepath: Path, pairs: pd.DataFrame, cross: pd.DataFrame,
                        stats: dict, today) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Ringkasan"
    ws["A1"] = "BUNDLE & CROSS-SELL — SKU YANG SERING DIBELI BERSAMA"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:E1")
    pct = (stats["multi_orders"] / stats["orders"] * 100) if stats["orders"] else 0
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  Berdasarkan SKU dalam satu Invoice.  "
                f"Support = jumlah order berisi keduanya; Confidence = peluang beli yang satu "
                f"kalau beli yang lain; Lift > 1 = lebih sering bareng dari kebetulan.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:E2")

    r = 4
    for label, val, fmt in [
        ("Total order (Invoice)", stats["orders"], FMT_NUM),
        (f"Order multi-SKU ({pct:.1f}%)", stats["multi_orders"], FMT_NUM),
        (f"Pasangan SKU (≥ {BASKET_MIN_PAIR_SUPPORT} order)", stats["pairs"], FMT_NUM),
    ]:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = BOLD_FONT
        c.number_format = fmt
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Top kandidat bundle (support tertinggi)").font = TITLE_FONT
    r += 1
    _style_header(ws, r, ["SKU A", "SKU B", "Support", "Conf A→B", "Lift"],
                  [34, 34, 11, 11, 9])
    r += 1
    for _, p in pairs.head(15).iterrows():
        ws.cell(row=r, column=1, value=p["sku_a"]).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=p["sku_b"]).font = NORMAL_FONT
        c = ws.cell(row=r, column=3, value=int(p["support"]))
        c.font = NORMAL_FONT
        c.number_format = FMT_NUM
        c2 = ws.cell(row=r, column=4, value=float(p["conf_ab"]))
        c2.font = NORMAL_FONT
        c2.number_format = FMT_PCT
        c3 = ws.cell(row=r, column=5, value=float(p["lift"]))
        c3.font = NORMAL_FONT
        c3.number_format = FMT_X
        r += 1
    if stats["pairs"] == 0:
        ws.cell(row=r, column=1, value="(belum ada pasangan yang memenuhi ambang support)").font = BOLD_FONT

    # --- Sheet 01: all pairs ---
    ws2 = wb.create_sheet("01_Pasangan_SKU")
    ws2["A1"] = "PASANGAN SKU SERING DIBELI BERSAMA"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:F1")
    _style_header(ws2, 3, ["SKU A", "SKU B", "Support (order)", "Conf A→B", "Conf B→A", "Lift"],
                  [34, 34, 14, 11, 11, 9])
    if len(pairs):
        rr = 4
        for _, p in pairs.iterrows():
            ws2.cell(row=rr, column=1, value=p["sku_a"]).font = NORMAL_FONT
            ws2.cell(row=rr, column=2, value=p["sku_b"]).font = NORMAL_FONT
            c = ws2.cell(row=rr, column=3, value=int(p["support"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=4, value=float(p["conf_ab"])); c.number_format = FMT_PCT; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=5, value=float(p["conf_ba"])); c.number_format = FMT_PCT; c.font = NORMAL_FONT
            c = ws2.cell(row=rr, column=6, value=float(p["lift"])); c.number_format = FMT_X; c.font = NORMAL_FONT
            if rr % 2 == 1:
                for ci in range(1, 7):
                    ws2.cell(row=rr, column=ci).fill = LIGHT_FILL
            rr += 1
        ws2.freeze_panes = "A4"

    # --- Sheet 02: per-SKU best partner (cross-sell prompt) ---
    ws3 = wb.create_sheet("02_Cross_Sell_per_SKU")
    ws3["A1"] = "CROSS-SELL: KALAU BELI INI, TAWARKAN ITU"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:E1")
    _style_header(ws3, 3, ["Kalau beli (SKU)", "Tawarkan (partner)", "Confidence", "Support", "Lift"],
                  [34, 34, 12, 11, 9])
    if len(cross):
        rr = 4
        for _, p in cross.iterrows():
            ws3.cell(row=rr, column=1, value=p["SKU"]).font = NORMAL_FONT
            ws3.cell(row=rr, column=2, value=p["partner"]).font = NORMAL_FONT
            c = ws3.cell(row=rr, column=3, value=float(p["confidence"])); c.number_format = FMT_PCT; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=4, value=int(p["support"])); c.number_format = FMT_NUM; c.font = NORMAL_FONT
            c = ws3.cell(row=rr, column=5, value=float(p["lift"])); c.number_format = FMT_X; c.font = NORMAL_FONT
            if rr % 2 == 1:
                for ci in range(1, 6):
                    ws3.cell(row=rr, column=ci).fill = LIGHT_FILL
            rr += 1
        ws3.freeze_panes = "A4"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
