"""Cash-flow restock plan (`--cashflow`).

Turns the per-SKU reorder metrics into a purchasing budget calendar: for each SKU
that will cross its reorder point within the horizon, it projects

  - WHEN to place the order (from current sisa stok, velocity, and the per-shop
    lead time already computed by the reorder analysis),
  - HOW MUCH to order (the suggested order qty),
  - the COST (qty × replacement HPP = the latest overseas lot price `hpp_pricing`,
    falling back to `hpp_wa`), and
  - WHICH supplier (the SKU's dominant `Toko` by purchase qty).

It then buckets the Rupiah by month and by supplier, so an importer who pays
suppliers upfront (and waits ~2.5 months for sea freight) can see how much
capital is needed and when. v1 plans the NEXT reorder per SKU within the horizon.

Output: output/Analisa_Cashflow_Restock.xlsx. Built entirely from the stok/jual
data — no template needed.
"""
from __future__ import annotations
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    BLUE_FILL_COLOR, CASHFLOW_HORIZON_MONTHS, FMT_NUM, FMT_RP, FONT_NAME,
    HEADER_BG_COLOR, HEADER_TEXT_COLOR, LIGHT_GRAY_COLOR, MIGRASI_PREFIX,
    RED_FILL_COLOR, TITLE_COLOR,
)

_DAYS_PER_MONTH = 365.25 / 12.0

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=HEADER_TEXT_COLOR, size=11)
HEADER_FILL = PatternFill("solid", start_color=HEADER_BG_COLOR)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color=TITLE_COLOR)
BIG_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color=TITLE_COLOR)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
RED_FILL = PatternFill("solid", start_color=RED_FILL_COLOR)
BLUE_FILL = PatternFill("solid", start_color=BLUE_FILL_COLOR)
LIGHT_FILL = PatternFill("solid", start_color=LIGHT_GRAY_COLOR)


def compute_primary_shop(stok: pd.DataFrame) -> pd.Series:
    """Per-SKU dominant supplier = the standardized `Toko` with the largest
    non-Migrasi purchase qty. This is the supplier the cash plan groups spend by."""
    if not {"SKU", "toko", "qty_beli"}.issubset(stok.columns):
        return pd.Series(dtype=object)
    s = stok[~stok["toko"].astype(str).str.startswith(MIGRASI_PREFIX, na=False)].copy()
    s = s[s["qty_beli"] > 0]
    if len(s) == 0:
        return pd.Series(dtype=object)
    qty = s.groupby(["SKU", "toko"])["qty_beli"].sum().reset_index()
    idx = qty.groupby("SKU")["qty_beli"].idxmax()
    return qty.loc[idx].set_index("SKU")["toko"]


def build_restock_plan(reorder_df: pd.DataFrame, hpp_agg: pd.DataFrame,
                       stok: pd.DataFrame, today: pd.Timestamp,
                       horizon_months: int = CASHFLOW_HORIZON_MONTHS) -> pd.DataFrame:
    """One row per planned (next) reorder due within `horizon_months`.

    Order timing = when projected stock crosses `rop_final`:
    `months_to_order = max(0, (sisa − rop_final) / velocity)`. STOCKOUT/below-ROP
    SKUs are due now (month 0). Order qty reuses the reorder formula (target cover
    + lead demand − stock at order time). Cost = qty × `hpp_pricing` (else `hpp_wa`)."""
    if reorder_df is None or len(reorder_df) == 0:
        return pd.DataFrame()

    hpp_pricing = (hpp_agg.set_index("SKU")["hpp_pricing"].to_dict()
                   if "hpp_pricing" in hpp_agg.columns else {})
    hpp_wa = (hpp_agg.set_index("SKU")["hpp_wa"].to_dict()
              if "hpp_wa" in hpp_agg.columns else {})
    primary_shop = compute_primary_shop(stok)

    rows = []
    for _, r in reorder_df.iterrows():
        vel = float(r["velocity_used"])
        if vel <= 0:                       # no demand → nothing to plan
            continue
        sisa = float(r["sisa_stok"])
        rop = float(r["rop_final"])
        daily = vel / _DAYS_PER_MONTH

        if sisa <= rop:                    # already at/below ROP → order now
            days_to_order = 0.0
            stock_at_order = sisa
        else:
            days_to_order = (sisa - rop) / daily if daily > 0 else float("inf")
            stock_at_order = rop
        months_to_order = days_to_order / _DAYS_PER_MONTH
        if months_to_order > horizon_months:   # beyond the plan window — budget later
            continue

        qty = max(0.0, float(r["target_qty_post_reorder"]) - stock_at_order
                  + float(r["lead_demand"]))
        qty = math.ceil(qty)
        if qty <= 0:
            continue

        sku = r["SKU"]
        unit = hpp_pricing.get(sku)
        if unit is None or pd.isna(unit) or unit <= 0:
            unit = hpp_wa.get(sku, np.nan)
        cost = qty * unit if (unit is not None and pd.notna(unit) and unit > 0) else np.nan

        order_date = today + pd.Timedelta(days=float(days_to_order))
        rows.append({
            "SKU": sku,
            "supplier": str(primary_shop.get(sku, "—")) or "—",
            "status": r["status"],
            "sisa_stok": sisa,
            "velocity": vel,
            "lead_months": float(r["lead_months"]),
            "months_to_order": months_to_order,
            "order_date": order_date,
            "order_month": order_date.strftime("%Y-%m"),
            "qty_order": qty,
            "unit_cost": unit,
            "order_cost": cost,
        })

    plan = pd.DataFrame(rows)
    if len(plan):
        plan = plan.sort_values(["order_date", "order_cost"],
                                ascending=[True, False]).reset_index(drop=True)
    return plan


def _months_axis(today: pd.Timestamp, horizon: int) -> list[str]:
    base = today.to_period("M")
    return [str(base + i) for i in range(horizon + 1)]


def summarize_by_month(plan: pd.DataFrame, months: list[str]) -> pd.DataFrame:
    """Total order cost + qty + SKU count per order month, over the full month axis."""
    out = []
    for m in months:
        sub = plan[plan["order_month"] == m] if len(plan) else plan
        out.append({
            "Bulan": m,
            "SKU": int(len(sub)),
            "Qty": int(sub["qty_order"].sum()) if len(sub) else 0,
            "Biaya": float(sub["order_cost"].sum(skipna=True)) if len(sub) else 0.0,
        })
    return pd.DataFrame(out)


def pivot_month_supplier(plan: pd.DataFrame, months: list[str]) -> pd.DataFrame:
    """Order cost matrix: supplier (rows) × order month (cols), with a Total column."""
    if plan is None or len(plan) == 0:
        return pd.DataFrame()
    pv = plan.pivot_table(index="supplier", columns="order_month",
                          values="order_cost", aggfunc="sum", fill_value=0.0)
    for m in months:                       # ensure every month column exists, in order
        if m not in pv.columns:
            pv[m] = 0.0
    pv = pv[months]
    pv["Total"] = pv.sum(axis=1)
    return pv.sort_values("Total", ascending=False)


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------
def _style_header(ws, row, headers, widths):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 26


def write_cashflow_report(filepath: Path, plan: pd.DataFrame, monthly: pd.DataFrame,
                          pivot: pd.DataFrame, months: list[str], today,
                          horizon: int = CASHFLOW_HORIZON_MONTHS) -> None:
    wb = Workbook()

    # --- Sheet 00: Ringkasan ---
    ws = wb.active
    ws.title = "00_Ringkasan"
    ws["A1"] = "RENCANA CASH-FLOW RESTOCK — MODAL BELI YANG DIBUTUHKAN"
    ws["A1"].font = BIG_TITLE_FONT
    ws.merge_cells("A1:D1")
    total = float(plan["order_cost"].sum(skipna=True)) if len(plan) else 0.0
    n_sku = int(len(plan))
    this_month = months[0] if months else today.strftime("%Y-%m")
    due_now = (float(plan[plan["order_month"] == this_month]["order_cost"].sum(skipna=True))
               if len(plan) else 0.0)
    n_no_hpp = int(plan["order_cost"].isna().sum()) if len(plan) else 0
    ws["A2"] = (f"Per {today.strftime('%d %B %Y')}  |  Horizon {horizon} bulan  |  "
                f"Biaya pakai HPP pengganti (lot luar negeri terakhir, fallback HPP_WA). "
                f"v1: merencanakan reorder BERIKUTNYA per SKU dalam horizon.")
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:D2")

    r = 4
    headline = [
        ("Total modal restock (horizon)", total, FMT_RP),
        (f"Jatuh tempo bulan ini ({this_month})", due_now, FMT_RP),
        ("Jumlah SKU yang perlu di-restock", n_sku, FMT_NUM),
    ]
    for label, val, fmt in headline:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = BOLD_FONT
        c.number_format = fmt
        r += 1
    if n_no_hpp:
        ws.cell(row=r, column=1,
                value=f"⚠ {n_no_hpp} SKU tanpa HPP (biaya tak terhitung — lihat Detail)").font = SUB_FONT
        r += 1

    # monthly breakdown table
    r += 1
    ws.cell(row=r, column=1, value="Rincian per bulan").font = TITLE_FONT
    r += 1
    _style_header(ws, r, ["Bulan", "Jumlah SKU", "Qty", "Biaya"], [14, 14, 12, 18])
    hdr_row = r
    r += 1
    for _, m in monthly.iterrows():
        ws.cell(row=r, column=1, value=m["Bulan"]).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=int(m["SKU"])).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=int(m["Qty"])).font = NORMAL_FONT
        ws.cell(row=r, column=3).number_format = FMT_NUM
        c = ws.cell(row=r, column=4, value=float(m["Biaya"]))
        c.font = NORMAL_FONT
        c.number_format = FMT_RP
        if m["Bulan"] == this_month and m["Biaya"] > 0:
            for cc in range(1, 5):
                ws.cell(row=r, column=cc).fill = RED_FILL       # due now = highlight
        r += 1
    # total row
    ws.cell(row=r, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=r, column=3, value=int(monthly["Qty"].sum())).font = BOLD_FONT
    ws.cell(row=r, column=3).number_format = FMT_NUM
    tc = ws.cell(row=r, column=4, value=float(monthly["Biaya"].sum()))
    tc.font = BOLD_FONT
    tc.number_format = FMT_RP
    for cc in range(1, 5):
        ws.cell(row=r, column=cc).fill = BLUE_FILL
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    # --- Sheet 01: Kalender per Bulan (supplier × bulan) ---
    ws2 = wb.create_sheet("01_Kalender_per_Bulan")
    ws2["A1"] = "KALENDER BELANJA — SUPPLIER × BULAN (Rp)"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells(f"A1:{get_column_letter(len(months) + 2)}1")
    if pivot is not None and len(pivot):
        headers = ["Supplier"] + months + ["Total"]
        widths = [22] + [13] * len(months) + [16]
        _style_header(ws2, 3, headers, widths)
        rr = 4
        for supplier, row in pivot.iterrows():
            ws2.cell(row=rr, column=1, value=str(supplier)).font = NORMAL_FONT
            for ci, m in enumerate(months, start=2):
                c = ws2.cell(row=rr, column=ci, value=float(row[m]))
                c.font = NORMAL_FONT
                c.number_format = FMT_RP
            tcell = ws2.cell(row=rr, column=len(months) + 2, value=float(row["Total"]))
            tcell.font = BOLD_FONT
            tcell.number_format = FMT_RP
            if rr % 2 == 1:
                for ci in range(1, len(months) + 1):
                    ws2.cell(row=rr, column=ci).fill = LIGHT_FILL
            rr += 1
        # column totals
        ws2.cell(row=rr, column=1, value="TOTAL").font = BOLD_FONT
        for ci, m in enumerate(months, start=2):
            c = ws2.cell(row=rr, column=ci, value=float(pivot[m].sum()))
            c.font = BOLD_FONT
            c.number_format = FMT_RP
            c.fill = BLUE_FILL
        gc = ws2.cell(row=rr, column=len(months) + 2, value=float(pivot["Total"].sum()))
        gc.font = BOLD_FONT
        gc.number_format = FMT_RP
        gc.fill = BLUE_FILL
        ws2.cell(row=rr, column=1).fill = BLUE_FILL
        ws2.freeze_panes = "B4"
    else:
        ws2["A3"] = "(tidak ada SKU yang perlu di-restock dalam horizon)"
        ws2["A3"].font = BOLD_FONT

    # --- Sheet 02: Detail per SKU ---
    ws3 = wb.create_sheet("02_Detail_per_SKU")
    ws3["A1"] = "DETAIL RENCANA RESTOCK PER SKU"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:K1")
    headers = ["SKU", "Supplier", "Status", "Sisa Stok", "Velocity/bln", "Lead (bln)",
               "Order dalam (bln)", "Tgl Order", "Qty Order", "HPP/pcs", "Total Biaya"]
    widths = [34, 18, 18, 11, 12, 10, 14, 13, 11, 13, 16]
    _style_header(ws3, 3, headers, widths)
    if plan is not None and len(plan):
        rr = 4
        for _, p in plan.iterrows():
            vals = [p["SKU"], p["supplier"], p["status"], round(p["sisa_stok"]),
                    round(p["velocity"], 1), round(p["lead_months"], 2),
                    round(p["months_to_order"], 1),
                    p["order_date"].strftime("%Y-%m-%d"), int(p["qty_order"]),
                    (round(p["unit_cost"]) if pd.notna(p["unit_cost"]) else None),
                    (round(p["order_cost"]) if pd.notna(p["order_cost"]) else None)]
            fmts = [None, None, None, FMT_NUM, FMT_NUM, FMT_NUM, FMT_NUM, None,
                    FMT_NUM, FMT_RP, FMT_RP]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
                c = ws3.cell(row=rr, column=ci, value=v)
                c.font = NORMAL_FONT
                if fmt:
                    c.number_format = fmt
                if rr % 2 == 1:
                    c.fill = LIGHT_FILL
            if p["order_month"] == this_month:                 # due now → red flag
                ws3.cell(row=rr, column=8).fill = RED_FILL
            rr += 1
        ws3.freeze_panes = "A4"
    else:
        ws3["A3"] = "(tidak ada SKU yang perlu di-restock dalam horizon)"
        ws3["A3"].font = BOLD_FONT

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"✓ Menulis laporan ke {filepath}")
